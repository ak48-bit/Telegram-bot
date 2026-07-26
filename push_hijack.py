"""Data & Hijack push — Excel COM screenshots + raw files to Telegram.
  data:   汇总截图 → 当日汇总截图 → 主Excel
  hijack: 当天数据汇总截图 → 劫持办公xlsx
  hr:     DAILY SUMMARY截图 → 劫持人事xlsx+HR文本

  NOTE: This script requires Windows (win32com / Excel COM).
        On Ubuntu it exits cleanly before importing any Windows-only modules.
"""

import os, sys, io, json, urllib.request, tempfile, traceback
from datetime import datetime

# ── Platform guard: MUST be before any win32com import ──
if sys.platform != "win32":
    print("This function is only available on Windows.")
    sys.exit(1)

import pythoncom
from win32com.client import GetActiveObject, Dispatch
import openpyxl

# ── Telegram credentials — unified getter ──
import _platform_config as _plat_hijack
TOKEN, CHAT_ID = _plat_hijack.get_telegram_credentials()
API = f"https://api.telegram.org/bot{TOKEN}"

DATA_FOLDER = r"C:\Users\ak481\OneDrive\Desktop\新建文件夹"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LOG_FILE = os.path.join(SCRIPT_DIR, "bot_log.txt")


def log(msg):
    s = f"[{datetime.now()}] [HijackPush] {msg}"
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(s + "\n")
    except Exception:
        pass
    print(s)


def api_call(method, payload):
    url = f"{API}/{method}"
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json"})
    try:
        resp = urllib.request.urlopen(req, timeout=30)
        return json.loads(resp.read())
    except urllib.error.HTTPError as e:
        log(f"api_call {method} HTTPError: {e.code}")
        return {"ok": False, "description": f"HTTP {e.code}"}
    except Exception as e:
        log(f"api_call {method} error: {e}")
        return {"ok": False, "description": str(e)}


def send_photo(file_path, caption=None):
    """Send a photo to Telegram using multipart form-data."""
    boundary = "----WebKitFormBoundary7MA4YWxkTrZu0gW"
    body = io.BytesIO()
    body.write(f"--{boundary}\r\n".encode())
    body.write(f'Content-Disposition: form-data; name="chat_id"\r\n\r\n'.encode())
    body.write(f"{CHAT_ID}\r\n".encode())
    if caption:
        body.write(f"--{boundary}\r\n".encode())
        body.write(f'Content-Disposition: form-data; name="caption"\r\n\r\n'.encode())
        body.write(f"{caption}\r\n".encode('utf-8'))
    body.write(f"--{boundary}\r\n".encode())
    body.write(f'Content-Disposition: form-data; name="photo"; filename="screenshot.png"\r\n'.encode())
    body.write(f"Content-Type: image/png\r\n\r\n".encode())
    with open(file_path, "rb") as f:
        body.write(f.read())
    body.write(f"\r\n--{boundary}--\r\n".encode())
    body.seek(0)

    url = f"{API}/sendPhoto"
    req = urllib.request.Request(url, data=body.read(),
                                 headers={"Content-Type": f"multipart/form-data; boundary={boundary}"})
    try:
        resp = urllib.request.urlopen(req, timeout=120)
        return json.loads(resp.read())
    except Exception as e:
        log(f"sendPhoto error: {e}")
        return {"ok": False, "description": str(e)}


def send_document(file_path, caption=None):
    """Send a document to Telegram using multipart form-data."""
    fname = os.path.basename(file_path)
    boundary = "----WebKitFormBoundary7MA4YWxkTrZu0gW"
    body = io.BytesIO()
    body.write(f"--{boundary}\r\n".encode())
    body.write(f'Content-Disposition: form-data; name="chat_id"\r\n\r\n'.encode())
    body.write(f"{CHAT_ID}\r\n".encode())
    if caption:
        body.write(f"--{boundary}\r\n".encode())
        body.write(f'Content-Disposition: form-data; name="caption"\r\n\r\n'.encode())
        body.write(f"{caption}\r\n".encode('utf-8'))
    body.write(f"--{boundary}\r\n".encode())
    body.write(f'Content-Disposition: form-data; name="document"; filename="{fname}"\r\n'.encode('utf-8'))
    body.write(f"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n".encode())
    with open(file_path, "rb") as f:
        body.write(f.read())
    body.write(f"\r\n--{boundary}--\r\n".encode())
    body.seek(0)

    url = f"{API}/sendDocument"
    req = urllib.request.Request(url, data=body.read(),
                                 headers={"Content-Type": f"multipart/form-data; boundary={boundary}"})
    try:
        resp = urllib.request.urlopen(req, timeout=120)
        return json.loads(resp.read())
    except Exception as e:
        log(f"sendDocument error: {e}")
        return {"ok": False, "description": str(e)}


def find_main_excel():
    """Find the main ground push Excel file. Returns (workbook_name, file_path)."""
    main_file = None
    for f in sorted(os.listdir(DATA_FOLDER), reverse=True):
        if not f.endswith('.xlsx'):
            continue
        if f.startswith('~$') or '副本' in f or 'Copy' in f:
            continue
        if '线上办公数据汇总' in f and '劫持' not in f and '平衡' not in f and '15天' not in f:
            main_file = f
            break
    if main_file:
        return main_file, os.path.join(DATA_FOLDER, main_file)
    return None, None


def find_hijack_office_excel():
    """Find hijack office Excel file."""
    for f in sorted(os.listdir(DATA_FOLDER), reverse=True):
        if not f.endswith('.xlsx') or f.startswith('~$') or '副本' in f:
            continue
        if '劫持' in f and '办公数据汇总' in f and '人事' not in f:
            return os.path.join(DATA_FOLDER, f)
    return None


def find_hijack_hr_excel():
    """Find hijack HR Excel file."""
    for f in sorted(os.listdir(DATA_FOLDER), reverse=True):
        if not f.endswith('.xlsx') or f.startswith('~$') or '副本' in f:
            continue
        if '劫持' in f and '人事数据汇总' in f:
            return os.path.join(DATA_FOLDER, f)
    return None


def take_sheet_screenshot(excel_path, sheet_name, output_path,
                         max_rows=None, max_cols=None, start_row=1, start_col=1):
    """Copy Excel range as picture, grab from clipboard, save clean PNG.
    start_row: first row to include (1-based).
    start_col: first column to include (1-based). Use 2 to skip empty col A."""
    try:
        import time
        pythoncom.CoInitialize()
        try:
            excel = GetActiveObject('Excel.Application')
        except Exception:
            excel = Dispatch('Excel.Application')
            excel.Visible = False

        # Find or open workbook — match by filename keyword to avoid encoding issues
        target_wb = None
        target_name = os.path.basename(excel_path)
        # Find or open workbook — prefer exact path match, then fall back to keyword match
        target_wb = None
        target_name = os.path.basename(excel_path)

        # First pass: try exact Name match among open workbooks
        for i in range(1, excel.Workbooks.Count + 1):
            try:
                wb = excel.Workbooks(i)
                try:
                    wb_name = str(wb.Name)
                except Exception:
                    continue
                if wb_name == target_name:
                    target_wb = wb
                    break
            except Exception:
                continue

        # Second pass: keyword match with exclusion filters
        if target_wb is None:
            keywords = []
            if '线上办公数据汇总' in target_name and '劫持' not in target_name:
                keywords = ['线上办公数据汇总', '线上办公']
            elif '劫持' in target_name and '办公' in target_name:
                keywords = ['劫持', '办公数据汇总']
            elif '劫持' in target_name and '人事' in target_name:
                keywords = ['劫持', '人事数据汇总']
            else:
                keywords = [target_name]

            for i in range(1, excel.Workbooks.Count + 1):
                try:
                    wb = excel.Workbooks(i)
                    try:
                        wb_name = str(wb.Name)
                    except Exception:
                        continue
                    # Match if ALL keywords appear AND exclude 15天/平衡/副本 variants
                    if all(kw in wb_name for kw in keywords) \
                       and '15天' not in wb_name \
                       and '平衡' not in wb_name \
                       and '副本' not in wb_name:
                        target_wb = wb
                        break
                except Exception:
                    continue

        if target_wb is None:
            try:
                target_wb = excel.Workbooks.Open(excel_path, ReadOnly=True)
            except Exception:
                # Fallback: try opening with short path
                import subprocess, locale
                target_wb = excel.Workbooks.Open(excel_path, ReadOnly=True)

        ws = target_wb.Sheets(sheet_name)
        used = ws.UsedRange
        rows = max_rows if max_rows else used.Rows.Count
        cols = max_cols if max_cols else used.Columns.Count

        # Build range with start_row and start_col offsets
        start_cell = ws.Cells(start_row, start_col)
        end_cell = ws.Cells(start_row + rows - 1, start_col + cols - 1)
        rng = ws.Range(start_cell, end_cell)

        # Copy as picture to clipboard (xlPrinter=1, xlPicture=-4147)
        rng.CopyPicture(1, 2)
        time.sleep(0.3)

        # Grab from clipboard — this is the exact table image, no margins
        from PIL import ImageGrab
        im = ImageGrab.grabclipboard()
        if im is None:
            log(f"Clipboard empty for {sheet_name}")
            return False

        # Crop white borders from the image
        im = _crop_white_borders(im)

        im.save(output_path, 'PNG')
        log(f"Screenshot saved: {sheet_name} (r{start_row}-{start_row+rows-1} x c{start_col}-{start_col+cols-1}) {im.size[0]}x{im.size[1]} -> {os.path.getsize(output_path)} bytes")
        return True

    except Exception as e:
        log(f"Screenshot {sheet_name} failed: {e}")
        traceback.print_exc()
        return False


def _crop_white_borders(im):
    """Crop white/light-gray borders from the image edges."""
    try:
        # Convert to RGB if needed
        if im.mode == 'RGBA':
            im = im.convert('RGB')
        # Find bounding box of non-white pixels
        # White = (255,255,255), allow slight gray up to (240,240,240)
        bg = im.getpixel((0, 0))
        # Use a threshold — anything within 25 of each channel is "background"
        threshold = 25
        def is_bg(pixel):
            return all(abs(int(pixel[i]) - int(bg[i])) < threshold for i in range(3))
        # Scan from edges
        w, h = im.size
        # Top
        top = 0
        for y in range(h):
            if not all(is_bg(im.getpixel((x, y))) for x in range(w)):
                top = y
                break
        # Bottom
        bottom = h - 1
        for y in range(h - 1, -1, -1):
            if not all(is_bg(im.getpixel((x, y))) for x in range(w)):
                bottom = y + 1
                break
        # Left
        left = 0
        for x in range(w):
            if not all(is_bg(im.getpixel((x, y))) for y in range(h)):
                left = x
                break
        # Right
        right = w - 1
        for x in range(w - 1, -1, -1):
            if not all(is_bg(im.getpixel((x, y))) for y in range(h)):
                right = x + 1
                break
        if left < right and top < bottom:
            return im.crop((left, top, right, bottom))
        return im
    except Exception:
        return im


def format_hr_report(hj_hr_path):
    """Read the latest day's HR data and format as text report matching template."""
    wb = openpyxl.load_workbook(hj_hr_path, data_only=True)
    ws = wb['DAILY REPORT']

    # Row 3 has cumulative data: B3=HR name/online count, C3=cumul resumes, D3=cumul interviews
    online_hr_count = ws.cell(row=3, column=2).value or 1
    try:
        online_hr_count = int(online_hr_count)
    except (ValueError, TypeError):
        online_hr_count = 1

    # Find the latest row with data (non-empty column C)
    latest_date = None
    latest_data = {}
    for r in range(5, ws.max_row + 1):
        c_val = ws.cell(row=r, column=3).value
        if c_val is not None and str(c_val).strip():
            try:
                c_val = int(c_val)
            except (ValueError, TypeError):
                pass
            if isinstance(c_val, (int, float)) and c_val > 0:
                date_val = ws.cell(row=r, column=1).value
                if date_val:
                    if isinstance(date_val, datetime):
                        latest_date = date_val.strftime("%Y/%m/%d")
                    else:
                        latest_date = str(date_val)[:10].replace('-', '/')

                def cell_int(col):
                    v = ws.cell(row=r, column=col).value
                    try:
                        return int(v) if v else 0
                    except (ValueError, TypeError):
                        return 0

                latest_data = {
                    "online_hr": online_hr_count,
                    "resumes": cell_int(3),
                    "interviews": cell_int(4),
                    "passed": cell_int(5),
                    "failed": cell_int(6),
                    "training": cell_int(7),
                    "onboarded": cell_int(8),
                    "backout": cell_int(9),
                }

    wb.close()

    if not latest_date:
        now = datetime.now()
        latest_date = now.strftime("%Y/%m/%d")
        latest_data = {
            "online_hr": online_hr_count,
            "resumes": 0, "interviews": 0,
            "passed": 0, "failed": 0, "training": 0,
            "onboarded": 0, "backout": 0,
        }

    lines = [
        f"{latest_date}",
        "=========    ",
        " - 线上办公（劫持）- 线上人事招聘日报      ",
        "   ",
        f"线上人事  :  {latest_data['online_hr']}",
        "   ",
        f"今日简历  :  {latest_data['resumes']}",
        f"今日面试  :  {latest_data['interviews']}",
        f"面试通过  :  {latest_data['passed']}",
        f"面试失败  :  {latest_data['failed']}",
        f"今日培训中  : {latest_data['training']}",
        f"正式上岗  :  {latest_data['onboarded']}",
        f"淘汰  :  {latest_data['backout']}",
    ]

    return "\n".join(lines)


def run_hijack_push():
    """Push hijack office data: 当天数据汇总截图 → 劫持办公xlsx."""
    log("=== Hijack Office push started ===")

    hj_office_path = find_hijack_office_excel()
    log(f"Hijack Office: {hj_office_path}")

    if not hj_office_path:
        log("ERROR: Hijack office Excel not found")
        return False

    tmp_dir = tempfile.gettempdir()
    ss_daily = os.path.join(tmp_dir, "hijack_push_当天数据汇总.png")

    # 1. 当天数据汇总 (rows 2-4, cols 1-32) → 2635x160
    ok1 = take_sheet_screenshot(hj_office_path, "当天数据汇总", ss_daily,
                                max_rows=3, max_cols=32, start_row=2)
    if ok1:
        result = send_document(ss_daily)
        log(f"Hijack Img 1 (当天数据汇总): {'OK' if result.get('ok') else result.get('description')}")
    else:
        log("WARNING: 当天数据汇总 screenshot failed")

    # 2. Hijack office xlsx
    if hj_office_path and os.path.exists(hj_office_path):
        result = send_document(hj_office_path)
        log(f"Hijack Doc 1 (劫持办公): {'OK' if result.get('ok') else result.get('description')}")
    else:
        log("WARNING: Hijack office file not found")

    # Clean up
    try:
        if os.path.exists(ss_daily):
            os.remove(ss_daily)
    except Exception:
        pass

    log("=== Hijack Office push completed ===")
    return True


def run_hr_push():
    """Push hijack HR data: DAILY SUMMARY截图 → 劫持人事xlsx+HR文本."""
    log("=== Hijack HR push started ===")

    hj_hr_path = find_hijack_hr_excel()
    log(f"Hijack HR: {hj_hr_path}")

    if not hj_hr_path:
        log("ERROR: Hijack HR Excel not found")
        return False

    tmp_dir = tempfile.gettempdir()
    ss_summary = os.path.join(tmp_dir, "hijack_push_DAILY_SUMMARY.png")

    # 1. DAILY SUMMARY (rows 1-24, cols 1-9) → 1549x852
    ok1 = take_sheet_screenshot(hj_hr_path, "DAILY SUMMARY", ss_summary,
                                max_rows=24, max_cols=9)
    if ok1:
        result = send_document(ss_summary)
        log(f"HR Img 1 (DAILY SUMMARY): {'OK' if result.get('ok') else result.get('description')}")
    else:
        log("WARNING: DAILY SUMMARY screenshot failed")

    # 2. Hijack HR xlsx + HR text caption
    hr_text = ""
    if hj_hr_path and os.path.exists(hj_hr_path):
        try:
            hr_text = format_hr_report(hj_hr_path)
        except Exception as e:
            log(f"HR report format error: {e}")
        result = send_document(hj_hr_path, caption=hr_text)
        log(f"HR Doc 1 (劫持人事): {'OK' if result.get('ok') else result.get('description')}")
    else:
        log("WARNING: Hijack HR file not found")

    # Clean up
    try:
        if os.path.exists(ss_summary):
            os.remove(ss_summary)
    except Exception:
        pass

    log("=== Hijack HR push completed ===")
    return True


def run_data_push():
    """Push data summary: 汇总截图 + 当日汇总截图 + 主Excel文件.
    Matches user's template format exactly."""
    log("=== Data push started ===")

    # ── Find main Excel ──
    main_name, main_path = find_main_excel()
    log(f"Main: {main_name}")

    if not main_path:
        log("ERROR: Main Excel not found")
        return False

    tmp_dir = tempfile.gettempdir()
    ss_huizong = os.path.join(tmp_dir, "data_push_汇总.png")
    ss_daily = os.path.join(tmp_dir, "data_push_当日汇总.png")

    # ── 1. 汇总 sheet screenshot ──
    ok1 = take_sheet_screenshot(main_path, "汇总", ss_huizong, max_rows=37, max_cols=32, start_col=2)
    if ok1:
        result = send_document(ss_huizong)
        log(f"Img 1 (汇总): {'OK' if result.get('ok') else result.get('description')}")
    else:
        log("WARNING: 汇总 screenshot failed")

    # ── 2. 当日汇总 sheet screenshot ──
    ok2 = take_sheet_screenshot(main_path, "当日汇总", ss_daily, max_rows=24, max_cols=28)
    if ok2:
        result = send_document(ss_daily)
        log(f"Img 2 (当日汇总): {'OK' if result.get('ok') else result.get('description')}")
    else:
        log("WARNING: 当日汇总 screenshot failed")

    # ── 3. Main Excel file ──
    if main_path and os.path.exists(main_path):
        result = send_document(main_path)
        log(f"Doc 1 (主数据文件): {'OK' if result.get('ok') else result.get('description')}")
    else:
        log("WARNING: Main file not found")

    # ── Clean up temp files ──
    for tmp in [ss_huizong, ss_daily]:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass

    log("=== Data push completed ===")
    return True


if __name__ == "__main__":
    mode = sys.argv[1] if len(sys.argv) > 1 else "data"
    if mode == "hijack":
        run_hijack_push()
    elif mode == "hr":
        run_hr_push()
    else:
        run_data_push()
