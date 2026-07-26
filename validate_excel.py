"""Server-side Excel validation — checks file integrity without sending Telegram."""
import os, sys, json

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

# Resolve data_folder and active_excel_file
cfg_path = os.path.join(SCRIPT_DIR, "config.json")
try:
    with open(cfg_path, "r", encoding="utf-8") as f:
        cfg = json.load(f)
except Exception as e:
    print(f"ERROR: Cannot read config.json: {e}")
    sys.exit(1)

data_folder = os.environ.get("DATA_FOLDER", "").strip()
if not data_folder:
    data_folder = cfg.get("data_folder", "")

active_file = cfg.get("active_excel_file", "")

if not data_folder:
    print("ERROR: data_folder not configured")
    sys.exit(1)
if not active_file:
    print("ERROR: active_excel_file not configured")
    sys.exit(1)

excel_path = os.path.join(data_folder, active_file)

errors = 0

# 1. File exists
if not os.path.isfile(excel_path):
    print(f"ERROR: File not found: {excel_path}")
    sys.exit(1)
print(f"OK: File exists: {excel_path}")

# 2. File size > 0
size = os.path.getsize(excel_path)
if size == 0:
    print(f"ERROR: File is empty (0 bytes)")
    sys.exit(1)
print(f"OK: File size: {size} bytes")

# 3. openpyxl can open
try:
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheets = wb.sheetnames
    print(f"OK: openpyxl opened, {len(sheets)} sheets")
except Exception as e:
    print(f"ERROR: Cannot open Excel: {e}")
    sys.exit(1)

# 4. Required sheets for dev platforms
try:
    import _platform_config as p
    dev = p.get_development_platforms()
except Exception as e:
    print(f"ERROR: Cannot load platform config: {e}")
    wb.close()
    sys.exit(1)

for code in dev:
    if code in sheets:
        print(f"OK: Sheet '{code}' exists")
    else:
        print(f"ERROR: Sheet '{code}' MISSING")
        errors += 1

# 5. Daily sheet check
daily_sn = None
for sn in sheets:
    if sn == "当日汇总":
        daily_sn = sn
        break
if not daily_sn:
    for sn in sheets:
        if "汇总" in sn and sn != "汇总":
            try:
                hdr = wb[sn].cell(row=4, column=1).value
                if hdr and "DATE" in str(hdr).upper():
                    daily_sn = sn
                    break
            except Exception:
                continue

if daily_sn:
    print(f"OK: Daily sheet found: '{daily_sn}'")
    ws = wb[daily_sn]
    rows = cfg.get("platform_rows", {})
    for code in dev:
        dr = rows.get(code, {}).get("daily_row")
        if dr:
            cell = str(ws.cell(row=dr, column=1).value or "")
            if code in cell:
                ftd = ws.cell(row=dr, column=7).value
                print(f"OK: {code} daily_row={dr} FTD={ftd}")
            else:
                print(f"ERROR: {code} daily_row={dr} content='{cell[:20]}' mismatch")
                errors += 1
else:
    print("ERROR: Daily summary sheet not found")
    errors += 1

# 6. Monthly sheet check
ws0 = wb[sheets[0]]
for code in dev:
    mr = rows.get(code, {}).get("monthly_row")
    if mr:
        cell = str(ws0.cell(row=mr, column=3).value or "")
        if code in cell:
            print(f"OK: {code} monthly_row={mr}")
        else:
            print(f"WARNING: {code} monthly_row={mr} content='{cell[:20]}'")
    else:
        print(f"WARNING: {code} missing monthly_row config")

wb.close()

print()
if errors == 0:
    print("VALIDATION PASSED")
    sys.exit(0)
else:
    print(f"VALIDATION FAILED ({errors} errors)")
    sys.exit(1)
