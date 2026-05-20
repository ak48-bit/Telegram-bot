import openpyxl

EXCEL = r"C:\Users\ak481\OneDrive\Desktop\新建文件夹\26年05月 线上办公数据汇总.xlsx"
TARGET_DATE = "2026-05-02"
SITES = ["PH09", "PH09-2", "PH25", "PH18", "PH30", "PH05", "PH16", "BD02", "BD05"]

wb = openpyxl.load_workbook(EXCEL, data_only=True)

print("=== 2026-05-02 各站点 首存人数 & 新客单价 ===")
print()
print(f"{'站点':<10} {'首存人数':>8} {'首存金额':>12} {'新客单价':>10} {'充值人数':>8} {'总充值':>12} {'总提款':>12} {'投产比':>8}")
print("-" * 85)

total_ftd = 0
total_ftd_amt = 0.0
total_dps = 0
total_dps_amt = 0.0
total_wdr = 0.0

for name in SITES:
    if name not in wb.sheetnames:
        print(f"{name:<10} {'SHEET NOT FOUND'}")
        continue

    ws = wb[name]
    target_row = None
    for row_idx in range(ws.max_row, 5, -1):
        date_val = ws.cell(row=row_idx, column=1).value
        if date_val is not None and hasattr(date_val, 'strftime') and date_val.strftime('%Y-%m-%d') == TARGET_DATE:
            ftd_val = ws.cell(row=row_idx, column=8).value
            if ftd_val is not None:
                target_row = row_idx
                break

    if target_row is None:
        print(f"{name:<10} {'NO DATA FOR 2026-05-02'}")
        continue

    def c(col):
        v = ws.cell(row=target_row, column=col).value
        if v is None:
            return 0
        try:
            return float(v)
        except (ValueError, TypeError):
            return 0

    ftd = int(c(8))
    ftd_amt = c(20)
    new_cust_avg = c(15)
    dps_ppl = int(c(11))
    total_dps_amt_v = c(23)
    total_wdr_amt_v = c(24)
    roi = c(34)

    total_ftd += ftd
    total_ftd_amt += ftd_amt
    total_dps += dps_ppl
    total_dps_amt += total_dps_amt_v
    total_wdr += total_wdr_amt_v

    roi_str = f"{roi:.1f}" if roi != 0 else "N/A"
    print(f"{name:<10} {ftd:>8} {ftd_amt:>12,.0f} {new_cust_avg:>10,.0f} {dps_ppl:>8} {total_dps_amt_v:>12,.0f} {total_wdr_amt_v:>12,.0f} {roi_str:>8}")

print("-" * 85)
avg_price = total_ftd_amt / total_ftd if total_ftd > 0 else 0
print(f"{'合计':<10} {total_ftd:>8} {total_ftd_amt:>12,.0f} {avg_price:>10,.0f} {total_dps:>8} {total_dps_amt:>12,.0f} {total_wdr:>12,.0f}")
print()

# Per-region summary
ph_ftd = 0; ph_ftd_amt = 0.0
bd_ftd = 0; bd_ftd_amt = 0.0
for name in SITES:
    if name not in wb.sheetnames:
        continue
    ws = wb[name]
    for row_idx in range(ws.max_row, 5, -1):
        date_val = ws.cell(row=row_idx, column=1).value
        if date_val is not None and hasattr(date_val, 'strftime') and date_val.strftime('%Y-%m-%d') == TARGET_DATE:
            f = float(ws.cell(row=row_idx, column=8).value or 0)
            a = float(ws.cell(row=row_idx, column=20).value or 0)
            if name.startswith("PH"):
                ph_ftd += int(f); ph_ftd_amt += a
            elif name.startswith("BD"):
                bd_ftd += int(f); bd_ftd_amt += a
            break

print("=== 区域汇总 ===")
print(f"菲律宾: 首存{ph_ftd}人, 首存金额{ph_ftd_amt:,.0f}, 均价{ph_ftd_amt/ph_ftd if ph_ftd>0 else 0:,.0f}")
print(f"孟加拉: 首存{bd_ftd}人, 首存金额{bd_ftd_amt:,.0f}, 均价{bd_ftd_amt/bd_ftd if bd_ftd>0 else 0:,.0f}")
print(f"总合计: 首存{total_ftd}人, 首存金额{total_ftd_amt:,.0f}, 均价{avg_price:,.0f}")

wb.close()
