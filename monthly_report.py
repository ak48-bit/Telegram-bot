import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

MARCH_FILE = r'C:\Users\ak481\OneDrive\Desktop\新建文件夹\26年03月01-31 线上办公数据汇总.xlsx'
APRIL_FILE = r'C:\Users\ak481\OneDrive\Desktop\新建文件夹\26年04月 线上办公数据汇.xlsx'
PLATFORMS = ['PH09', 'PH09-2', 'PH25', 'PH18', 'PH05', 'PH16', 'BD02', 'BD05']

def get_monthly(filepath, target_month):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {}
    for platform in PLATFORMS:
        if platform not in wb.sheetnames:
            result[platform] = None
            continue
        ws = wb[platform]
        m = {'reg': 0, 'ftd': 0, 'dps_ppl': 0, 'ftd_amt': 0.0, 'dps_amt': 0.0, 'wdr_amt': 0.0, 'diff': 0.0}
        found = False
        for row_idx in range(6, ws.max_row + 1):
            date_val = ws.cell(row=row_idx, column=1).value
            if date_val is None or not hasattr(date_val, 'strftime'):
                continue
            if date_val.strftime('%Y-%m') != target_month:
                continue
            ftd_val = ws.cell(row=row_idx, column=8).value
            if ftd_val is None:
                continue
            try:
                float(ftd_val)
            except (ValueError, TypeError):
                continue
            found = True
            def cell(c):
                v = ws.cell(row=row_idx, column=c).value
                if v is None:
                    return 0.0
                try:
                    return float(v)
                except (ValueError, TypeError):
                    return 0.0
            m['reg'] += int(cell(7))
            m['ftd'] += int(cell(8))
            m['dps_ppl'] += int(cell(11))
            m['ftd_amt'] += cell(20)
            m['dps_amt'] += cell(23)
            m['wdr_amt'] += cell(24)
            m['diff'] += cell(25)
        result[platform] = m if found else None
    wb.close()
    return result

def fmt_k(v):
    if abs(v) >= 1_000_000:
        return f'{v/1_000_000:.2f}M'
    if abs(v) >= 1_000:
        return f'{v/1_000:.0f}K'
    return f'{v:,.0f}'

mar_data = get_monthly(MARCH_FILE, '2026-03')
apr_data = get_monthly(APRIL_FILE, '2026-04')

print()
print('=' * 90)
print('  各平台 3月 vs 4月 月度数据汇总')
print('=' * 90)

for platform in PLATFORMS:
    print(f'\n--- {platform} ---')
    print(f'{"指标":<12} {"3月":>12} {"4月":>12} {"变化":>12} {"变化率":>10}')
    print('-' * 58)
    mar = mar_data.get(platform)
    apr = apr_data.get(platform)
    metrics = [
        ('总注册', 'reg'), ('总首存', 'ftd'), ('总充值', 'dps_ppl'),
        ('首存金额', 'ftd_amt'), ('总充值', 'dps_amt'), ('总提款', 'wdr_amt'),
        ('存提差', 'diff'),
    ]
    for name, key in metrics:
        v3 = mar[key] if mar else 0
        v4 = apr[key] if apr else 0
        chg = v4 - v3
        pct = (chg / abs(v3) * 100) if v3 != 0 else 0
        print(f'{name:<12} {fmt_k(v3):>12} {fmt_k(v4):>12} {fmt_k(chg):>12} {pct:>+9.1f}%')

# Grand totals
print()
print('=' * 90)
print('  8站合计汇总')
print('=' * 90)
print(f'{"指标":<12} {"3月":>12} {"4月":>12} {"变化":>12} {"变化率":>10}')
print('-' * 58)

grand_mar = {'reg': 0, 'ftd': 0, 'dps_ppl': 0, 'ftd_amt': 0.0, 'dps_amt': 0.0, 'wdr_amt': 0.0, 'diff': 0.0}
grand_apr = {'reg': 0, 'ftd': 0, 'dps_ppl': 0, 'ftd_amt': 0.0, 'dps_amt': 0.0, 'wdr_amt': 0.0, 'diff': 0.0}

for platform in PLATFORMS:
    for key in grand_mar:
        mar = mar_data.get(platform)
        apr = apr_data.get(platform)
        if mar:
            grand_mar[key] += mar[key]
        if apr:
            grand_apr[key] += apr[key]

for name, key in metrics:
    v3 = grand_mar[key]
    v4 = grand_apr[key]
    chg = v4 - v3
    pct = (chg / abs(v3) * 100) if v3 != 0 else 0
    print(f'{name:<12} {fmt_k(v3):>12} {fmt_k(v4):>12} {fmt_k(chg):>12} {pct:>+9.1f}%')

print()
