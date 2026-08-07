"""
Platform Configuration — Single Source of Truth
================================================
All platform identities and row mappings are read from config.json ONLY.

Three independent categories (flat lists in config.json):
  development_platforms  → daily push, monthly, FTD, English
  hijack_platforms       → hijack office / HR push
  disabled_platforms     → permanently disabled

Row mappings live separately in platform_rows.
Region is derived from the platform code prefix (PH / BD / MM).

No Python file may hardcode any platform name.
"""

import os
import json
from datetime import datetime

try:
    import _runtime
except ImportError:
    _runtime = None

# ══════════════════════════════════════════════════════════════════════
#  Constants
# ══════════════════════════════════════════════════════════════════════

CONFIG_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "config.json"
)

REGION_PREFIXES = ("PH", "BD", "MM")


# ══════════════════════════════════════════════════════════════════════
#  Telegram credentials (unified — call only when actually sending)
# ══════════════════════════════════════════════════════════════════════

def _load_wfhdp_env():
    """Load .wfhdp.env from script directory into os.environ (if file exists).
    Does NOT override existing env vars. Simple KEY=VALUE parser.
    Returns list of warnings (empty if all OK)."""
    warnings = []
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".wfhdp.env")
    if not os.path.isfile(env_path):
        return warnings
    try:
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, _, value = line.partition("=")
                key = key.strip()
                value = value.strip().strip('"').strip("'")
                if key and key not in os.environ:
                    os.environ[key] = value
    except PermissionError:
        warnings.append(f"Cannot read .wfhdp.env: permission denied")
    except Exception as e:
        warnings.append(f"Cannot read .wfhdp.env: {e}")
    return warnings


def get_telegram_credentials():
    """Return (token, chat_id) from env or .wfhdp.env.
    Priority: OS environment > .wfhdp.env file.
    Raises RuntimeError if missing.
    """
    _load_wfhdp_env()
    token = os.environ.get("TELEGRAM_BOT_TOKEN", "").strip()
    chat_id = os.environ.get("TELEGRAM_CHAT_ID", "").strip()
    if not token:
        raise RuntimeError("TELEGRAM_BOT_TOKEN is not configured")
    if not chat_id:
        raise RuntimeError("TELEGRAM_CHAT_ID is not configured")
    return token, int(chat_id)


# ══════════════════════════════════════════════════════════════════════
#  Cache / loading
# ══════════════════════════════════════════════════════════════════════

_cfg = None


def _clear_cache():
    """Clear internal caches (for hot reload)."""
    global _cfg
    _cfg = None


def _load():
    """Lazy-load config.json into _cfg."""
    global _cfg
    if _cfg is not None:
        return
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            _cfg = json.load(f)
    else:
        _cfg = {}


def get_config():
    """Return the full config dict (live reference — call after reload)."""
    _load()
    return _cfg


# ══════════════════════════════════════════════════════════════════════
#  Version helpers
# ══════════════════════════════════════════════════════════════════════

def get_config_version():
    """Return (version, updated_at, config_path)."""
    _load()
    v = _cfg.get("config_version")
    u = _cfg.get("updated_at", "unknown")
    return v, u, CONFIG_FILE


# ══════════════════════════════════════════════════════════════════════
#  Region derivation
# ══════════════════════════════════════════════════════════════════════

def get_region(code):
    """Derive region from platform code prefix: PH09→'PH', BD05→'BD'."""
    if not code:
        return None
    upper = code.upper()
    for prefix in REGION_PREFIXES:
        if upper.startswith(prefix):
            return prefix
    return None


# ══════════════════════════════════════════════════════════════════════
#  Platform identity (reads from live _cfg)
# ══════════════════════════════════════════════════════════════════════

def _list(name):
    _load()
    return _cfg.get(name, [])


def get_development_platforms():
    return list(_list("development_platforms"))


def get_hijack_platforms():
    return list(_list("hijack_platforms"))


def get_disabled_platforms():
    return list(_list("disabled_platforms"))


def get_all_configured_platforms():
    _load()
    return list(_cfg.get("platform_rows", {}).keys())


def get_platforms_by_region(region):
    return [p for p in get_development_platforms() if get_region(p) == region]


def get_admin_ids():
    """Return admin Telegram IDs. Env var ADMIN_TELEGRAM_IDS takes priority
    (comma-separated), falls back to config.json admin_telegram_ids."""
    env_val = os.environ.get("ADMIN_TELEGRAM_IDS", "").strip()
    if env_val:
        return [x.strip() for x in env_val.split(",") if x.strip()]
    _load()
    return _cfg.get("admin_telegram_ids", [])


def _get_idle_platforms():
    """Platforms in platform_rows but not in dev/hijack/disabled."""
    all_configured = set(get_all_configured_platforms())
    categorized = set(get_development_platforms() + get_hijack_platforms() + get_disabled_platforms())
    return sorted(all_configured - categorized)


# ══════════════════════════════════════════════════════════════════════
#  Row helpers
# ══════════════════════════════════════════════════════════════════════

def _rows():
    _load()
    return _cfg.get("platform_rows", {})


def get_daily_row_map():
    return {code: info.get("daily_row") for code, info in _rows().items()}


def get_monthly_row_map():
    return {code: info.get("monthly_row") for code, info in _rows().items()}


def get_daily_rows(region=None):
    all_rows = _rows()
    if region:
        codes = [c for c in all_rows if get_region(c) == region]
    else:
        codes = list(all_rows.keys())
    return sorted(all_rows[c].get("daily_row") for c in codes if all_rows[c].get("daily_row") is not None)


def get_monthly_rows(region=None):
    all_rows = _rows()
    if region:
        codes = [c for c in all_rows if get_region(c) == region]
    else:
        codes = list(all_rows.keys())
    return sorted(all_rows[c].get("monthly_row") for c in codes if all_rows[c].get("monthly_row") is not None)


# ══════════════════════════════════════════════════════════════════════
#  Excel file resolution (UNIFIED — used by all checks and push)
# ══════════════════════════════════════════════════════════════════════

# active_month.json — unified month data source
# Priority: Railway env vars → /data/active_month.json → git active_month.json
ACTIVE_MONTH_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "active_month.json"
)


def _resolve_active_month_path():
    """Find active_month.json. Priority: /data/active_month.json → git root."""
    try:
        import _runtime
        data_root = _runtime.resolve_data_root()
        data_path = os.path.join(data_root, "active_month.json")
        if os.path.isfile(data_path):
            return data_path
    except Exception:
        pass
    return ACTIVE_MONTH_FILE


def _load_active_month():
    """Read active_month config.
    Priority: env vars → /data/active_month.json → git active_month.json.
    Returns (dict, None) on success.
    Returns (None, error_msg) if a config file exists but is malformed.
    Returns (None, None) if no config (V1.0 compatibility → fallback)."""
    # 1. Railway env vars — MUST be configured in pairs, else fail
    env_dev = os.environ.get("ACTIVE_DEVELOPMENT_FILE", "").strip()
    env_hij = os.environ.get("ACTIVE_HIJACK_FILE", "").strip()
    if env_dev or env_hij:
        if not env_dev or not env_hij:
            missing = "hijack" if not env_hij else "development"
            return None, (f"ACTIVE_*_FILE 环境变量必须成对配置，缺少 {missing}")
        return {"development": env_dev, "hijack": env_hij}, None

    # 2. /data/active_month.json then git active_month.json
    path = _resolve_active_month_path()
    if not os.path.isfile(path):
        return None, None
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data, None
    except json.JSONDecodeError as e:
        return None, f"active_month.json JSON 格式错误: {str(e)[:60]}"
    except OSError as e:
        return None, f"active_month.json 读取失败: {str(e)[:60]}"


def _resolve_data_folder():
    """Resolve DATA_FOLDER: env var → config.json → platform default.
    Delegates to _runtime for Railway/Linux handling."""
    try:
        import _runtime
        return _runtime.excel_dir()
    except ImportError:
        pass
    env_val = os.environ.get("DATA_FOLDER", "").strip()
    if env_val:
        return env_val
    _load()
    cfg_folder = _cfg.get("data_folder", "").strip()
    if cfg_folder:
        return cfg_folder
    return os.path.dirname(os.path.abspath(__file__))


def _scan_latest(data_folder, keyword, exclude_kw=None):
    """Scan for most recently modified xlsx matching keyword. Returns path or None."""
    if not data_folder or not os.path.isdir(data_folder):
        return None
    best, best_mtime = None, 0
    for f in os.listdir(data_folder):
        if not f.endswith('.xlsx') or f.startswith('~$'):
            continue
        if keyword not in f:
            continue
        if exclude_kw and exclude_kw in f:
            continue
        path = os.path.join(data_folder, f)
        mtime = os.path.getmtime(path)
        if mtime > best_mtime:
            best_mtime = mtime
            best = path
    return best


def _detect_data_date(path):
    """Detect data cutoff date from an Excel file.
    Priority: datetime cells → title month + platform sheet latest day.
    Returns 'YYYY-MM-DD' or None."""
    import re
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]

        # 1. Direct datetime cells in rows 1-5
        for r in range(1, 6):
            for c in range(1, 20):
                v = ws.cell(row=r, column=c).value
                if v and hasattr(v, 'strftime'):
                    wb.close()
                    return v.strftime('%Y-%m-%d')

        # 2. Title month + platform sheet latest day
        # Year priority: filename "26年" → 2026; fallback current year
        year = None
        m_year = re.search(r'(?:^|\D)(\d{2})年', os.path.basename(path))
        if m_year:
            year = 2000 + int(m_year.group(1))
        if not year:
            year = datetime.now().year
        month = None
        for r in range(1, 4):
            for c in range(1, 8):
                v = str(ws.cell(row=r, column=c).value or "")
                m = re.search(r'(\d{1,2})\s*月', v)
                if m:
                    month = int(m.group(1))
                    break
            if month:
                break

        latest_day = 0
        for sn in wb.sheetnames:
            if not re.match(r'^(PH|BD|MM)\d', sn):
                continue
            try:
                pws = wb[sn]
                for row_idx in range(pws.max_row, 5, -1):
                    d = pws.cell(row=row_idx, column=1).value
                    if not d or not hasattr(d, 'strftime'):
                        continue
                    if month and d.month != month:
                        continue
                    # Prefer sheet's real year over filename/current
                    if year is None:
                        year = d.year
                    # Check multiple business fields non-zero
                    fields = [pws.cell(row=row_idx, column=c).value for c in (7, 8, 20, 11, 23, 24)]
                    has_data = False
                    for fv in fields:
                        try:
                            if fv is not None and float(fv) != 0:
                                has_data = True
                                break
                        except (ValueError, TypeError):
                            pass
                    if has_data and d.day > latest_day:
                        latest_day = d.day
            except Exception:
                continue
        wb.close()
        if latest_day > 0 and month:
            return f"{year}-{month:02d}-{latest_day:02d}"
        return None
    except Exception:
        return None


def _verify_excel(path, kind, data_folder):
    """Validate an Excel file for the given kind.
    kind: 'development' | 'hijack'
    Returns (ok, errors_list)."""
    errors = []
    fname = os.path.basename(path)
    if not os.path.isfile(path):
        return False, [f"文件不存在: {fname}"]

    # Development must not be a hijack file
    if kind == "development" and "劫持" in fname:
        return False, [f"development 误指向劫持文件: {fname}"]
    # Hijack must contain 劫持
    if kind == "hijack" and "劫持" not in fname:
        return False, [f"hijack 文件名称缺少「劫持」: {fname}"]

    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        wb.close()
    except Exception as e:
        return False, [f"Excel 无法读取: {fname} ({str(e)[:60]})"]

    # data_date detection
    date_found = _detect_data_date(path)
    if not date_found:
        return False, [f"data_date 无法识别: {fname}"]

    return True, errors


def get_active_excel(kind):
    """Resolve the active Excel file for kind ('development'|'hijack').
    Priority: active_month.json → config.json → auto-scan.
    Returns (path, fname, data_date, errors) or (None, None, None, errors).
    """
    errors = []
    data_folder = _resolve_data_folder()

    candidate = None
    am_error = None
    # 1. active_month.json
    am, am_error = _load_active_month()
    if am_error:
        # File exists but malformed → fail clearly, no silent fallback
        return None, None, None, [am_error]
    if am is not None:
        # active_month.json EXISTS → all fields required, no silent fallback
        missing = []
        if "version" not in am:
            missing.append("version")
        for field in ("development", "hijack"):
            val = am.get(field)
            if not val or not str(val).strip():
                missing.append(field)
        if missing:
            return None, None, None, [
                f"active_month.json 缺少必填字段: {', '.join(missing)}"]
        candidate = am[kind]
    # 2. config.json active_excel_file (development only) — only if active_month absent
    if am is None and candidate is None and kind == "development":
        _load()
        candidate = _cfg.get("active_excel_file", "").strip()

    resolved_path = None
    fname = None
    if candidate:
        fname = candidate
        resolved_path = os.path.join(data_folder, candidate)
    else:
        # 3. auto-scan fallback
        if kind == "development":
            resolved_path = _scan_latest(data_folder, "线上办公数据汇总", "劫持")
        else:
            resolved_path = _scan_latest(data_folder, "劫持（线上办公数据汇总）")
        if resolved_path:
            fname = os.path.basename(resolved_path)

    if not resolved_path:
        return None, None, None, [f"{kind}: 未找到活动 Excel 文件"]

    ok, verr = _verify_excel(resolved_path, kind, data_folder)
    if not ok:
        return None, None, None, verr

    # Extract data_date (robust detection)
    data_date = _detect_data_date(resolved_path)

    return resolved_path, fname, data_date, errors


def get_active_excel_verbose(kind):
    """Detailed active Excel info for status output."""
    path, fname, data_date, errors = get_active_excel(kind)
    if not path:
        return {"success": False, "kind": kind, "errors": errors}
    return {
        "success": True, "kind": kind, "path": path, "fname": fname,
        "data_date": data_date, "size": os.path.getsize(path) if os.path.isfile(path) else 0,
    }


def find_main_excel(data_folder):
    """Return the active Excel file path as configured in config.json.
    Uses active_excel_file if set, otherwise falls back to searching.
    Returns (path, filename) or (None, None).
    """
    if not data_folder or not os.path.isdir(data_folder):
        return None, None

    _load()
    active_file = _cfg.get("active_excel_file", "").strip()

    if active_file:
        path = os.path.join(data_folder, active_file)
        if os.path.isfile(path):
            return path, active_file
        else:
            return None, None  # configured file missing → fail

    # No active_excel_file configured: search for current month files
    now = datetime.now()
    year_2d = now.strftime("%y")
    patterns = [f"{year_2d}年{now.strftime('%m')}月", f"{year_2d}年{now.month}月"]

    found = []
    for f in sorted(os.listdir(data_folder)):
        if not f.endswith('.xlsx'):
            continue
        if f.startswith('~$') or '副本' in f or ' - Copy' in f:
            continue
        if '线上办公数据汇总' not in f or '劫持' in f:
            continue
        for pat in patterns:
            if pat in f:
                found.append(f)
                break

    if len(found) == 0:
        return None, None
    if len(found) == 1:
        return os.path.join(data_folder, found[0]), found[0]
    # Multiple candidates and no active_excel_file → fail (must configure)
    return None, None


def get_other_monthly_files(data_folder):
    """Return list of other same-month Excel files NOT being used.
    For INFO display in startup_check / /platforms.
    """
    if not data_folder or not os.path.isdir(data_folder):
        return []

    _load()
    active_file = _cfg.get("active_excel_file", "").strip()

    now = datetime.now()
    year_2d = now.strftime("%y")
    patterns = [f"{year_2d}年{now.strftime('%m')}月", f"{year_2d}年{now.month}月"]

    all_found = []
    for f in sorted(os.listdir(data_folder)):
        if not f.endswith('.xlsx'):
            continue
        if f.startswith('~$') or '副本' in f or ' - Copy' in f:
            continue
        if '线上办公数据汇总' not in f or '劫持' in f:
            continue
        for pat in patterns:
            if pat in f:
                all_found.append(f)
                break

    if active_file:
        return [f for f in all_found if f != active_file]
    return all_found


def _find_daily_sheet(wb):
    """Locate 当日汇总 sheet in a workbook. Returns sheet name or None."""
    sheets = wb.sheetnames
    for sn in sheets:
        if sn == "当日汇总":
            return sn
    for sn in sheets:
        if "汇总" in sn and sn != "汇总":
            try:
                hdr = wb[sn].cell(row=4, column=1).value
                if hdr and "DATE" in str(hdr).upper():
                    return sn
            except Exception:
                continue
    if len(sheets) > 3:
        return sheets[3]
    return None


# ══════════════════════════════════════════════════════════════════════
#  Value classification (None vs 0 vs empty)
# ══════════════════════════════════════════════════════════════════════

def _is_truly_empty(v):
    """None or empty string → True.  0, 0.0, False → False."""
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def _safe_read(ws, row, col):
    try:
        return ws.cell(row=row, column=col).value
    except Exception:
        return None


# ══════════════════════════════════════════════════════════════════════
#  Platform classification
# ══════════════════════════════════════════════════════════════════════

def _classify(code):
    upper = code.upper() if code else ""
    if upper in (p.upper() for p in get_development_platforms()):
        return "开发"
    if upper in (p.upper() for p in get_hijack_platforms()):
        return "劫持"
    if upper in (p.upper() for p in get_disabled_platforms()):
        return "停用"
    if upper in (p.upper() for p in get_all_configured_platforms()):
        return "闲置"
    return None


# ══════════════════════════════════════════════════════════════════════
#  Config structure validation (operates on a candidate dict)
# ══════════════════════════════════════════════════════════════════════

def validate_config_structure(cfg):
    """Validate a config dict structure. Returns (ok, errors_list)."""
    errors = []

    required = ["development_platforms", "hijack_platforms", "disabled_platforms", "platform_rows"]
    for key in required:
        if key not in cfg:
            errors.append(f"缺少必需字段: {key}")

    if errors:
        return False, errors

    dev = set(cfg.get("development_platforms", []))
    hij = set(cfg.get("hijack_platforms", []))
    dis = set(cfg.get("disabled_platforms", []))

    for p in dev & hij:
        errors.append(f"平台 {p} 同时出现在 development 和 hijack")
    for p in dev & dis:
        errors.append(f"平台 {p} 同时出现在 development 和 disabled")
    for p in hij & dis:
        errors.append(f"平台 {p} 同时出现在 hijack 和 disabled")

    rows = cfg.get("platform_rows", {})
    needs_rows = dev | dis
    for p in sorted(needs_rows):
        if p not in rows:
            errors.append(f"平台 {p} 在列表中但缺少 platform_rows 配置")
        else:
            r = rows[p]
            if r.get("daily_row") is None:
                errors.append(f"平台 {p} 缺少 daily_row")
            if r.get("monthly_row") is None:
                errors.append(f"平台 {p} 缺少 monthly_row")

    return len(errors) == 0, errors


def validate_config_with_excel(cfg, data_folder):
    """Validate candidate config against current Excel.
    Returns (ok, errors, warnings).
    """
    errors = []
    warnings = []

    excel_path, _ = find_main_excel(data_folder)
    if excel_path is None:
        errors.append("未找到地推 Excel 文件")
        return False, errors, warnings

    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except ImportError:
        warnings.append("openpyxl 未安装，跳过 Excel 交叉验证")
        return True, errors, warnings
    except Exception as e:
        errors.append(f"无法打开 Excel: {str(e)[:100]}")
        return False, errors, warnings

    try:
        sheets = set(wb.sheetnames)
        daily_sn = _find_daily_sheet(wb)
        if daily_sn is None:
            errors.append("当日汇总 sheet 未找到")
            return False, errors, warnings

        ws_daily = wb[daily_sn]
        ws_monthly = wb[wb.sheetnames[0]]
        rows = cfg.get("platform_rows", {})

        for p in cfg.get("development_platforms", []):
            r = rows.get(p, {})
            dr = r.get("daily_row")
            mr = r.get("monthly_row")

            # Sheet check
            if p not in sheets:
                errors.append(f"{p}: Excel 缺少 sheet")
                continue

            # Daily row check
            if dr is not None:
                cell = str(_safe_read(ws_daily, dr, 1) or "")
                if p not in cell:
                    errors.append(f"{p}: daily_row={dr} 内容='{cell[:20]}' 不匹配")

            # Monthly row check — platform code is in column 3
            if mr is not None:
                cell = str(_safe_read(ws_monthly, mr, 3) or "")
                if p not in cell:
                    warnings.append(f"{p}: monthly_row={mr} col2='{cell[:20]}' 不匹配")

        # Duplicate file check already done by find_main_excel

        wb.close()
    except Exception as e:
        errors.append(f"Excel 验证异常: {str(e)[:100]}")
        return False, errors, warnings

    return len(errors) == 0, errors, warnings


# ══════════════════════════════════════════════════════════════════════
#  Safe atomic reload
# ══════════════════════════════════════════════════════════════════════

def reload_config(data_folder=None):
    """Safely reload config.json: validate candidate → only replace if OK.
    Returns (success, message, old_version, new_version).
    """
    old_v, old_u, _ = get_config_version()

    if not os.path.exists(CONFIG_FILE):
        return False, f"❌ 配置文件不存在: {CONFIG_FILE}", old_v, None

    # ── 1. Read candidate ──
    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            new_cfg = json.load(f)
    except json.JSONDecodeError as e:
        return False, f"❌ JSON 解析错误: {e}", old_v, None
    except Exception as e:
        return False, f"❌ 读取配置文件失败: {e}", old_v, None

    new_v = new_cfg.get("config_version")
    new_u = new_cfg.get("updated_at", "unknown")

    # ── 2. Structure validation ──
    ok, errors = validate_config_structure(new_cfg)
    if not ok:
        msg = "❌ 配置结构验证失败，保留旧配置:\n" + "\n".join(f"  • {e}" for e in errors)
        return False, msg, old_v, new_v

    # ── 3. Excel cross-validation (if data_folder provided) ──
    excel_warnings = []
    if data_folder:
        ok_excel, excel_errs, excel_warns = validate_config_with_excel(new_cfg, data_folder)
        if not ok_excel:
            msg = "❌ Excel 交叉验证失败，保留旧配置:\n" + "\n".join(f"  • {e}" for e in excel_errs)
            return False, msg, old_v, new_v
        excel_warnings = excel_warns

    # ── 4. Replace cache atomically ──
    global _cfg
    saved = _cfg
    try:
        _cfg = new_cfg
    except Exception as e:
        _cfg = saved
        return False, f"❌ 缓存替换失败: {e}", old_v, new_v

    # ── 5. Build result ──
    msg_lines = [
        "✅ 配置已重新加载",
        f"旧版本: {old_v}  新版本: {new_v}",
        f"更新时间: {new_u}",
        "",
        f"开发平台 ({len(get_development_platforms())}): {', '.join(get_development_platforms())}",
        f"劫持平台 ({len(get_hijack_platforms())}): {', '.join(get_hijack_platforms()) or '(无)'}",
        f"停用平台 ({len(get_disabled_platforms())}): {', '.join(get_disabled_platforms()) or '(无)'}",
    ]
    idle = _get_idle_platforms()
    if idle:
        msg_lines.append(f"闲置平台 ({len(idle)}): {', '.join(idle)}")

    if excel_warnings:
        msg_lines.append(f"\n⚠️ {len(excel_warnings)} 个 Excel Warning:")
        for w in excel_warnings:
            msg_lines.append(f"  • {w}")

    error_count = 0  # we already blocked on errors
    msg_lines.append(f"\nERROR: {error_count}  WARNING: {len(excel_warnings)}")
    msg_lines.append("旧配置继续生效: 否（已替换为新配置）")

    return True, "\n".join(msg_lines), old_v, new_v


# ══════════════════════════════════════════════════════════════════════
#  Single platform diagnostics  (/check_platform)
# ══════════════════════════════════════════════════════════════════════

def check_single_platform(code, data_folder):
    """Comprehensive single-platform check against Excel.
    Returns dict with diagnostic info.
    """
    code_upper = code.upper() if code else ""

    # Find original-case code from config
    original_code = None
    for c in get_all_configured_platforms():
        if c.upper() == code_upper:
            original_code = c
            break
    if original_code is None:
        for c in get_hijack_platforms():
            if c.upper() == code_upper:
                original_code = c
                break

    if original_code is None:
        all_avail = sorted(set(
            get_development_platforms() + get_hijack_platforms() +
            get_disabled_platforms() + _get_idle_platforms()
        ))
        return {
            "code": code, "found": False, "status": "ERROR",
            "message": f"平台 '{code}' 未找到", "available": all_avail,
        }

    code = original_code
    category = _classify(code)
    region = get_region(code)
    row_info = _rows().get(code, {})
    daily_row = row_info.get("daily_row")
    monthly_row = row_info.get("monthly_row")

    result = {
        "code": code, "found": True, "category": category,
        "region": region,
        "is_dev": code in get_development_platforms(),
        "is_hijack": code in get_hijack_platforms(),
        "daily_row": daily_row, "monthly_row": monthly_row,
        "status": "PASS", "warnings": [], "errors": [], "data": {},
    }

    # Hijack platforms skip sheet/row checks
    if category == "劫持":
        result["sheet_exists"] = "N/A (劫持)"
        result["daily_row_content"] = "N/A"
        result["monthly_row_content"] = "N/A"
        return result

    excel_path, _ = find_main_excel(data_folder)
    if excel_path is None:
        result["status"] = "WARNING"
        result["errors"].append("未找到 Excel 文件")
        return result

    result["excel_file"] = os.path.basename(excel_path)

    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except ImportError:
        result["warnings"].append("openpyxl 未安装")
        return result
    except Exception as e:
        result["status"] = "ERROR"
        result["errors"].append(f"无法打开 Excel: {str(e)[:100]}")
        return result

    try:
        sheets = wb.sheetnames

        # Sheet
        result["sheet_exists"] = code in sheets
        if not result["sheet_exists"]:
            result["status"] = "ERROR"
            result["errors"].append(f"Excel 中缺少 {code} sheet")

        # Daily row
        daily_sn = _find_daily_sheet(wb)
        if daily_sn and daily_row is not None:
            ws = wb[daily_sn]
            cell_val = str(_safe_read(ws, daily_row, 1) or "")
            result["daily_row_content"] = cell_val[:40]

            if code not in cell_val:
                result["status"] = "ERROR"
                result["errors"].append(f"daily_row={daily_row} 内容不匹配: '{cell_val[:30]}'")

            # Data fields
            fields = {
                "注册": 6, "FTD": 7, "首存金额": 19,
                "总充值": 22, "总提款": 23, "存提差": 24, "ROI": 27,
            }
            all_none = True
            any_none = False
            for fname, fcol in fields.items():
                v = _safe_read(ws, daily_row, fcol)
                result["data"][fname] = v
                if _is_truly_empty(v):
                    any_none = True
                else:
                    all_none = False

            if all_none and category == "开发":
                result["status"] = "ERROR"
                result["errors"].append("所有关键字段为空")
            elif any_none and category == "开发":
                if result["status"] == "PASS":
                    result["status"] = "WARNING"
                result["warnings"].append("部分字段为空")

            # Latest data date from own sheet
            if result["sheet_exists"] and code in sheets:
                ws_plat = wb[code]
                last_date = None
                for row_idx in range(ws_plat.max_row, 5, -1):
                    d = ws_plat.cell(row=row_idx, column=1).value
                    f = ws_plat.cell(row=row_idx, column=8).value
                    if d is not None and hasattr(d, 'strftime') and f is not None:
                        try:
                            float(f)
                            last_date = d
                            break
                        except (ValueError, TypeError):
                            continue
                if last_date:
                    result["data"]["最后数据日期"] = (
                        last_date.strftime("%Y-%m-%d")
                        if hasattr(last_date, 'strftime') else str(last_date)[:10]
                    )
                elif category == "开发":
                    result["warnings"].append("独立 sheet 中无最新数据日期")
        elif daily_sn is None:
            if category == "开发":
                result["status"] = "ERROR"
                result["errors"].append("当日汇总 sheet 未找到")

        # Monthly row — platform code is in column 3
        if monthly_row is not None and sheets:
            ws0 = wb[wb.sheetnames[0]]
            cell_val = str(_safe_read(ws0, monthly_row, 3) or "")
            result["monthly_row_content"] = cell_val[:40]
            if code not in cell_val:
                if result["status"] == "PASS":
                    result["status"] = "WARNING"
                result["warnings"].append(f"monthly_row={monthly_row} col3 不匹配: '{cell_val[:30]}'")

        wb.close()
    except Exception as e:
        result["status"] = "ERROR"
        result["errors"].append(f"Excel 读取异常: {str(e)[:100]}")

    # Final status resolution
    if result["errors"]:
        result["status"] = "ERROR"
    elif result["warnings"] and result["status"] == "PASS":
        result["status"] = "WARNING"

    return result


def format_check_result(result):
    """Format check_single_platform result as Telegram message."""
    if not result.get("found"):
        msg = f"❌ 平台 '{result['code']}' 未找到\n\n可查询的平台:\n"
        for p in result.get("available", []):
            msg += f"  {p}\n"
        return msg.strip()

    icon = {"PASS": "✅", "WARNING": "⚠️", "ERROR": "❌"}.get(result["status"], "❓")

    lines = [
        f"{icon} <b>平台检查: {result['code']}</b>", "",
        f"分类: {result['category']}",
        f"区域: {result['region'] or 'N/A'}",
        f"参与开发推送: {'是' if result.get('is_dev') else '否'}",
        f"参与劫持推送: {'是' if result.get('is_hijack') else '否'}",
        f"每日行号: {result.get('daily_row', 'N/A')}",
        f"月度行号: {result.get('monthly_row', 'N/A')}",
        f"独立 Sheet: {result.get('sheet_exists', 'N/A')}",
        f"每日行号内容: {result.get('daily_row_content', 'N/A')}",
        f"月度行号内容: {result.get('monthly_row_content', 'N/A')}",
    ]

    data = result.get("data", {})
    if data:
        lines.append("")
        lines.append("<b>今日数据:</b>")
        for k, v in data.items():
            if k == "最后数据日期":
                continue
            if v is None:
                lines.append(f"  {k}: ⚠️ None")
            elif isinstance(v, float):
                lines.append(f"  {k}: {v:,.0f}" if abs(v) >= 1 else f"  {k}: {v:.2f}")
            else:
                lines.append(f"  {k}: {v}")
        ld = data.get("最后数据日期")
        if ld:
            lines.append(f"  最后数据日期: {ld}")

    if result.get("warnings"):
        lines.append("")
        lines.append("<b>⚠️ Warnings:</b>")
        for w in result["warnings"]:
            lines.append(f"  • {w}")

    if result.get("errors"):
        lines.append("")
        lines.append("<b>❌ Errors:</b>")
        for e in result["errors"]:
            lines.append(f"  • {e}")

    lines.append("")
    lines.append(f"检查结果: {result['status']}")
    if result.get("excel_file"):
        lines.append(f"Excel: {result['excel_file']}")

    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════
#  /status and /data_status — admin commands
# ══════════════════════════════════════════════════════════════════════

def format_bot_status(data_folder=None):
    """/status — bot + system + platform summary. Read-only."""
    import sys, subprocess, platform
    from datetime import datetime

    lines = ["📊 <b>Bot 状态</b>", ""]

    # Git
    git_commit, git_branch = "N/A", "N/A"
    try:
        r = subprocess.run(["git", "log", "-1", "--oneline"], capture_output=True,
                           text=True, timeout=5, cwd=os.path.dirname(os.path.abspath(__file__)))
        if r.returncode == 0:
            git_commit = r.stdout.strip()
        r2 = subprocess.run(["git", "branch", "--show-current"], capture_output=True,
                            text=True, timeout=5, cwd=os.path.dirname(os.path.abspath(__file__)))
        if r2.returncode == 0:
            git_branch = r2.stdout.strip()
    except Exception:
        pass

    # Python + PID
    py_ver = platform.python_version()
    pid = os.getpid()

    # Config version
    v, u, _ = get_config_version()
    v_str = str(v) if v is not None else "N/A"

    lines.append("━━━━━━━━━━━━━━━━")
    lines.append("🟢 <b>运行状态</b>")
    lines.append(f"├ 进程: <b>运行中</b>（当前命令由 Bot 进程响应）")
    lines.append(f"├ PID: <code>{pid}</code>")
    lines.append(f"├ Python: <code>{py_ver}</code>")
    lines.append(f"└ 时间: <code>{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</code>")
    lines.append("")
    lines.append("🛠 <b>版本</b>")
    lines.append(f"├ Commit: <code>{git_commit}</code>")
    lines.append(f"├ 分支: <code>{git_branch}</code>")
    lines.append(f"└ 配置版本: <code>{v_str}</code>")

    # Platform lists
    dev = get_development_platforms()
    hij = get_hijack_platforms()
    dis = get_disabled_platforms()
    lines.append("")
    lines.append("🏷 <b>站点范围</b>")
    lines.append(f"├ 开发 ({len(dev)}): {', '.join(dev)}")
    lines.append(f"├ 劫持 ({len(hij)}): {', '.join(hij)}")
    lines.append(f"└ 停用 ({len(dis)}): {', '.join(dis)}")

    # Instance count + runtime env — cross-platform
    try:
        import _runtime as _rt
        instance_count = _rt.instance_count()
        runtime_env = _rt.runtime_label()
    except Exception:
        instance_count = "未检查"
        runtime_env = "未知"

    # Setup mode indicator
    setup_mode = False
    try:
        import _runtime as _rt
        setup_mode = _rt.is_setup_mode()
    except Exception:
        setup_mode = False

    lines.append("")
    lines.append("🌐 <b>环境</b>")
    lines.append(f"├ 运行环境: <code>{runtime_env}</code>")
    lines.append(f"├ bot_listener 实例数: <code>{instance_count}</code>")
    lines.append(f"└ 模式: <code>{'SETUP MODE' if setup_mode else '正式'}</code>")

    if setup_mode:
        lines.append("")
        lines.append("⚠️ <b>SETUP MODE</b> — 仅允许上传 Excel，业务推送已禁用")

    return "\n".join(lines)


def format_data_status(data_folder=None):
    """/data_status — Excel + snapshot status. Read-only."""
    from datetime import datetime, timedelta

    lines = ["📊 <b>数据状态</b>", ""]

    if data_folder is None:
        data_folder = _resolve_data_folder()

    def _ok_icon(cond):
        return "✅" if cond else "❌"

    # Development
    dev_info = get_active_excel_verbose("development")
    lines.append("━━━━━━━━━━━━━━━━")
    lines.append("📁 <b>开发数据</b>")
    if dev_info.get("success"):
        lines.append(f"├ 文件: <code>{dev_info['fname']}</code>")
        lines.append(f"├ data_date: <code>{dev_info.get('data_date', 'N/A')}</code>")
        lines.append(f"└ 存在: {_ok_icon(dev_info.get('size', 0) > 0)}")
        dev_date = dev_info.get("data_date")
    else:
        lines.append(f"└ 错误: {dev_info.get('errors', [])}")
        dev_date = None

    # Latest dev snapshot — only standard names, sorted by parsed date
    import re as _re_snap
    archive_dir = _runtime.archive_dir() if _runtime else os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'comparison_archive')
    dev_snaps = []
    if os.path.isdir(archive_dir):
        for f in os.listdir(archive_dir):
            m = _re_snap.match(r'^development_(\d{4}-\d{2}-\d{2})\.xlsx$', f)
            if m:
                dev_snaps.append((m.group(1), f))
    dev_snaps.sort(key=lambda x: x[0])
    latest_dev_snap = dev_snaps[-1][1] if dev_snaps else "(无)"
    lines.append(f"├ 最新快照: <code>{latest_dev_snap}</code>")
    lines.append("")

    # Hijack
    hij_info = get_active_excel_verbose("hijack")
    lines.append("━━━━━━━━━━━━━━━━")
    lines.append("📁 <b>劫持数据</b>")
    if hij_info.get("success"):
        lines.append(f"├ 文件: <code>{hij_info['fname']}</code>")
        lines.append(f"├ data_date: <code>{hij_info.get('data_date', 'N/A')}</code>")
        lines.append(f"└ 存在: {_ok_icon(hij_info.get('size', 0) > 0)}")
        hij_date = hij_info.get("data_date")
    else:
        lines.append(f"└ 错误: {hij_info.get('errors', [])}")
        hij_date = None

    hij_snaps = []
    if os.path.isdir(archive_dir):
        for f in os.listdir(archive_dir):
            m = _re_snap.match(r'^hijack_(\d{4}-\d{2}-\d{2})\.xlsx$', f)
            if m:
                hij_snaps.append((m.group(1), f))
    hij_snaps.sort(key=lambda x: x[0])
    latest_hij_snap = hij_snaps[-1][1] if hij_snaps else "(无)"
    lines.append(f"└ 最新快照: <code>{latest_hij_snap}</code>")
    lines.append("")

    # Yesterday / last-month snapshots + compare feasibility
    def _snap_exists(kind, date_str):
        return os.path.isfile(os.path.join(archive_dir, f"{kind}_{date_str}.xlsx"))

    lines.append("━━━━━━━━━━━━━━━━")
    lines.append("📈 <b>对比可用性</b>")
    if dev_date:
        try:
            dt = datetime.strptime(dev_date, '%Y-%m-%d')
            y = (dt - timedelta(days=1)).strftime('%Y-%m-%d')
            m = (dt.replace(month=dt.month - 1) if dt.month > 1
                 else dt.replace(year=dt.year - 1, month=12)).strftime('%Y-%m-%d')
            dev_y = _snap_exists('development', y)
            dev_m = _snap_exists('development', m)
            lines.append(f"├ 开发昨日 {y}: {_ok_icon(dev_y)}")
            lines.append(f"├ 开发上月 {m}: {_ok_icon(dev_m)}")
            lines.append(f"└ 开发对比图: {'可以生成' if dev_y and dev_m else '部分/无法生成（缺失显示警告）'}")
        except Exception:
            pass

    if hij_date:
        try:
            dt = datetime.strptime(hij_date, '%Y-%m-%d')
            y = (dt - timedelta(days=1)).strftime('%Y-%m-%d')
            m = (dt.replace(month=dt.month - 1) if dt.month > 1
                 else dt.replace(year=dt.year - 1, month=12)).strftime('%Y-%m-%d')
            lines.append(f"├ 劫持昨日 {y}: {_ok_icon(_snap_exists('hijack', y))}")
            lines.append(f"└ 劫持上月 {m}: {_ok_icon(_snap_exists('hijack', m))}")
        except Exception:
            pass

    return "\n".join(lines)


def _snapshot_detail(kind, date_str):
    """Check a single snapshot file: exists, size, opens, internal date match.
    Returns a formatted line."""
    archive_dir = _runtime.archive_dir() if _runtime else os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'comparison_archive')
    fname = f"{kind}_{date_str}.xlsx"
    path = os.path.join(archive_dir, fname)

    if not os.path.isfile(path):
        return f"  ❌ <code>{fname}</code> — 不存在"

    size = os.path.getsize(path)
    if size == 0:
        return f"  ⚠️ <code>{fname}</code> — 存在但 0 字节"

    # Try open
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        wb.close()
    except Exception as e:
        return f"  ⚠️ <code>{fname}</code> — 存在 {size:,}B 但无法打开 ({str(e)[:40]})"

    # Internal date
    internal = _detect_data_date(path)
    date_match = "匹配" if internal == date_str else (f"不匹配({internal})" if internal else "无法识别")
    ok_mark = "✅" if internal == date_str else "⚠️"
    return f"  {ok_mark} <code>{fname}</code> — {size:,}B · 可打开 · 内部日期:{date_match}"


def format_snapshot_check(data_folder=None):
    """Snapshot check — target-date-derived existence + integrity. Read-only."""
    lines = ["📊 <b>快照检查</b>", ""]

    # Resolve both current Excels
    dev_info = get_active_excel_verbose("development")
    hij_info = get_active_excel_verbose("hijack")

    def _targets(date_str):
        from datetime import timedelta
        dt = datetime.strptime(date_str, '%Y-%m-%d')
        y = (dt - timedelta(days=1)).strftime('%Y-%m-%d')
        m = (dt.replace(month=dt.month - 1) if dt.month > 1
             else dt.replace(year=dt.year - 1, month=12)).strftime('%Y-%m-%d')
        return y, m

    # Development
    lines.append("━━━━━━━━━━━━━━━━")
    lines.append("📁 <b>开发快照</b>")
    if dev_info.get("success"):
        dev_date = dev_info.get("data_date")
        lines.append(f"├ 当前 data_date: <code>{dev_date}</code>")
        if dev_date:
            y, m = _targets(dev_date)
            lines.append(f"├ 昨日目标 <code>{y}</code>:")
            lines.append(_snapshot_detail("development", y))
            lines.append(f"└ 上月目标 <code>{m}</code>:")
            lines.append(_snapshot_detail("development", m))
    else:
        lines.append(f"└ ❌ {dev_info.get('errors', [])}")
    lines.append("")

    # Hijack
    lines.append("━━━━━━━━━━━━━━━━")
    lines.append("📁 <b>劫持快照</b>")
    if hij_info.get("success"):
        hij_date = hij_info.get("data_date")
        lines.append(f"├ 当前 data_date: <code>{hij_date}</code>")
        if hij_date:
            y, m = _targets(hij_date)
            lines.append(f"├ 昨日目标 <code>{y}</code>:")
            lines.append(_snapshot_detail("hijack", y))
            lines.append(f"└ 上月目标 <code>{m}</code>:")
            lines.append(_snapshot_detail("hijack", m))
    else:
        lines.append(f"└ ❌ {hij_info.get('errors', [])}")
    lines.append("")

    return "\n".join(lines)


def format_platform_status(data_folder=None):
    """Upgraded /platforms output with runtime Sheet/Daily/Monthly checks."""
    v, u, _ = get_config_version()
    v_str = str(v) if v is not None else "N/A"

    excel_path, excel_name = find_main_excel(data_folder) if data_folder else (None, None)

    lines = [
        "📋 <b>平台配置状态</b>",
        f"Config Version: {v_str}",
        f"Updated At: {u}",
    ]
    if excel_name:
        lines.append(f"正式文件: {excel_name}")
    elif data_folder:
        _load()
        af = _cfg.get("active_excel_file", "").strip()
        if af:
            lines.append(f"⚠️ 配置文件缺失: {af}")
    lines.append("")

    # Open Excel once for all status checks
    wb = None
    excel_path = None
    if data_folder:
        excel_path, _ = find_main_excel(data_folder)
        if excel_path:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(excel_path, data_only=True)
            except Exception:
                wb = None

    def _sheet_ok(code):
        if wb is None: return "?"
        return "✔" if code in wb.sheetnames else "✘"

    def _daily_ok(code):
        if wb is None: return "?"
        r = _rows().get(code, {}).get("daily_row")
        if r is None: return "?"
        sn = _find_daily_sheet(wb)
        if sn is None: return "?"
        try:
            v = str(wb[sn].cell(row=r, column=1).value or "")
            return "✔" if code in v else "⚠"
        except Exception:
            return "⚠"

    def _monthly_ok(code):
        if wb is None: return "?"
        r = _rows().get(code, {}).get("monthly_row")
        if r is None: return "?"
        try:
            # Platform code is in column 3 of the monthly cost sheet
            v = str(wb[wb.sheetnames[0]].cell(row=r, column=3).value or "")
            return "✔" if code in v else "⚠"
        except Exception:
            return "⚠"

    dev = get_development_platforms()
    lines.append("<b>开发平台:</b>")
    for p in dev:
        s, d, m = _sheet_ok(p), _daily_ok(p), _monthly_ok(p)
        lines.append(f"  {p} | ACTIVE | Sheet {s} | Daily {d} | Monthly {m} | Push ✔")
    if not dev:
        lines.append("  (无)")
    lines.append("")

    hij = get_hijack_platforms()
    lines.append("<b>劫持平台:</b>")
    for p in hij:
        lines.append(f"  {p} | ACTIVE | Hijack Push ✔ | Development Push ✘")
    if not hij:
        lines.append("  (无)")
    lines.append("")

    dis = get_disabled_platforms()
    lines.append("<b>停用平台:</b>")
    for p in dis:
        s = _sheet_ok(p)
        lines.append(f"  {p} | DISABLED | Sheet {s} | Push ✘")
    if not dis:
        lines.append("  (无)")
    lines.append("")

    idle = _get_idle_platforms()
    if idle:
        lines.append("<b>闲置平台:</b>")
        for p in idle:
            lines.append(f"  {p} | IDLE | Rows Configured ✔ | Push ✘")
        lines.append("")

    if wb:
        wb.close()

    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════
#  Startup check — enhanced
# ══════════════════════════════════════════════════════════════════════

def startup_check(data_folder):
    """Enhanced startup check. Returns list of status strings."""
    warnings = []

    # Version
    v, u, p = get_config_version()
    if v is None:
        warnings.append("⚠️ config.json 缺少 config_version 字段")
    else:
        warnings.append(f"✅ Config version: {v}, updated: {u}")
    warnings.append(f"   Path: {p}")

    # Structure
    _load()
    for key in ("development_platforms", "hijack_platforms", "disabled_platforms", "platform_rows"):
        if key not in _cfg:
            warnings.append(f"⚠️ config.json 缺少字段: {key}")

    dev = get_development_platforms()
    hij = get_hijack_platforms()
    dis = get_disabled_platforms()
    rows = _rows()
    all_cfg = set(rows.keys())

    # Duplicates
    for p in set(dev) & set(hij):
        warnings.append(f"⚠️ {p} 同时出现在 development 和 hijack")
    for p in set(dev) & set(dis):
        warnings.append(f"⚠️ {p} 同时出现在 development 和 disabled")
    for p in set(hij) & set(dis):
        warnings.append(f"⚠️ {p} 同时出现在 hijack 和 disabled")

    # Row completeness
    needs_rows = set(dev + dis)
    for p in sorted(needs_rows - all_cfg):
        warnings.append(f"⚠️ {p} 在平台列表中但缺少 platform_rows 配置")

    for code, info in rows.items():
        if info.get("daily_row") is None:
            warnings.append(f"⚠️ {code} 缺少 daily_row")
        if info.get("monthly_row") is None:
            warnings.append(f"⚠️ {code} 缺少 monthly_row")

    # Excel cross-check
    if not data_folder or not os.path.isdir(data_folder):
        return warnings

    try:
        import openpyxl
    except ImportError:
        warnings.append("⚠️ openpyxl 未安装，跳过 Excel 检查")
        return warnings

    excel_path, excel_name = find_main_excel(data_folder)

    if excel_path is None:
        _load()
        af = _cfg.get("active_excel_file", "").strip()
        if af:
            warnings.append(f"❌ 配置文件不存在: {af}")
        else:
            warnings.append("⚠️ 未找到地推 Excel 文件且未配置 active_excel_file")
        return warnings

    warnings.append(f"📂 正式文件: {excel_name}")

    # Show other same-month files as INFO
    other = get_other_monthly_files(data_folder)
    if other:
        warnings.append(f"ℹ️ 发现其他同月 Excel 文件，但未使用:")
        for o in other:
            warnings.append(f"   - {o}")

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        warnings.append(f"❌ 无法打开 Excel: {str(e)[:100]}")
        return warnings

    try:
        sheets = set(wb.sheetnames)
        daily_sn = _find_daily_sheet(wb)

        for p in dev:
            r = rows.get(p, {})
            dr = r.get("daily_row")
            mr = r.get("monthly_row")

            if p not in sheets:
                warnings.append(f"❌ {p}: Excel 缺少 sheet → ERROR")
                continue
            warnings.append(f"✅ {p}: Sheet 存在")

            # Daily row + data fields
            if daily_sn and dr:
                ws = wb[daily_sn]
                cell = str(_safe_read(ws, dr, 1) or "")
                if p not in cell:
                    warnings.append(f"❌ {p}: daily_row={dr} 内容='{cell[:20]}' 不匹配 → ERROR")
                else:
                    ftd = _safe_read(ws, dr, 7)
                    reg = _safe_read(ws, dr, 6)
                    diff = _safe_read(ws, dr, 24)

                    if _is_truly_empty(ftd) and _is_truly_empty(reg) and _is_truly_empty(diff):
                        warnings.append(f"❌ {p}: 关键字段全部为空 → ERROR")
                    else:
                        missing = []
                        if _is_truly_empty(ftd): missing.append("FTD=None")
                        if _is_truly_empty(reg): missing.append("Reg=None")
                        if _is_truly_empty(diff): missing.append("Diff=None")

                        if missing:
                            warnings.append(f"⚠️ {p}: 部分字段为空 ({', '.join(missing)}) → WARNING")
                        else:
                            # Check latest date
                            if p in sheets:
                                ws_plat = wb[p]
                                last_date = None
                                for row_idx in range(ws_plat.max_row, 5, -1):
                                    d = ws_plat.cell(row=row_idx, column=1).value
                                    f = ws_plat.cell(row=row_idx, column=8).value
                                    if d and hasattr(d, 'strftime') and f is not None:
                                        try:
                                            float(f)
                                            last_date = d
                                            break
                                        except (ValueError, TypeError):
                                            continue
                                if last_date:
                                    ds = last_date.strftime("%Y-%m-%d") if hasattr(last_date, 'strftime') else str(last_date)
                                    today = datetime.now().strftime("%Y-%m-%d")
                                    if ds != today:
                                        warnings.append(f"⚠️ {p}: 数据日期={ds}（不是今天 {today}）→ WARNING")
                                    else:
                                        warnings.append(f"✅ {p}: PASS (FTD={ftd}, Reg={reg})")
                                else:
                                    warnings.append(f"⚠️ {p}: 无最新数据日期 → WARNING")
            elif daily_sn is None:
                warnings.append(f"⚠️ {p}: 当日汇总 sheet 未找到")

            # Monthly row — column 3
            if mr is not None and sheets:
                ws0 = wb[wb.sheetnames[0]]
                cell = str(_safe_read(ws0, mr, 3) or "")
                if p not in cell:
                    warnings.append(f"⚠️ {p}: monthly_row={mr} col3 不匹配 → WARNING")

        # Excel platforms not in config
        hijack_codes = set(get_hijack_platforms())
        excel_plats = {
            s for s in sheets
            if any(s.upper().startswith(pre) for pre in REGION_PREFIXES)
            and s not in hijack_codes
        }
        unconfigured = excel_plats - all_cfg
        if unconfigured:
            warnings.append(f"⚠️ Excel 中存在但 config 未配置: {', '.join(sorted(unconfigured))}")

        wb.close()
    except Exception as e:
        warnings.append(f"❌ Excel 检查异常: {str(e)[:100]}")

    err_count = sum(1 for w in warnings if "→ ERROR" in w)
    warn_count = sum(1 for w in warnings if "→ WARNING" in w)
    warnings.append(f"\n  检查完成: {len(warnings)} 条消息 ({err_count} ERROR, {warn_count} WARNING)")

    return warnings


# ══════════════════════════════════════════════════════════════════════
#  Pre-push guard — FAIL-CLOSED
# ══════════════════════════════════════════════════════════════════════

def pre_push_guard(data_folder):
    """Check before sending push. FAIL-CLOSED: blocks on any structural
    or data error. Only allows push when at least one dev platform has
    valid data and no structural errors exist.

    Returns (ok: bool, warnings: list, block_reason: str).
    """
    warnings = []

    # ── 1. Dev list non-empty ──
    dev = get_development_platforms()
    if not dev:
        return False, [], "开发平台列表为空"

    # ── 2. No duplicates ──
    if len(dev) != len(set(dev)):
        return False, [], "开发平台列表存在重复项"

    # ── 3. All have platform_rows ──
    rows = _rows()
    for p in dev:
        if p not in rows:
            return False, [], f"平台 {p} 缺少 platform_rows 配置"
        if rows[p].get("daily_row") is None:
            return False, [], f"平台 {p} 缺少 daily_row"

    # ── 4. Excel file exists ──
    try:
        import openpyxl
    except ImportError:
        return False, [], "openpyxl 不可用，无法验证 Excel 数据"

    excel_path, excel_name = find_main_excel(data_folder)
    if excel_path is None:
        _load()
        active_file = _cfg.get("active_excel_file", "").strip()
        if active_file:
            return False, [], f"配置的正式 Excel 文件不存在:\n{active_file}\n\n请检查 新建文件夹 目录"
        # No active_excel_file and multiple/no candidates
        other = get_other_monthly_files(data_folder)
        if other:
            return False, [], f"存在多个同月候选文件且未配置 active_excel_file:\n" + "\n".join(f"  - {f}" for f in other) + "\n\n请在 config.json 中指定 active_excel_file"
        return False, [], "未找到地推 Excel 文件"

    # ── 5. Open Excel ──
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        return False, [], f"无法打开 Excel: {str(e)[:100]}"

    # ── 6. Find daily sheet ──
    try:
        daily_sn = _find_daily_sheet(wb)
        if daily_sn is None:
            wb.close()
            return False, [], "当日汇总 sheet 未找到"

        ws = wb[daily_sn]

        # ── 7. Check each dev platform ──
        all_empty = True
        any_valid = False
        platform_warnings = []

        for p in dev:
            dr = rows[p].get("daily_row")
            if dr is None:
                platform_warnings.append(f"{p}: 缺少 daily_row")
                continue

            # Check row content matches
            cell = str(_safe_read(ws, dr, 1) or "")
            if p not in cell:
                # Row mismatch — check if ALL platforms have this problem
                continue

            # Read data
            ftd = _safe_read(ws, dr, 7)
            reg = _safe_read(ws, dr, 6)

            ftd_ok = not _is_truly_empty(ftd)
            reg_ok = not _is_truly_empty(reg)

            if ftd_ok or reg_ok:
                all_empty = False
                any_valid = True

        wb.close()

        # ── 8. Decision ──
        if not any_valid:
            return False, [], "所有开发平台均无可读取的有效数据"

        if all_empty:
            return False, [], "所有开发平台数据为空"

        # At least one platform has data → OK, include per-platform warnings
        return True, platform_warnings, ""

    except Exception as e:
        # Guard自身异常 → FAIL-CLOSED
        try:
            wb.close()
        except Exception:
            pass
        return False, [], f"Pre-push guard 异常: {str(e)[:100]}"
