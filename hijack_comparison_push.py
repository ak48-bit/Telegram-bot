"""PH33 Hijack Comparison — generates combined comparison image.
Formal production module. Does NOT depend on generate_previews.py.
"""

import os, sys, io
from datetime import datetime, timedelta

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

import openpyxl
from PIL import Image, ImageDraw, ImageFont

FONT = 'C:/Windows/Fonts/msyh.ttc'
FONT_BOLD = 'C:/Windows/Fonts/msyhbd.ttc'
ARCHIVE_DIR = os.path.join(SCRIPT_DIR, 'data', 'comparison_archive')
OUT_DIR = os.path.join(SCRIPT_DIR, 'data', 'generated')
os.makedirs(OUT_DIR, exist_ok=True)

HIJACK_FIELDS = [
    ('线上组长', 2), ('线上人数', 3), ('数据消耗', 4), ('注册人数', 5),
    ('劫持人数', 6), ('每日平均劫持人数', 7), ('平均劫持成本', 8),
    ('新客首充金额', 11), ('新客首存总额', 12), ('新客总提款', 13), ('新客存提差', 14),
    ('当天累计存款', 15), ('当天累计提款', 16), ('当天存提差', 17),
]
FIELD_LABELS = [l for l, _ in HIJACK_FIELDS]


def _read_hijack_row(filepath):
    """Read PH33 当天数据汇总 row 4."""
    if not os.path.isfile(filepath):
        raise FileNotFoundError(f"文件不存在: {os.path.basename(filepath)}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if '当天数据汇总' not in wb.sheetnames:
        wb.close()
        raise ValueError(f"缺少 当天数据汇总 sheet: {os.path.basename(filepath)}")
    ws = wb['当天数据汇总']
    data = {}
    for label, col in HIJACK_FIELDS:
        v = ws.cell(row=4, column=col).value
        try: data[label] = float(v) if v is not None else 0.0
        except: data[label] = 0.0
    wb.close()
    return data


def _fv(v):
    if v is None: return '—'
    if isinstance(v, float) and abs(v) >= 1000: return f'{v:,.0f}'
    if isinstance(v, float) and v != int(v): return f'{v:,.2f}'
    if isinstance(v, float): return f'{int(v):,}'
    return str(v)


def _fd(v):
    if v is None: return '—'
    return f'{v:+,.0f}' if abs(v) >= 1 else f'{v:+,.2f}'


def _fp(pct):
    if pct is None: return '—'
    return f'{pct:+.1f}%'


def _fc(cur, prev):
    if cur is None or prev is None: return 0, None, '—'
    df = cur - prev
    if prev == 0: return df, None, '—'
    pc = (df / prev) * 100
    return df, pc, '↑' if df > 0 else ('↓' if df < 0 else '—')


def _build_section_image(title, subtitle, cur_data, prev_data):
    """Build one comparison section image (orange theme)."""
    ft = ImageFont.truetype(FONT_BOLD, 18)
    fs = ImageFont.truetype(FONT, 14)
    fh = ImageFont.truetype(FONT_BOLD, 13)
    fb = ImageFont.truetype(FONT, 12)
    ff = ImageFont.truetype(FONT, 10)

    headers = ['指标', '当前日', '对比日', '增减值', '增减比例', '变化']
    cw = [150, 120, 120, 120, 100, 60]
    rh, hh, th, sh = 28, 32, 42, 26

    rows = []
    for label in FIELD_LABELS:
        cv = cur_data.get(label); pv = prev_data.get(label)
        df, pc, ar = _fc(cv, pv)
        rows.append((label, _fv(cv), _fv(pv), _fd(df), _fp(pc), ar))

    tw = sum(cw)
    tth = th + sh + hh + len(rows) * rh + 32
    DB = (55, 55, 60); SB = (250, 240, 225); W = (255, 255, 255)
    GB = (235, 255, 235); RB = (255, 235, 235); GYB = (245, 245, 245)
    DK = (33, 37, 41); GY = (140, 140, 140)

    img = Image.new('RGB', (tw, tth), W); d = ImageDraw.Draw(img)
    d.rectangle([0, 0, tw - 1, th], fill=DB)
    tw2 = ft.getbbox(title)[2]; d.text(((tw - tw2) // 2, 10), title, fill=W, font=ft)
    y = th
    d.rectangle([0, y, tw - 1, y + sh], fill=SB)
    sw = fs.getbbox(subtitle)[2]; d.text(((tw - sw) // 2, y + 4), subtitle, fill=DK, font=fs)
    y += sh; x = 0
    for ci, h in enumerate(headers):
        wc = cw[ci]
        d.rectangle([x, y, x + wc - 1, y + hh], fill=DB)
        tw3 = fh.getbbox(h)[2]; d.text((x + (wc - tw3) // 2, y + 7), h, fill=W, font=fh); x += wc
    y += hh
    for ri, row_data in enumerate(rows):
        ar = row_data[5]
        bg = GB if ar == '↑' else (RB if ar == '↓' else (GYB if ri % 2 == 0 else W)); x = 0
        for ci, val in enumerate(row_data):
            wc = cw[ci]
            d.rectangle([x, y, x + wc - 1, y + rh], fill=bg, outline=(230, 230, 230))
            tw4 = fb.getbbox(str(val))[2]; d.text((x + (wc - tw4) // 2, y + 5), str(val), fill=DK, font=fb); x += wc
        y += rh
    ts = datetime.now().strftime('%Y-%m-%d %H:%M')
    footer = f'WFHDPbot  |  Generated: {ts}'
    fw6 = ff.getbbox(footer)[2]; d.text((tw - fw6 - 10, tth - 22), footer, fill=GY, font=ff)
    buf = io.BytesIO(); img.save(buf, format='PNG', optimize=True)
    return buf.getvalue()


def _build_warning_section(target_date_str):
    """Build warning section for missing snapshot."""
    img = Image.new('RGB', (670, 120), (255, 255, 255))
    d = ImageDraw.Draw(img)
    fw = ImageFont.truetype(FONT, 13)
    msg1 = f'⚠️ 缺少 {target_date_str} PH33 劫持历史快照'
    msg2 = '本次无法生成上月同日数据对比（未使用其他日期代替）'
    tw1 = fw.getbbox(msg1)[2]; tw2 = fw.getbbox(msg2)[2]
    d.text(((670 - tw1) // 2, 30), msg1, fill=(180, 120, 0), font=fw)
    d.text(((670 - tw2) // 2, 60), msg2, fill=(180, 120, 0), font=fw)
    buf = io.BytesIO(); img.save(buf, format='PNG', optimize=True)
    return buf.getvalue()


def generate_hijack_comparison(target_date_str):
    """Generate combined PH33 comparison long image.
    Args: target_date_str = '2026-07-26'
    Returns: (success, error_message, image_bytes, caption, metadata)
    """
    try:
        target_dt = datetime.strptime(target_date_str, '%Y-%m-%d')
        prev_day_dt = target_dt - timedelta(days=1)
        prev_month_dt = (target_dt.replace(month=target_dt.month - 1) if target_dt.month > 1
                         else target_dt.replace(year=target_dt.year - 1, month=12))
    except Exception as e:
        return False, f"日期格式错误: {target_date_str}", None, "", {}

    prev_day_str = prev_day_dt.strftime('%Y-%m-%d')
    prev_month_str = prev_month_dt.strftime('%Y-%m-%d')

    cur_path = os.path.join(ARCHIVE_DIR, f'hijack_{target_date_str}.xlsx')
    prev_day_path = os.path.join(ARCHIVE_DIR, f'hijack_{prev_day_str}.xlsx')
    prev_month_path = os.path.join(ARCHIVE_DIR, f'hijack_{prev_month_str}.xlsx')

    # Read current and previous day
    try:
        cur_data = _read_hijack_row(cur_path)
    except Exception as e:
        return False, f"读取当前日快照失败: {e}", None, "", {}

    try:
        prev_day_data = _read_hijack_row(prev_day_path)
    except Exception as e:
        return False, f"读取前一日快照失败: {e}", None, "", {}

    # Build day comparison section
    day_title = f'{target_dt.month}月{target_dt.day}日 vs {prev_day_dt.month}月{prev_day_dt.day}日 | PH33劫持数据对比'
    day_sub = f'数据口径：PH33 当天数据汇总'
    img_day = _build_section_image(day_title, day_sub, cur_data, prev_day_data)

    # Build month section
    month_title = f'{target_dt.month}月{target_dt.day}日 vs {prev_month_dt.month}月{prev_month_dt.day}日 | PH33劫持数据对比'
    month_sub = f'数据口径：PH33 当天数据汇总'

    if os.path.isfile(prev_month_path):
        try:
            prev_month_data = _read_hijack_row(prev_month_path)
            img_month = _build_section_image(month_title, month_sub, cur_data, prev_month_data)
            month_available = True
        except Exception as e:
            img_month = _build_warning_section(prev_month_str)
            month_available = False
    else:
        img_month = _build_warning_section(prev_month_str)
        month_available = False

    # Combine vertically
    img_day_pil = Image.open(io.BytesIO(img_day))
    img_month_pil = Image.open(io.BytesIO(img_month))
    combined = Image.new('RGB', (
        max(img_day_pil.width, img_month_pil.width),
        img_day_pil.height + img_month_pil.height
    ), (255, 255, 255))
    combined.paste(img_day_pil, (0, 0))
    combined.paste(img_month_pil, (0, img_day_pil.height))

    buf = io.BytesIO()
    combined.save(buf, format='PNG', optimize=True)
    img_bytes = buf.getvalue()

    caption = f'🛡️ PH33 劫持数据对比 — {target_dt.month}月{target_dt.day}日'
    metadata = {
        'target_date': target_date_str,
        'prev_day_date': prev_day_str,
        'prev_month_date': prev_month_str,
        'month_available': month_available,
        'cur_file': cur_path,
        'prev_day_file': prev_day_path,
        'prev_month_file': prev_month_path if month_available else None,
        'image_size': len(img_bytes),
    }

    return True, "", img_bytes, caption, metadata


def generate_hijack_comparison_file(target_date_str, output_dir=None):
    """Generate and save to file. Returns (success, error, filepath, caption, metadata)."""
    if output_dir is None:
        output_dir = OUT_DIR
    os.makedirs(output_dir, exist_ok=True)
    ok, err, img_bytes, caption, meta = generate_hijack_comparison(target_date_str)
    if not ok:
        return False, err, None, caption, meta
    fname = f"hijack_compare_{target_date_str}.png"
    fpath = os.path.join(output_dir, fname)
    with open(fpath, 'wb') as f:
        f.write(img_bytes)
    return True, "", fpath, caption, meta
