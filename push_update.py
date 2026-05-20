import os, json, urllib.request, re
from datetime import datetime
import openpyxl
from PIL import Image, ImageDraw, ImageFont

FONT_PATH = "C:/Windows/Fonts/msyh.ttc"
FONT_BOLD_PATH = "C:/Windows/Fonts/msyhbd.ttc"

TARGET_DIR = os.path.dirname(os.path.abspath(__file__))

# Load config
CONFIG_FILE = os.path.join(TARGET_DIR, "config.json")
def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}
CONFIG = load_config()
DATA_FOLDER = r"C:\Users\ak481\OneDrive\Desktop\新建文件夹"


def find_monthly_file(target_date, keywords, exclude_kw=None):
    """Find Excel file in DATA_FOLDER matching year+month and keywords."""
    if target_date is None:
        return None
    # Extract year and month from datetime or string
    if hasattr(target_date, 'strftime'):
        year_2d = target_date.strftime("%y")
        month = target_date.month
        month_str = target_date.strftime("%m")
        month_no_pad = str(month)
    else:
        parts = target_date.split("-")
        year_2d = parts[0][-2:]
        month = int(parts[1])
        month_str = f"{month:02d}"
        month_no_pad = str(month)

    # Build patterns: "26年05月", "26年5月", "2605", "265"
    patterns = [
        f"{year_2d}年{month_str}月",
        f"{year_2d}年{month_no_pad}月",
    ]

    for f in sorted(os.listdir(DATA_FOLDER)):
        if not f.endswith('.xlsx'):
            continue
        # Skip backup copies
        if '副本' in f or ' - Copy' in f:
            continue
        # Normalize brackets
        fname = f.replace("（", "(").replace("）", ")")
        # Check all keywords present
        kw_ok = True
        for kw in keywords:
            kw_norm = kw.replace("（", "(").replace("）", ")")
            if kw_norm not in fname:
                kw_ok = False
                break
        if not kw_ok:
            continue
        # Check exclude keywords
        if exclude_kw:
            ex_ok = True
            for ekw in exclude_kw:
                ekw_norm = ekw.replace("（", "(").replace("）", ")")
                if ekw_norm in fname:
                    ex_ok = False
                    break
            if not ex_ok:
                continue
        # Check year+month match
        for pat in patterns:
            if pat in fname:
                return os.path.join(DATA_FOLDER, f)
        # Fuzzy: year+month digits
        ym_digits = f"{year_2d}{month_str}"  # "2604"
        ym_digits_nopad = f"{year_2d}{month_no_pad}"  # "264"
        if ym_digits in fname or ym_digits_nopad in fname:
            return os.path.join(DATA_FOLDER, f)
    return None

TELEGRAM_BOT_TOKEN = "8731392429:AAFb6QywB4NG4TDTmeOtzDbS7IR_G95JzAI"
TELEGRAM_CHAT_ID = "-1003899337250"

PH_PLATFORMS = ["PH09", "PH09-2", "PH25", "PH18", "PH30", "PH05", "PH16"]
BD_PLATFORMS = ["BD02", "BD05"]
MM_PLATFORMS = ["MM01"]
ALL_PLATFORMS = PH_PLATFORMS + BD_PLATFORMS + MM_PLATFORMS

# Column display widths (visual)
COLS = [
    ("站点", 8),
    ("注册", 6),
    ("首存", 6),
    ("充值", 6),
    ("提款", 6),
    ("首存金额", 10),
    ("总充值", 8),
    ("总提款", 8),
    ("充提差", 8),
    ("新客单价", 10),
    ("投产比", 8),
    ("状态", 6),
]

TIPS = [
    "检查今日零单站点，立刻联系了解原因",
    "看昨日各站点FTD环比变化",
    "关注提款率超过70%的站点",
    "查看是否有连续3天掉量的站点",
    "检查新版本数据对比表是否更新",
    "抽查1-2个高FTD站点的玩家来源",
    "确认今天的异常预警已处理",
    "提醒团队多用1个社交平台",
    "对比本周和上周的各站点FTD趋势",
    "看看有没有人在用虚拟号或同IP的迹象",
    "确认职级考核标准是否已发给全员",
    "整理今天的数据，准备明天早上的汇报",
]


def _cell_float(ws, row, col):
    v = ws.cell(row=row, column=col).value
    if v is None:
        return 0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0


def _find_last_data_row(ws):
    for row_idx in range(ws.max_row, 5, -1):
        date_val = ws.cell(row=row_idx, column=1).value
        ftd_val = ws.cell(row=row_idx, column=8).value
        if date_val is not None and hasattr(date_val, 'strftime') and ftd_val is not None:
            try:
                float(ftd_val)
                return row_idx
            except (ValueError, TypeError):
                continue
    return None


def _build_dict(ws, row):
    c = lambda col: _cell_float(ws, row, col)
    return {
        "date": ws.cell(row=row, column=1).value,
        "register": int(c(7)),
        "ftd": int(c(8)),
        "dps_ppl": int(c(11)),
        "ftd_amount": c(20),
        "total_dps_amt": c(23),
        "total_wdr_amt": c(24),
        "diff": c(25),
        "new_cust_avg": c(15),
        "wdr_ppl": int(c(12)),
        "roi": c(34),
    }


def _find_data_row_by_date(ws, target_date_str):
    """Find row matching a specific date string like '2026-05-06'."""
    for row_idx in range(6, ws.max_row + 1):
        date_val = ws.cell(row=row_idx, column=1).value
        if date_val is not None and hasattr(date_val, 'strftime'):
            if date_val.strftime("%Y-%m-%d") == target_date_str:
                ftd_val = ws.cell(row=row_idx, column=8).value
                if ftd_val is not None:
                    try:
                        float(ftd_val)
                        return row_idx
                    except (ValueError, TypeError):
                        return None
    return None


def read_platform_data(ws, target_date=None):
    if target_date:
        row = _find_data_row_by_date(ws, target_date)
    else:
        row = _find_last_data_row(ws)
    if row is None:
        return None
    return _build_dict(ws, row)


def read_platform_prev(ws, target_date=None):
    if target_date:
        current = _find_data_row_by_date(ws, target_date)
        if current is None:
            return None
        # Try previous row with data
        for row_idx in range(current - 1, 5, -1):
            date_val = ws.cell(row=row_idx, column=1).value
            ftd_val = ws.cell(row=row_idx, column=8).value
            if date_val is not None and hasattr(date_val, 'strftime') and ftd_val is not None:
                try:
                    float(ftd_val)
                    return _build_dict(ws, row_idx)
                except (ValueError, TypeError):
                    continue
        return None
    else:
        latest = _find_last_data_row(ws)
        if latest is None:
            return None
        for row_idx in range(latest - 1, 5, -1):
            date_val = ws.cell(row=row_idx, column=1).value
            ftd_val = ws.cell(row=row_idx, column=8).value
            if date_val is not None and hasattr(date_val, 'strftime') and ftd_val is not None:
                try:
                    float(ftd_val)
                    return _build_dict(ws, row_idx)
                except (ValueError, TypeError):
                    continue
        return None


def read_platform_monthly(ws):
    """Sum all daily rows for month-to-date cumulative totals."""
    latest = _find_last_data_row(ws)
    if latest is None:
        return None
    c = lambda row, col: _cell_float(ws, row, col)
    total_reg = 0
    total_ftd = 0
    total_dps_ppl = 0
    total_ftd_amt = 0.0
    total_wdr_ppl = 0
    for row_idx in range(6, latest + 1):
        d = ws.cell(row=row_idx, column=1).value
        ftd = ws.cell(row=row_idx, column=8).value
        if d is None or not hasattr(d, 'strftime') or ftd is None:
            continue
        try:
            float(ftd)
        except (ValueError, TypeError):
            continue
        total_reg += int(c(row_idx, 7))
        total_ftd += int(c(row_idx, 8))
        total_dps_ppl += int(c(row_idx, 11))
        total_ftd_amt += c(row_idx, 20)
        total_wdr_ppl += int(c(row_idx, 12))
    latest_d = _build_dict(ws, latest)
    latest_d["register"] = total_reg
    latest_d["ftd"] = total_ftd
    latest_d["dps_ppl"] = total_dps_ppl
    latest_d["ftd_amount"] = total_ftd_amt
    latest_d["wdr_ppl"] = total_wdr_ppl
    if total_ftd > 0:
        latest_d["new_cust_avg"] = total_ftd_amt / total_ftd
    return latest_d


def read_platform_monthly_full(ws):
    """Sum all daily rows for month-to-date, returning full 28-column data."""
    latest = _find_last_data_row(ws)
    if latest is None:
        return None
    c = lambda row, col: _cell_float(ws, row, col)
    # Summable columns
    s_reg = 0
    s_ftd = 0
    s_level1_ftd = 0
    s_dps_ppl = 0
    s_wdr_ppl = 0
    s_level1_dps = 0.0
    s_level1_wdr = 0.0
    s_ftd_amt = 0.0
    s_new_dps = 0.0
    s_new_wdr = 0.0
    s_daily_dps = 0.0
    s_daily_wdr = 0.0
    s_diff = 0.0
    s_valid_bets = 0.0
    s_invest = 0.0
    s_office_pay = 0.0
    s_online_pay = 0.0
    s_field_to_online = 0.0
    for row_idx in range(6, latest + 1):
        d = ws.cell(row=row_idx, column=1).value
        ftd = ws.cell(row=row_idx, column=8).value
        if d is None or not hasattr(d, 'strftime') or ftd is None:
            continue
        try:
            float(ftd)
        except (ValueError, TypeError):
            continue
        s_reg += int(c(row_idx, 7))
        s_ftd += int(c(row_idx, 8))
        s_level1_ftd += int(c(row_idx, 10))
        s_dps_ppl += int(c(row_idx, 11))
        s_wdr_ppl += int(c(row_idx, 12))
        s_invest += c(row_idx, 5)
        s_valid_bets += c(row_idx, 17)
        s_level1_dps += c(row_idx, 18)
        s_level1_wdr += c(row_idx, 19)
        s_ftd_amt += c(row_idx, 20)
        s_new_dps += c(row_idx, 21)
        s_new_wdr += c(row_idx, 22)
        s_daily_dps += c(row_idx, 23)
        s_daily_wdr += c(row_idx, 24)
        s_diff += c(row_idx, 25)
        s_office_pay += c(row_idx, 2)
        s_online_pay += c(row_idx, 4)
        s_field_to_online += c(row_idx, 3)
    # Derived rates from sums
    conv = s_ftd / s_reg if s_reg > 0 else 0
    reg_cost = s_invest / s_reg if s_reg > 0 else 0
    ftd_cost = s_invest / s_ftd if s_ftd > 0 else 0
    new_cust_avg = s_ftd_amt / s_ftd if s_ftd > 0 else 0
    avg_order = s_daily_dps / s_dps_ppl if s_dps_ppl > 0 else 0
    per_capita_dev = s_ftd / c(latest, 30) if c(latest, 30) > 0 else 0
    # Cost & return ratios
    total_cost = s_office_pay + s_online_pay + s_invest
    roas = s_diff / total_cost if total_cost > 0 else 0
    roi = (s_daily_dps - s_daily_wdr) / total_cost if total_cost > 0 else 0
    return {
        "date": ws.cell(row=latest, column=1).value,
        "register": s_reg,
        "ftd": s_ftd,
        "dps_ppl": s_dps_ppl,
        "wdr_ppl": s_wdr_ppl,
        "ftd_amount": s_ftd_amt,
        "total_dps_amt": s_daily_dps,
        "total_wdr_amt": s_daily_wdr,
        "diff": s_diff,
        "new_cust_avg": new_cust_avg,
        "roi": roi,
        "office_staff": s_office_pay,
        "online_staff": s_online_pay,
        "invest_cost": s_invest,
        "per_capita_dev": per_capita_dev,
        "conversion_rate": conv,
        "level1_ftd": s_level1_ftd,
        "retained_ppl": s_wdr_ppl,
        "reg_cost": reg_cost,
        "ftd_cost": ftd_cost,
        "avg_order_value": avg_order,
        "valid_bets": s_valid_bets,
        "level1_dps": s_level1_dps,
        "level1_wdr": s_level1_wdr,
        "daily_dps": s_new_dps,
        "daily_wdr": s_new_wdr,
        "cumul_dps": s_daily_dps,
        "cumul_wdr": s_daily_wdr,
        "total_ppl": int(c(latest, 30)),
        "roas": roas,
        "turnover": c(latest, 35),
    }


def read_headcount(wb):
    ws = wb[wb.sheetnames[0]]
    hc = {}
    for row_idx in range(4, ws.max_row + 1):
        platform = ws.cell(row=row_idx, column=2).value
        if platform is None:
            continue
        if isinstance(platform, str) and platform.strip():
            office = ws.cell(row=row_idx, column=3).value or 0
            online = ws.cell(row=row_idx, column=5).value or 0
            try:
                office = int(float(office))
            except (ValueError, TypeError):
                office = 0
            try:
                online = int(float(online))
            except (ValueError, TypeError):
                online = 0
            hc[platform.strip()] = {"office": office, "online": online}
    return hc


# ── Ground Push Daily Summary ──────────────────────────────────────

def read_ground_push_daily(wb):
    """Read today's ground push data from 当日汇总 sheet.

    Returns dict {site_name: data_dict} for all 10 platforms with all 28 columns.
    """
    # Find 当日汇总 sheet (index 3, may have garbled name)
    target_sn = None
    for sn in wb.sheetnames:
        if sn == "当日汇总" or "汇总" in sn:
            try:
                ws_test = wb[sn]
                hdr = ws_test.cell(row=4, column=1).value
                if hdr and "DATE" in str(hdr).upper():
                    target_sn = sn
                    break
            except Exception:
                continue
    if target_sn is None and len(wb.sheetnames) > 3:
        target_sn = wb.sheetnames[3]
    if target_sn is None:
        return None

    ws = wb[target_sn]
    site_rows = {
        "PH09": 5, "PH09-2": 6, "PH25": 7, "PH18": 8, "PH30": 9, "PH05": 10, "PH16": 11,
        "BD02": 17, "BD05": 18,
        "MM01": 23,
    }
    today = {}
    for name, row in site_rows.items():
        c = lambda col: _cell_float(ws, row, col)
        ftd = int(c(7))
        reg = int(c(6))
        dps_ppl = int(c(10))
        wdr_ppl = int(c(11))
        ftd_amt = c(19)
        daily_dps = c(20)
        daily_wdr = c(21)
        cumul_dps = c(22)
        cumul_wdr = c(23)
        diff = c(24)
        roi_raw = c(27)
        if roi_raw is None:
            roi_raw = 0
        roas_raw = c(26)
        if roas_raw is None:
            roas_raw = 0
        new_cust_avg = c(14) if c(14) else (ftd_amt / ftd if ftd > 0 else 0)
        d = {
            "date": None,
            "register": reg,
            "ftd": ftd,
            "dps_ppl": dps_ppl,
            "wdr_ppl": wdr_ppl,
            "ftd_amount": ftd_amt,
            "total_dps_amt": cumul_dps,
            "total_wdr_amt": cumul_wdr,
            "diff": diff,
            "new_cust_avg": new_cust_avg,
            "roi": roi_raw,
            # New 28-column fields
            "office_staff": c(2),
            "online_staff": c(3),
            "invest_cost": c(4),
            "per_capita_dev": c(5),
            "conversion_rate": c(8),
            "level1_ftd": int(c(9)),
            "retained_ppl": int(c(11)),
            "reg_cost": c(12),
            "ftd_cost": c(13),
            "avg_order_value": c(15),
            "valid_bets": c(16),
            "level1_dps": c(17),
            "level1_wdr": c(18),
            "daily_dps": daily_dps,
            "daily_wdr": daily_wdr,
            "cumul_dps": cumul_dps,
            "cumul_wdr": cumul_wdr,
            "total_ppl": int(c(25)),
            "roas": roas_raw,
            "turnover": c(28),
        }
        d["status"] = evaluate_status(d)
        d["fraud_risks"] = evaluate_fraud_risks(d)
        today[name] = d
    return today


# ── Price Summary ───────────────────────────────────────────────────

def read_price_summary(wb):
    """Read 单价汇总 sheet, return formatted text with ceiled values."""
    # Find 单价汇总 sheet
    target = None
    for sn in wb.sheetnames:
        if "单价" in sn or "單價" in sn:
            target = sn
            break
    if target is None:
        return None
    import math
    ws = wb[target]
    c2 = lambda row: ws.cell(row=row, column=2).value

    day = c2(2) or ""
    lines = [
        "单价汇总",
        f"2026/05/{day}",
        "",
        "线上办公部门",
        f"内部编制: {c2(5)}",
        f"菲律宾总编制: {c2(6)}",
        f"孟加拉总编制: {c2(7)}",
        "─" * 40,
        "菲律宾",
        "",
    ]

    for name, row in [("PH09", 12), ("PH09-2", 16), ("PH25", 20),
                       ("PH18", 24), ("PH30", 28), ("PH05", 32), ("PH16", 36)]:
        ftd = c2(row + 1) or 0
        price = c2(row + 2) or 0
        try:
            price = math.ceil(float(price))
        except (ValueError, TypeError):
            price = 0
        lines.append(name)
        lines.append(f"首存人数: {ftd}")
        lines.append(f"新客单价: {price}")
        lines.append("")

    lines.append("─" * 40)
    lines.append("孟加拉")
    lines.append("")

    for name, row in [("BD02", 42), ("BD05", 46)]:
        ftd = c2(row + 1) or 0
        price = c2(row + 2) or 0
        try:
            price = math.ceil(float(price))
        except (ValueError, TypeError):
            price = 0
        lines.append(name)
        lines.append(f"首存人数: {ftd}")
        lines.append(f"新客单价: {price}")
        lines.append("")

    lines.append("─" * 40)
    lines.append(f"菲律宾首存人数: {c2(51)}")
    lines.append(f"孟加拉首存人数: {c2(52)}")

    return "\n".join(lines)


# ── HJ Office Data ─────────────────────────────────────────────────

def read_hj_office(wb):
    """Read today's HJ office data from PH33 sheet latest daily row.

    PH33 columns (row 5 English headers, 1-indexed):
     1:DATE  2:Company Staff  3:TOTAL CALL WFH Staff  4:AVERAGE FTD
     5:ON-SITE STAFF SALARIES  6:WFH STAFF SALARY  7:PAYMOUNT AMOUNT
     8:INVERSMENT COST  9:Data consumption  10:total register
     11:TOTAL FTD  12:TOTAL DEPOSITOR  13:REDEPOSITOR  14:RECALL COST
     15:New player ARPU  16:1st deposit AMOUNT  17:ftd total 1st AMOUNT
     18:TOTAL WDR AMOUNT  19:DPSWDR DIFFERENCE
     20:daily/accumulative total deposit  21:daily/accumulative withdraw
     22:retained player total deposit  23:retained player ARPU
     24:daily/accumulative dep - wtd  25:new player profit & lose
     26:daily/accumulative PROFIT & LOSS  27:conversion rate
     28:hijack ratio  29:new player roi  30:daily/accumulative ROI
    """
    if "PH33" not in wb.sheetnames:
        return None
    ws = wb["PH33"]
    # Find latest row with FTD data (col 11)
    latest_row = None
    for row_idx in range(ws.max_row, 5, -1):
        ftd_val = ws.cell(row=row_idx, column=11).value
        if ftd_val is not None:
            try:
                if float(ftd_val) != 0:
                    latest_row = row_idx
                    break
            except (ValueError, TypeError):
                continue
    if latest_row is None:
        return None
    c = lambda col: _cell_float(ws, latest_row, col)
    date_val = ws.cell(row=latest_row, column=1).value
    field_staff = int(c(2))
    online_staff = int(c(3))
    total_staff = field_staff + online_staff
    ftd_count = int(c(11))
    reg_count = int(c(10))
    dps_ppl = int(c(12))
    retained_ppl = int(c(13))
    total_dps = c(20)
    total_wdr = c(21)
    diff = c(24)
    total_salary = int(c(5)) + int(c(6))
    topup_fee = total_dps * 0.035
    platform_fee = diff * 0.20 if diff > 0 else 0
    total_cost = total_salary + topup_fee + platform_fee
    return {
        "platform": "PH33",
        "date": date_val,
        "field_staff": field_staff,
        "field_to_online": 0,
        "online_staff": online_staff,
        "total_staff": total_staff,
        "avg_dev": ftd_count / total_staff if total_staff > 0 else 0,
        "registrations": reg_count,
        "ftd": ftd_count,
        "ftd_amount": c(16),
        "total_ftd_amount": c(17),
        "total_first_wdr": c(18),
        "new_cust_dps": c(17),
        "new_cust_wdr": c(18),
        "new_cust_diff": c(19),
        "level1_ftd": int(c(13)),
        "level1_dps": c(22),
        "level1_wdr": 0,
        "dps_ppl": dps_ppl,
        "retained_ppl": retained_ppl,
        "total_dps": total_dps,
        "total_wdr": total_wdr,
        "diff": diff,
        "online_salary": c(6),
        "platform_pl": c(25),
        "topup_fee": topup_fee,
        "platform_fee": platform_fee,
        "field_salary": c(5),
        "field_to_online_cost": 0,
        "net_profit": c(26),
        "new_cust_avg": c(15),
        "avg_cust_value": c(23),
        "expense_roi": c(29),
        "roi": c(30),
        "conversion_rate": c(27),
        "total_cost": total_cost,
        "total_salary": total_salary,
    }


def read_hj_daily_summary(wb):
    """Read today's hijack data from 当天数据汇总 sheet."""
    if "当天数据汇总" not in wb.sheetnames:
        return None
    ws = wb["当天数据汇总"]
    c = lambda col: _cell_float(ws, 4, col)
    return {
        "platform": str(ws.cell(row=4, column=1).value or "PH33").strip(),
        "team_leaders": int(c(2)),
        "online_staff": int(c(3)),
        "total_staff": int(c(2)) + int(c(3)),
        "data_consumption": c(4),
        "registrations": int(c(5)),
        "ftd": int(c(6)),
        "avg_monthly_ftd": c(7),
        "avg_hijack_cost": c(8),
        "depositors": int(c(9)),
        "loyal_customers": int(c(10)),
        "first_deposit_amount": c(11),
        "total_first_deposit": c(12),
        "total_first_withdrawal": c(13),
        "new_cust_diff": c(14),
        "cumulative_deposits": c(15),
        "cumulative_withdrawals": c(16),
        "cumulative_diff": c(17),
        "est_online_salary": c(18),
        "est_leader_salary": c(19),
        "daily_platform_pl": c(20),
        "daily_topup_fee": c(21),
        "daily_platform_fee": c(22),
        "cumulative_platform_pl": c(23),
        "cumulative_topup_fee": c(24),
        "cumulative_platform_fee": c(25),
        "net_profit": c(26),
        "new_cust_arpu": c(27),
        "loyal_cust_arpu": c(28),
        "conversion_rate": c(29),
        "hijack_rate": c(30),
        "daily_roi": c(31),
        "cumulative_roi": c(32),
    }


def read_hj_monthly_summary(wb):
    """Read monthly cumulative hijack data from 当月数据汇总 sheet (May row)."""
    if "当月数据汇总" not in wb.sheetnames:
        return None
    ws = wb["当月数据汇总"]
    c = lambda col: _cell_float(ws, 5, col)
    return {
        "platform": str(ws.cell(row=5, column=2).value or "PH33").strip(),
        "team_leaders": int(c(3)),
        "online_staff": int(c(4)),
        "total_staff": int(c(3)) + int(c(4)),
        "data_consumption": c(5),
        "registrations": int(c(6)),
        "ftd": int(c(7)),
        "avg_monthly_ftd": c(8),
        "avg_hijack_cost": c(9),
        "depositors": int(c(10)),
        "loyal_customers": int(c(11)),
        "first_deposit_amount": c(12),
        "total_first_deposit": c(13),
        "total_first_withdrawal": c(14),
        "new_cust_diff": c(15),
        "cumulative_deposits": c(16),
        "cumulative_withdrawals": c(17),
        "cumulative_diff": c(18),
        "est_training_salary": c(19),
        "est_online_salary": c(20),
        "daily_platform_pl": c(21),
        "daily_topup_fee": c(22),
        "daily_platform_fee": c(23),
        "cumulative_platform_pl": c(24),
        "cumulative_topup_fee": c(25),
        "cumulative_platform_fee": c(26),
        "net_profit": c(27),
        "new_cust_arpu": c(28),
        "loyal_cust_arpu": c(29),
    }


DAILY_HJ_HEADERS = [
    "平台", "培训组长", "线上人员", "数据消耗", "注册数量", "FTD",
    "月均FTD", "平均劫持成本", "充值数量", "老客数量",
    "首存金额", "首存总金额", "首提总金额", "新客存提差",
    "累计存款", "累计提款", "累计存提差",
    "预线上薪资", "预组长薪资",
    "当天平台盈亏", "当天手续费3.5%", "当天平台费20%",
    "累计平台盈亏", "累计手续费3.5%", "累计平台费20%",
    "净利润", "新客客单价", "老客客单价",
    "转化率", "劫持率", "当天投产比", "累计投产比",
]

MONTHLY_HJ_HEADERS = [
    "平台", "培训组长", "线上人员", "数据消耗", "注册数量", "FTD",
    "月均FTD", "平均劫持成本", "充值数量", "老客数量",
    "首存金额", "首存总金额", "首提总金额", "存提差",
    "累计存款", "累计提款", "累计存提差",
    "预培训薪资", "预线上薪资",
    "当天平台盈亏", "当天手续费3.5%", "当天平台费20%",
    "累计平台盈亏", "累计手续费3.5%", "累计平台费20%",
    "净利润", "新客客单价", "老客客单价",
]


def build_hj_full_row_image(title, headers, values):
    """Render a single-row horizontal table image for hijack data."""
    font_hdr = _get_font(13, bold=True)
    font_body = _get_font(12, bold=False)
    font_title = _get_font(18, bold=True)
    font_footer = _get_font(11, bold=False)

    BG = (255, 255, 255)
    HDR_BG = (45, 62, 80)
    HDR_FG = (255, 255, 255)
    GRID = (222, 226, 230)
    TITLE_FG = (33, 37, 41)
    FOOTER_FG = (140, 140, 140)
    DATA_BG = (249, 251, 253)

    row_h = 30
    title_h = 44
    footer_h = 26

    # Column widths
    col_widths = []
    for ci, h in enumerate(headers):
        w = font_hdr.getbbox(h)[2] + 20
        w = max(w, font_body.getbbox(str(values[ci]))[2] + 20)
        col_widths.append(w)

    total_w = sum(col_widths)
    total_h = title_h + row_h * 2 + footer_h + 6

    img = Image.new("RGB", (total_w, total_h), BG)
    draw = ImageDraw.Draw(img)

    tw_title = font_title.getbbox(title)[2]
    draw.text(((total_w - tw_title) // 2, 10), title, fill=TITLE_FG, font=font_title)

    y = title_h

    # Header row
    x = 0
    for ci, h in enumerate(headers):
        cw = col_widths[ci]
        draw.rectangle([x, y, x + cw - 1, y + row_h], fill=HDR_BG, outline=HDR_BG)
        tw = font_hdr.getbbox(h)[2]
        draw.text((x + (cw - tw) // 2, y + 6), h, fill=HDR_FG, font=font_hdr)
        x += cw
    y += row_h

    # Data row
    x = 0
    for ci, v in enumerate(values):
        cw = col_widths[ci]
        draw.rectangle([x, y, x + cw - 1, y + row_h], fill=DATA_BG, outline=GRID)
        tw = font_body.getbbox(str(v))[2]
        draw.text((x + (cw - tw) // 2, y + 6), str(v), fill=TITLE_FG, font=font_body)
        x += cw

    footer_text = f"@WFHDPbot | {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    draw.text((14, y + row_h + 5), footer_text, fill=FOOTER_FG, font=font_footer)

    import io
    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=True)
    return buf.getvalue()


def build_hj_daily_full_image(daily_data, date_str):
    """Render ALL daily hijack fields (32 cols) in one row."""
    if daily_data is None:
        return None
    def _f(v):
        if v is None: return "0"
        if isinstance(v, float):
            if abs(v) >= 1000: return f"{v:,.0f}"
            if abs(v) >= 100: return f"{v:,.2f}" if v != int(v) else f"{v:,.0f}"
            if v != int(v): return f"{v:.2f}"
            return f"{v:,.0f}"
        return str(v)
    values = [
        daily_data["platform"],
        str(daily_data["team_leaders"]), str(daily_data["online_staff"]),
        _f(daily_data["data_consumption"]), str(daily_data["registrations"]),
        str(daily_data["ftd"]), _f(daily_data["avg_monthly_ftd"]),
        _f(daily_data["avg_hijack_cost"]), str(daily_data["depositors"]),
        str(daily_data["loyal_customers"]), _f(daily_data["first_deposit_amount"]),
        _f(daily_data["total_first_deposit"]), _f(daily_data["total_first_withdrawal"]),
        fmt_full_signed(daily_data["new_cust_diff"]), _f(daily_data["cumulative_deposits"]),
        _f(daily_data["cumulative_withdrawals"]), fmt_full_signed(daily_data["cumulative_diff"]),
        _f(daily_data["est_online_salary"]), _f(daily_data["est_leader_salary"]),
        fmt_full_signed(daily_data["daily_platform_pl"]), _f(daily_data["daily_topup_fee"]),
        _f(daily_data["daily_platform_fee"]), fmt_full_signed(daily_data["cumulative_platform_pl"]),
        _f(daily_data["cumulative_topup_fee"]), _f(daily_data["cumulative_platform_fee"]),
        fmt_full_signed(daily_data["net_profit"]), _f(daily_data["new_cust_arpu"]),
        _f(daily_data["loyal_cust_arpu"]), f"{daily_data['conversion_rate']*100:.1f}%",
        f"{daily_data['hijack_rate']*100:.2f}%", f"{daily_data['daily_roi']*100:.1f}%",
        f"{daily_data['cumulative_roi']*100:.1f}%",
    ]
    return build_hj_full_row_image(f"劫持运营 — 当天数据汇总 | {date_str}", DAILY_HJ_HEADERS, values)


def build_hj_monthly_full_image(monthly_data, date_str):
    """Render ALL monthly hijack fields (28 cols) in one row."""
    if monthly_data is None:
        return None
    def _f(v):
        if v is None: return "0"
        if isinstance(v, float):
            if abs(v) >= 1000: return f"{v:,.0f}"
            if abs(v) >= 100: return f"{v:,.2f}" if v != int(v) else f"{v:,.0f}"
            if v != int(v): return f"{v:.2f}"
            return f"{v:,.0f}"
        return str(v)
    values = [
        monthly_data["platform"],
        str(monthly_data["team_leaders"]), str(monthly_data["online_staff"]),
        _f(monthly_data["data_consumption"]), str(monthly_data["registrations"]),
        str(monthly_data["ftd"]), _f(monthly_data["avg_monthly_ftd"]),
        _f(monthly_data["avg_hijack_cost"]), str(monthly_data["depositors"]),
        str(monthly_data["loyal_customers"]), _f(monthly_data["first_deposit_amount"]),
        _f(monthly_data["total_first_deposit"]), _f(monthly_data["total_first_withdrawal"]),
        fmt_full_signed(monthly_data["new_cust_diff"]), _f(monthly_data["cumulative_deposits"]),
        _f(monthly_data["cumulative_withdrawals"]), fmt_full_signed(monthly_data["cumulative_diff"]),
        _f(monthly_data.get("est_training_salary", 0)), _f(monthly_data.get("est_online_salary", 0)),
        fmt_full_signed(monthly_data["daily_platform_pl"]), _f(monthly_data["daily_topup_fee"]),
        _f(monthly_data["daily_platform_fee"]), fmt_full_signed(monthly_data["cumulative_platform_pl"]),
        _f(monthly_data.get("cumulative_topup_fee", 0)), _f(monthly_data.get("cumulative_platform_fee", 0)),
        fmt_full_signed(monthly_data.get("net_profit", 0)), _f(monthly_data.get("new_cust_arpu", 0)),
        _f(monthly_data.get("loyal_cust_arpu", 0)),
    ]
    return build_hj_full_row_image(f"劫持运营 — 当月数据汇总 | 5月", MONTHLY_HJ_HEADERS, values)


def read_hj_office_prev(wb):
    """Read yesterday's HJ office data from PH33 sheet daily rows."""
    if "PH33" not in wb.sheetnames:
        return None
    ws = wb["PH33"]
    today = read_hj_office(wb)
    if today is None:
        return None
    # Find latest row with data, then go one row up for previous day
    latest_row = None
    for row_idx in range(ws.max_row, 5, -1):
        ftd_val = ws.cell(row=row_idx, column=11).value
        if ftd_val is not None:
            try:
                if float(ftd_val) != 0:
                    latest_row = row_idx
                    break
            except (ValueError, TypeError):
                continue
    if latest_row is None or latest_row <= 6:
        return None
    prev_row = latest_row - 1
    c = lambda col: _cell_float(ws, prev_row, col)
    return {
        "date": ws.cell(row=prev_row, column=1).value,
        "registrations": int(c(10)),
        "ftd": int(c(11)),
        "depositors": int(c(12)),
        "new_cust_ftd_amt": c(16),
        "new_cust_total_dps": c(17),
        "new_cust_total_wdr": c(18),
        "new_cust_diff": c(19),
        "cumulative_dps": c(20),
        "cumulative_wdr": c(21),
        "roi": c(30),
        "daily_roi": c(29),
    }


# ── HJ HR Data ────────────────────────────────────────────────────

def read_hj_hr(wb):
    """Read latest HJ HR daily report row."""
    if "DAILY REPORT" not in wb.sheetnames:
        return None
    ws = wb["DAILY REPORT"]
    # Find latest row with data
    latest_row = None
    for row_idx in range(ws.max_row, 4, -1):
        val = ws.cell(row=row_idx, column=3).value
        if val is not None:
            try:
                if float(val) != 0:
                    latest_row = row_idx
                    break
            except (ValueError, TypeError):
                continue
    if latest_row is None:
        return None
    c = lambda col: _cell_float(ws, latest_row, col)
    return {
        "date": ws.cell(row=latest_row, column=1).value,
        "hr_name": str(ws.cell(row=latest_row, column=2).value or "HR").strip(),
        "resumes": int(c(3)),
        "interviews": int(c(4)),
        "passed": int(c(5)),
        "failed": int(c(6)),
        "in_training": int(c(7)),
        "officially_started": int(c(8)),
        "backout": int(c(9)),
    }


def read_hj_hr_prev(wb, today_date):
    """Read previous day's HR data for DoD comparison."""
    if "DAILY REPORT" not in wb.sheetnames:
        return None
    ws = wb["DAILY REPORT"]
    prev_row = None
    for row_idx in range(ws.max_row, 4, -1):
        val = ws.cell(row=row_idx, column=3).value
        d = ws.cell(row=row_idx, column=1).value
        if val is not None and d is not None:
            try:
                if float(val) != 0 and hasattr(d, "strftime") and d.strftime("%Y-%m-%d") != (today_date.strftime("%Y-%m-%d") if hasattr(today_date, "strftime") else ""):
                    prev_row = row_idx
                    break
            except (ValueError, TypeError):
                continue
    if prev_row is None:
        return None
    c = lambda col: _cell_float(ws, prev_row, col)
    return {
        "date": ws.cell(row=prev_row, column=1).value,
        "resumes": int(c(3)),
        "interviews": int(c(4)),
        "passed": int(c(5)),
        "failed": int(c(6)),
        "in_training": int(c(7)),
        "officially_started": int(c(8)),
        "backout": int(c(9)),
    }


def read_hj_hr_monthly(wb):
    """Sum all rows in DAILY REPORT for month-to-date totals."""
    if "DAILY REPORT" not in wb.sheetnames:
        return None
    ws = wb["DAILY REPORT"]
    totals = {"resumes": 0, "interviews": 0, "passed": 0, "failed": 0,
              "in_training": 0, "officially_started": 0, "backout": 0}
    for row_idx in range(5, ws.max_row + 1):
        c = lambda col: _cell_float(ws, row_idx, col)
        resumes = int(c(3))
        if resumes == 0 and c(4) == 0 and c(5) == 0 and c(6) == 0 and c(7) == 0 and c(8) == 0:
            continue
        totals["resumes"] += resumes
        totals["interviews"] += int(c(4))
        totals["passed"] += int(c(5))
        totals["failed"] += int(c(6))
        totals["in_training"] += int(c(7))
        totals["officially_started"] += int(c(8))
        totals["backout"] += int(c(9))
    return totals


def evaluate_status(d):
    if d["ftd"] == 0:
        return "critical"
    if d["wdr_ppl"] > d["dps_ppl"] and d["dps_ppl"] > 0:
        ratio = d["wdr_ppl"] / d["dps_ppl"]
        if ratio > 1.5:
            return "critical"
        return "warning"
    if 0 < d["ftd"] < 10:
        return "warning"
    if d["roi"] < 0:
        return "warning"
    return "ok"


def evaluate_fraud_risks(d):
    risks = []
    reg = d.get("register", 0)
    ftd = d.get("ftd", 0)
    dps = d.get("total_dps_amt", 0)
    wdr = d.get("total_wdr_amt", 0)
    diff = d.get("diff", 0)
    if reg > 0 and ftd > 0:
        conversion = ftd / reg
        if conversion > 0.7:
            risks.append(f"注册转首存率={conversion:.0%}(>70%红线)")
    if dps > 0 and ftd > 0:
        wdr_rate = wdr / dps
        if wdr_rate > 0.9:
            risks.append(f"提款率={wdr_rate:.0%}(>90%红线)")
        diff_ratio = diff / dps
        if diff_ratio < 0.1:
            risks.append(f"存提差占比={diff_ratio:.0%}(<10%红线)")
    return risks


def fmt_k(v):
    if abs(v) >= 1_000_000:
        return f"{v/1_000_000:.1f}M"
    if abs(v) >= 1_000:
        return f"{v/1_000:.0f}K"
    return f"{v:,.0f}"


def fmt_k_signed(v):
    """Format with + or - sign."""
    s = fmt_k(v)
    if v > 0:
        return f"+{s}"
    return s


def fmt_full(v):
    """Format number with commas, no K/M abbreviation."""
    return f"{v:,.0f}"


def fmt_full_signed(v):
    """Format with + or - sign, full number."""
    if v > 0:
        return f"+{v:,.0f}"
    return f"{v:,.0f}"


def visual_width(s):
    """Approximate display width: CJK/emoji = 2, ASCII = 1."""
    w = 0
    for ch in str(s):
        cp = ord(ch)
        if cp > 0x2000:  # broad: CJK, emoji, symbols
            w += 2
        else:
            w += 1
    return w


def pad_cell(val, width):
    """Pad a value to the given visual width."""
    s = str(val)
    vw = visual_width(s)
    return s + " " * max(0, width - vw)


# ── Box table builder ──────────────────────────────────────────────

def h_line(left, mid, right, fill):
    parts = [fill * w for _, w in COLS]
    return left + mid.join(parts) + right


def data_row(row_values):
    return "│" + "│".join(pad_cell(str(v), w) for v, (_, w) in zip(row_values, COLS)) + "│"


def header_row():
    return "│" + "│".join(pad_cell(name, w) for name, w in COLS) + "│"


def build_site_row(name, d, icon):
    roi_str = f"{d['roi']:.1f}" if d['roi'] != 0 else "N/A"
    new_cust = f"{d['new_cust_avg']:,.0f}" if d['new_cust_avg'] else "N/A"
    return data_row([
        name,
        d["register"],
        d["ftd"],
        d["dps_ppl"],
        d["wdr_ppl"],
        fmt_k(d["ftd_amount"]),
        fmt_k(d["total_dps_amt"]),
        fmt_k(d["total_wdr_amt"]),
        fmt_k_signed(d["diff"]),
        new_cust,
        roi_str,
        icon,
    ])


def build_box_table(data_all):
    """Build a single box-drawing table for all sites."""
    lines = []
    lines.append(h_line("┌", "┬", "┐", "─"))
    lines.append(header_row())

    for i, (name, d) in enumerate(data_all.items()):
        lines.append(h_line("├", "┼", "┤", "─"))
        icon = {"critical": "🔴", "warning": "🟡"}.get(d["status"], "🟢")
        lines.append(build_site_row(name, d, icon))

    lines.append(h_line("└", "┴", "┘", "─"))
    return "\n".join(lines)


# ── Telegram sender ────────────────────────────────────────────────

def esc_html(s):
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def send_telegram(text, parse_mode="HTML"):
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    payload = json.dumps({
        "chat_id": TELEGRAM_CHAT_ID,
        "text": text,
        "parse_mode": parse_mode,
        "disable_web_page_preview": True,
    }).encode("utf-8")
    req = urllib.request.Request(url, data=payload, headers={"Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return json.loads(resp.read()).get("ok", False)
    except Exception as e:
        print(f"  [Telegram send error] {e}")
        return False


def send_telegram_parts(parts):
    ok = 0
    for i, text in enumerate(parts):
        if send_telegram(text.strip()):
            ok += 1
        else:
            print(f"  [Telegram] Part {i+1} failed")
    return ok


# ── Image renderer ────────────────────────────────────────────────

def _get_font(size, bold=False):
    path = FONT_BOLD_PATH if bold else FONT_PATH
    return ImageFont.truetype(path, size)


def render_table_image(title, headers, rows, col_widths=None):
    """Render a table as PNG image. Returns bytes."""
    font_hdr = _get_font(14, bold=True)
    font_body = _get_font(13, bold=False)
    font_title = _get_font(18, bold=True)
    font_footer = _get_font(11, bold=False)

    # Detect multi-line headers (contain \n)
    max_header_lines = 1
    for h in headers:
        lines = h.split("\n")
        max_header_lines = max(max_header_lines, len(lines))

    # Calc column widths if not provided
    if col_widths is None:
        col_widths = []
        for ci, h in enumerate(headers):
            h_lines = h.split("\n")
            max_w = max(font_hdr.getbbox(line)[2] for line in h_lines) + 28
            for row in rows:
                tw = font_body.getbbox(str(row[ci]))[2] + 28
                max_w = max(max_w, tw)
            col_widths.append(max_w)

    row_h = 32
    hdr_row_h = 18 * max_header_lines + 8  # Taller header for multi-line
    title_h = 44
    footer_h = 26
    pad_x = 14

    total_w = sum(col_widths)
    total_h = title_h + hdr_row_h + len(rows) * row_h + footer_h + 6

    # Colors
    BG = (255, 255, 255)
    HDR_BG = (45, 62, 80)
    HDR_FG = (255, 255, 255)
    ROW_ODD = (249, 251, 253)
    ROW_EVEN = (255, 255, 255)
    GRID = (222, 226, 230)
    TITLE_FG = (33, 37, 41)
    FOOTER_FG = (140, 140, 140)
    RED_FG = (220, 53, 69)
    GREEN_FG = (40, 167, 69)
    ORANGE_FG = (255, 152, 0)

    img = Image.new("RGB", (total_w, total_h), BG)
    draw = ImageDraw.Draw(img)

    # Title — centered
    tw_title = font_title.getbbox(title)[2]
    draw.text(((total_w - tw_title) // 2, 10), title, fill=TITLE_FG, font=font_title)

    y = title_h

    # Header row — supports multi-line headers
    x = 0
    for ci, h in enumerate(headers):
        cw = col_widths[ci]
        draw.rectangle([x, y, x + cw - 1, y + hdr_row_h], fill=HDR_BG, outline=HDR_BG)
        h_lines = h.split("\n")
        line_h = 18
        start_y = y + (hdr_row_h - len(h_lines) * line_h) // 2
        for li, line in enumerate(h_lines):
            tw = font_hdr.getbbox(line)[2]
            draw.text((x + (cw - tw) // 2, start_y + li * line_h), line, fill=HDR_FG, font=font_hdr)
        x += cw
    y += hdr_row_h

    font_section = _get_font(14, bold=True)
    SECTION_BG = (235, 240, 245)
    SUBTOTAL_BG = (220, 230, 240)
    SECTION_FG = (60, 80, 100)

    # Data rows
    for ri, row in enumerate(rows):
        # Detect section row: only first column has content, rest empty
        has_section = bool(str(row[0]).strip()) and all(not str(c).strip() for c in row[1:])
        # Detect subtotal row: first column contains "小计"
        is_subtotal = "小计" in str(row[0])

        if has_section:
            # Section label row — centered text, no grid
            draw.rectangle([0, y, total_w - 1, y + row_h], fill=SECTION_BG, outline=GRID)
            s = str(row[0])
            tw = font_section.getbbox(s)[2]
            draw.text(((total_w - tw) // 2, y + 6), s, fill=SECTION_FG, font=font_section)
            y += row_h
            continue

        bg = SUBTOTAL_BG if is_subtotal else (ROW_ODD if ri % 2 == 0 else ROW_EVEN)
        x = 0
        for ci, val in enumerate(row):
            cw = col_widths[ci]
            draw.rectangle([x, y, x + cw - 1, y + row_h], fill=bg, outline=GRID)

            s = str(val)
            fg = (33, 37, 41)
            if is_subtotal:
                fg = (40, 60, 90)

            tw = font_body.getbbox(s)[2]
            draw.text((x + (cw - tw) // 2, y + 7), s, fill=fg, font=font_body)
            x += cw
        y += row_h

    # Footer
    footer_text = f"@WFHDPbot | {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    draw.text((pad_x, y + 5), footer_text, fill=FOOTER_FG, font=font_footer)

    # Save to bytes
    import io
    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=True)
    return buf.getvalue()


# ── Telegram photo sender ──────────────────────────────────────────

def send_telegram_photo(image_bytes, caption=""):
    """Send a PNG image to Telegram."""
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendPhoto"
    import io
    buf = io.BytesIO()
    buf.write(image_bytes)
    buf.seek(0)

    # Manually construct multipart
    boundary = "---boundary" + os.urandom(8).hex()
    body = []
    body.append(f"--{boundary}".encode())
    body.append(f'Content-Disposition: form-data; name="chat_id"\r\n\r\n{TELEGRAM_CHAT_ID}'.encode())
    body.append(f"--{boundary}".encode())
    body.append(f'Content-Disposition: form-data; name="caption"\r\n\r\n{caption}'.encode())
    body.append(f"--{boundary}".encode())
    body.append(f'Content-Disposition: form-data; name="photo"; filename="table.png"\r\nContent-Type: image/png\r\n'.encode())
    body.append(buf.read())
    body.append(f"--{boundary}--".encode())
    data = b"\r\n".join(body)

    req = urllib.request.Request(url, data=data)
    for k, v in {
        "Content-Type": f"multipart/form-data; boundary={boundary}",
    }.items():
        req.add_header(k, v)
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            return json.loads(resp.read()).get("ok", False)
    except Exception as e:
        print(f"  [Telegram photo send error] {e}")
        return False


def send_telegram_document(image_bytes, caption="", filename="table.png"):
    """Send a PNG image as Document (lossless) to Telegram."""
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"
    import io
    buf = io.BytesIO()
    buf.write(image_bytes)
    buf.seek(0)

    boundary = "---boundary" + os.urandom(8).hex()
    body = []
    body.append(f"--{boundary}".encode())
    body.append(f'Content-Disposition: form-data; name="chat_id"\r\n\r\n{TELEGRAM_CHAT_ID}'.encode())
    body.append(f"--{boundary}".encode())
    body.append(f'Content-Disposition: form-data; name="caption"\r\n\r\n{caption}'.encode())
    body.append(f"--{boundary}".encode())
    body.append(f'Content-Disposition: form-data; name="document"; filename="{filename}"\r\nContent-Type: image/png\r\n'.encode())
    body.append(buf.read())
    body.append(f"--{boundary}--".encode())
    data = b"\r\n".join(body)

    req = urllib.request.Request(url, data=data)
    for k, v in {
        "Content-Type": f"multipart/form-data; boundary={boundary}",
    }.items():
        req.add_header(k, v)
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            return json.loads(resp.read()).get("ok", False)
    except Exception as e:
        print(f"  [Telegram document send error] {e}")
        return False


# ── Build table images ─────────────────────────────────────────────

def build_ground_push_image(today, latest_date):
    """Render ground push data table as image."""
    headers = ["站点", "注册", "首存", "充值", "提款", "首存金额", "总充值", "总提款", "充提差", "新客单价", "投产比", "状态"]

    rows = []
    for name in ALL_PLATFORMS:
        if name not in today:
            continue
        d = today[name]
        roi_str = f"{d['roi']:.1f}" if d['roi'] != 0 else "N/A"
        new_cust = f"{d['new_cust_avg']:,.0f}" if d['new_cust_avg'] else "N/A"
        icon = {"critical": "🔴", "warning": "🟡"}.get(d["status"], "🟢")
        rows.append([
            name,
            str(d["register"]),
            str(d["ftd"]),
            str(d["dps_ppl"]),
            str(d["wdr_ppl"]),
            fmt_full(d["ftd_amount"]),
            fmt_full(d["total_dps_amt"]),
            fmt_full(d["total_wdr_amt"]),
            fmt_full(d["diff"]),
            new_cust,
            roi_str,
            icon,
        ])

    cfg = load_config()
    t = cfg.get("titles", {}).get("daily", "线上办公数据汇总")
    return render_table_image(f"📊 {t} — {latest_date}", headers, rows)


def build_monthly_image(monthly, latest_date):
    """Render month-to-date cumulative table as image."""
    headers = ["站点", "注册", "首存", "充值", "提款", "首存金额", "总充值", "总提款", "充提差", "新客单价", "投产比", "状态"]

    rows = []
    for name in ALL_PLATFORMS:
        if name not in monthly:
            continue
        d = monthly[name]
        roi_str = f"{d['roi']:.1f}" if d['roi'] != 0 else "N/A"
        new_cust = f"{d['new_cust_avg']:,.0f}" if d['new_cust_avg'] else "N/A"
        icon = {"critical": "🔴", "warning": "🟡"}.get(d["status"], "🟢")
        rows.append([
            name,
            str(d["register"]),
            str(d["ftd"]),
            str(d["dps_ppl"]),
            str(d["wdr_ppl"]),
            fmt_full(d["ftd_amount"]),
            fmt_full(d["total_dps_amt"]),
            fmt_full(d["total_wdr_amt"]),
            fmt_full(d["diff"]),
            new_cust,
            roi_str,
            icon,
        ])

    month_start = f"{latest_date[:-2]}01"
    cfg = load_config()
    t = cfg.get("titles", {}).get("monthly", "当月累计汇总")
    return render_table_image(f"📅 {t} {month_start}至{latest_date[-2:]}", headers, rows)


def build_headcount_image(today, latest_date):
    """Render staff headcount and efficiency table as image."""
    headers = ["站点", "编制", "办公", "远程", "人均FTD", "人均充提差"]
    rows = []
    for name in ALL_PLATFORMS:
        if name not in today:
            continue
        d = today[name]
        office = d.get("office", 0)
        online = d.get("online", 0)
        total_hc = office + online
        if total_hc > 0:
            avg_ftd = f"{d['ftd'] / total_hc:.1f}"
            avg_diff = f"{d['diff'] / total_hc:,.0f}"
        else:
            avg_ftd = "N/A"
            avg_diff = "N/A"
        rows.append([
            name,
            str(total_hc),
            str(office),
            str(online),
            avg_ftd,
            avg_diff,
        ])
    cfg = load_config()
    t = cfg.get("titles", {}).get("headcount", "编制与人效")
    return render_table_image(f"👥 {t} — {latest_date}", headers, rows)


# ── Ground Push 32-Column Mapping ─────────────────────────────────

# 28-column display spec (user-specified headers)
DAILY_28_HEADERS = [
    "日期", "现场人员", "线上人员", "投放费用(U)", "人均开发",
    "注册人数", "首存人数", "转化率%", "一级首存", "充值人数",
    "复存人数", "注册成本(U)", "首存成本(U)", "新客单价", "人均客单价",
    "有效投注", "一级存款", "一级提款", "首存金额", "新客首存总金额",
    "新客首提总金额", "当天存款", "当天提款", "存提差(P)", "总人数",
    "支出\n回报率%", "投资\n回报率%", "打码倍数",
]


def _daily_to_28row(name, d):
    """Format a daily summary dict into 28 display columns."""
    na = "-"

    def pct(v):
        if v is None:
            return na
        return f"{v*100:.1f}%" if abs(v) < 10 else f"{v*100:.0f}%"

    conv = d.get("conversion_rate", 0)
    conv_str = pct(conv) if conv else "0%"
    roas = d.get("roas", 0)
    roi = d.get("roi", 0)

    return [
        name,  # 日期
        fmt_full(d.get("office_staff", 0)),  # 现场人员
        fmt_full(d.get("online_staff", 0)),  # 线上人员
        fmt_full(d.get("invest_cost", 0)),  # 投放费用(U)
        fmt_full(d.get("per_capita_dev", 0)),  # 人均开发
        str(d["register"]),  # 注册人数
        str(d["ftd"]),  # 首存人数
        conv_str,  # 转化率%
        str(d.get("level1_ftd", 0)),  # 一级首存
        str(d["dps_ppl"]),  # 充值人数
        str(d.get("retained_ppl", 0)),  # 复存人数
        fmt_full(d.get("reg_cost", 0)),  # 注册成本(U)
        fmt_full(d.get("ftd_cost", 0)),  # 首存成本(U)
        fmt_full(d["new_cust_avg"]),  # 新客单价
        fmt_full(d.get("avg_order_value", 0)),  # 人均客单价
        fmt_full(d.get("valid_bets", 0)),  # 有效投注
        fmt_full(d.get("level1_dps", 0)),  # 一级存款
        fmt_full(d.get("level1_wdr", 0)),  # 一级提款
        fmt_full(d["ftd_amount"]),  # 首存金额
        fmt_full(d.get("daily_dps", 0)),  # 新客首存总金额
        fmt_full(d.get("daily_wdr", 0)),  # 新客首提总金额
        fmt_full(d.get("cumul_dps", d["total_dps_amt"])),  # 当天存款
        fmt_full(d.get("cumul_wdr", d["total_wdr_amt"])),  # 当天提款
        fmt_full_signed(d["diff"]),  # 存提差(P)
        str(d.get("total_ppl", 0) or (d.get("office", 0) + d.get("online", 0))),  # 总人数
        pct(roas) if roas else na,  # 支出回报率%
        pct(roi) if roi else "0%",  # 投资回报率%
        fmt_full(d.get("turnover", 0)),  # 打码倍数
    ]


def build_daily_full_image(today, latest_date):
    """Single wide image with all 28 columns x 10 platforms."""
    rows = []
    for name in ALL_PLATFORMS:
        if name not in today:
            continue
        rows.append(_daily_to_28row(name, today[name]))
    cfg = load_config()
    t = cfg.get("titles", {}).get("daily", "线上办公数据汇总")
    return render_table_image(f"📊 {t} — {latest_date}", DAILY_28_HEADERS, rows)


def build_monthly_full_image(monthly_full, latest_date):
    """Single wide image with all 28 columns x 10 platforms (monthly cumulative)."""
    if not monthly_full:
        return None
    rows = []
    for name in ALL_PLATFORMS:
        if name not in monthly_full:
            continue
        rows.append(_daily_to_28row(name, monthly_full[name]))
    cfg = load_config()
    t = cfg.get("titles", {}).get("monthly", "当月累计汇总")
    return render_table_image(f"📊 {t} — {latest_date}", DAILY_28_HEADERS, rows)


# ── Exact Excel sheet replicas ─────────────────────────────────────

def _safe_cell(ws, row, col):
    """Read cell value, return 0 for error/None."""
    v = ws.cell(row=row, column=col).value
    if v is None:
        return 0
    if isinstance(v, str) and ('DIV/0' in v or 'VALUE' in v):
        return 0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0


def _fmt_excel(v, is_pct=False):
    """Format a value as it would appear in Excel."""
    if isinstance(v, float):
        if is_pct:
            return f"{v*100:.1f}%" if abs(v) < 10 else f"{v*100:.0f}%"
        if abs(v) >= 1_000_000:
            return f"{v:,.0f}"
        if abs(v) >= 100:
            return f"{v:,.2f}" if v != int(v) else f"{v:,.0f}"
        if abs(v) < 100 and v != int(v):
            return f"{v:.2f}"
        return f"{v:,.0f}"
    return str(v) if v else "0"


def _fmt_excel_signed(v):
    """Format with + sign for positive."""
    s = _fmt_excel(v)
    if isinstance(v, (int, float)) and v > 0:
        return f"+{s}"
    return s


def read_daily_sheet_rows(wb):
    """Read ALL rows from 当日汇总 sheet exactly as displayed.
    Returns list of (label, [28 col strings], is_subtotal).
    """
    # Find 当日汇总 sheet
    target_sn = None
    for sn in wb.sheetnames:
        if sn == "当日汇总" or "汇总" in sn:
            try:
                ws_test = wb[sn]
                hdr = ws_test.cell(row=4, column=1).value
                if hdr and "DATE" in str(hdr).upper():
                    target_sn = sn
                    break
            except Exception:
                continue
    if target_sn is None and len(wb.sheetnames) > 3:
        target_sn = wb.sheetnames[3]
    if target_sn is None:
        return None, None

    ws = wb[target_sn]

    # Read all data rows: PH(5-12), BD(17-19), MM(23-24)
    # Section labels between groups
    rows = []

    # PH section
    rows.append(("section", "菲律宾 05月", None))
    ph_sites = [5, 6, 7, 8, 9, 10, 11]  # PH09 to PH16
    for r in ph_sites:
        name = str(ws.cell(row=r, column=1).value or "")
        vals = [_safe_cell(ws, r, c) for c in range(1, 29)]
        if vals[5] == 0 and vals[6] == 0:  # Skip sites with no register/FTD
            continue
        rows.append(("data", name, vals))
    # PH subtotal (row 12)
    r = 12
    rows.append(("subtotal", "PH小计", [_safe_cell(ws, r, c) for c in range(1, 29)]))

    # BD section
    rows.append(("section", "孟加拉 05月", None))
    for r in [17, 18]:
        name = str(ws.cell(row=r, column=1).value or "")
        vals = [_safe_cell(ws, r, c) for c in range(1, 29)]
        if vals[5] == 0 and vals[6] == 0:
            continue
        rows.append(("data", name, vals))
    # BD subtotal (row 19)
    r = 19
    rows.append(("subtotal", "BD小计", [_safe_cell(ws, r, c) for c in range(1, 29)]))

    # MM section
    rows.append(("section", "缅甸 05月", None))
    r = 23
    name = str(ws.cell(row=r, column=1).value or "")
    vals = [_safe_cell(ws, r, c) for c in range(1, 29)]
    rows.append(("data", name, vals))
    # MM subtotal (row 24)
    r = 24
    rows.append(("subtotal", "MM小计", [_safe_cell(ws, r, c) for c in range(1, 29)]))

    return rows, ws


def _daily_row_to_display(rows):
    """Convert raw daily rows to display format for render_table_image."""
    display_rows = []
    for row_type, label, vals in rows:
        if row_type == "section":
            display_rows.append([label] + [""] * (len(DAILY_28_HEADERS) - 1))
        elif row_type == "data" or row_type == "subtotal":
            rv = [
                label,  # 日期
                _fmt_excel(vals[1]),   # 现场人员
                _fmt_excel(vals[2]),   # 线上人员
                _fmt_excel(vals[3]),   # 投放费用(U)
                _fmt_excel(vals[4]),   # 人均开发
                _fmt_excel(vals[5]),   # 注册人数
                _fmt_excel(vals[6]),   # 首存人数
                _fmt_excel(vals[7], is_pct=True),  # 转化率%
                _fmt_excel(vals[8]),   # 一级首存
                _fmt_excel(vals[9]),   # 充值人数
                _fmt_excel(vals[10]),  # 复存人数
                _fmt_excel(vals[11]),  # 注册成本(U)
                _fmt_excel(vals[12]),  # 首存成本(U)
                _fmt_excel(vals[13]),  # 新客单价
                _fmt_excel(vals[14]),  # 人均客单价
                _fmt_excel(vals[15]),  # 有效投注
                _fmt_excel(vals[16]),  # 一级存款
                _fmt_excel(vals[17]),  # 一级提款
                _fmt_excel(vals[18]),  # 首存金额
                _fmt_excel(vals[19]),  # 新客首存总金额
                _fmt_excel(vals[20]),  # 新客首提总金额
                _fmt_excel(vals[21]),  # 当天存款
                _fmt_excel(vals[22]),  # 当天提款
                _fmt_excel_signed(vals[23]),  # 存提差(P)
                _fmt_excel(vals[24]),  # 总人数
                _fmt_excel(vals[25], is_pct=True),  # 支出回报率%
                _fmt_excel(vals[26], is_pct=True),  # 投资回报率%
                _fmt_excel(vals[27]),  # 打码倍数
            ]
            display_rows.append(rv)
    return display_rows


def build_daily_full_image_from_sheet(rows, latest_date):
    """Render 当日汇总 image exactly matching the Excel sheet."""
    if not rows:
        return None
    display_rows = _daily_row_to_display(rows)
    cfg = load_config()
    t = cfg.get("titles", {}).get("daily", "线上办公数据汇总")
    return render_table_image(f"📊 {t} — {latest_date}", DAILY_28_HEADERS, display_rows)


# ── Monthly cost sheet (33 columns) ───────────────────────────────

MONTHLY_33_HEADERS = [
    "平台", "现场人员", "现场转线上", "投放费用", "人均开发",
    "总注册", "总开发人数", "首存金额", "首存总金额", "首提总金额",
    "新客总存", "新客总提", "新客存提差", "一级首存", "一级存款",
    "一级提款", "充值人数", "复存人数", "存款", "提款",
    "存提差", "线上人员工资", "平台盈亏", "充值手续费", "平台服务费",
    "现场人员工资", "现场转线上费用", "净利润", "新客客单价", "人均客单价",
    "支出回报率", "投资回报率",
]


def read_monthly_sheet_rows(wb):
    """Read ALL rows from monthly cost sheet (Sheet[0]).
    Returns list of (label, [33 col strings], is_subtotal).
    """
    ws = wb[wb.sheetnames[0]]
    rows = []

    # PH section (rows 4-16)
    rows.append(("section", "菲律宾 05月", None))
    for r in [4, 5, 6, 7, 10, 11, 13]:  # PH09,PH09-2,PH25,PH18,PH30,PH05,PH16 (skip inactive 8,9,12)
        name = str(ws.cell(row=r, column=2).value or "")
        vals = [_safe_cell(ws, r, c) for c in range(2, 35)]
        if vals[5] == 0 and vals[6] == 0:  # Skip empty sites
            continue
        rows.append(("data", name, vals))
    # PH subtotal (row 16)
    r = 16
    rows.append(("subtotal", "PH小计", [_safe_cell(ws, r, c) for c in range(2, 35)]))

    # BD section (rows 20-26)
    rows.append(("section", "孟加拉 05月", None))
    for r in [20, 23]:  # BD02, BD05 (skip inactive 19,21,22,24)
        name = str(ws.cell(row=r, column=2).value or "")
        vals = [_safe_cell(ws, r, c) for c in range(2, 35)]
        if vals[5] == 0 and vals[6] == 0:
            continue
        rows.append(("data", name, vals))
    # BD subtotal (row 26)
    r = 26
    rows.append(("subtotal", "BD小计", [_safe_cell(ws, r, c) for c in range(2, 35)]))

    # MM section (rows 29-31)
    rows.append(("section", "缅甸 05月", None))
    r = 29
    name = str(ws.cell(row=r, column=2).value or "")
    vals = [_safe_cell(ws, r, c) for c in range(2, 35)]
    rows.append(("data", name, vals))
    # MM subtotal (row 31)
    r = 31
    rows.append(("subtotal", "MM小计", [_safe_cell(ws, r, c) for c in range(2, 35)]))

    return rows


def _monthly_row_to_display(rows):
    """Convert raw monthly rows to display format."""
    display_rows = []
    for row_type, label, vals in rows:
        if row_type == "section":
            display_rows.append([label] + [""] * (len(MONTHLY_33_HEADERS) - 1))
        elif row_type == "data" or row_type == "subtotal":
            rv = [
                label,  # 平台
                _fmt_excel(vals[1]),   # 现场人员
                _fmt_excel(vals[2]),   # 现场转线上
                _fmt_excel(vals[3]),   # 投放费用
                _fmt_excel(vals[4]),   # 人均开发
                _fmt_excel(vals[5]),   # 总注册
                _fmt_excel(vals[6]),   # 总开发人数
                _fmt_excel(vals[7]),   # 首存金额
                _fmt_excel(vals[8]),   # 首存总金额
                _fmt_excel(vals[9]),   # 首提总金额
                _fmt_excel(vals[10]),  # 新客总存
                _fmt_excel(vals[11]),  # 新客总提
                _fmt_excel_signed(vals[12]),  # 新客存提差
                _fmt_excel(vals[13]),  # 一级首存
                _fmt_excel(vals[14]),  # 一级存款
                _fmt_excel(vals[15]),  # 一级提款
                _fmt_excel(vals[16]),  # 充值人数
                _fmt_excel(vals[17]),  # 复存人数
                _fmt_excel(vals[18]),  # 存款
                _fmt_excel(vals[19]),  # 提款
                _fmt_excel_signed(vals[20]),  # 存提差
                _fmt_excel(vals[21]),  # 线上人员工资
                _fmt_excel_signed(vals[22]),  # 平台盈亏
                _fmt_excel(vals[23]),  # 充值手续费
                _fmt_excel(vals[24]),  # 平台服务费
                _fmt_excel(vals[25]),  # 现场人员工资
                _fmt_excel(vals[26]),  # 现场转线上费用
                _fmt_excel_signed(vals[27]),  # 净利润
                _fmt_excel(vals[28]),  # 新客客单价
                _fmt_excel(vals[29]),  # 人均客单价
                _fmt_excel(vals[30], is_pct=True),  # 支出回报率
                _fmt_excel(vals[31], is_pct=True),  # 投资回报率
            ]
            display_rows.append(rv)
    return display_rows


def build_monthly_full_image_from_sheet(rows, date_label):
    """Render monthly cost sheet image exactly matching the Excel."""
    if not rows:
        return None
    display_rows = _monthly_row_to_display(rows)
    cfg = load_config()
    t = cfg.get("titles", {}).get("monthly", "当月累计汇总")
    return render_table_image(f"📊 {t} — {date_label}", MONTHLY_33_HEADERS, display_rows)


def _build_hj_horizontal(title, hj_office, fields):
    """Build a horizontal table image for HJ office with given fields.

    fields: list of (header, value_str)
    """
    if hj_office is None:
        return None
    headers = [h for h, v in fields]
    rows = [[v for h, v in fields]]
    return render_table_image(title, headers, rows)


def build_hj_staff_dev_image(hj_office, hj_date_str):
    """人员 & 开发指标 (cols 1-8)."""
    fields = [
        ("现场人数", str(hj_office["field_staff"])),
        ("现场转线上", str(hj_office["field_to_online"])),
        ("线上人数", str(hj_office["online_staff"])),
        ("人均开发", f"{hj_office['avg_dev']:,.0f}" if hj_office['avg_dev'] else "N/A"),
        ("总注册", str(hj_office["registrations"])),
        ("FTD", str(hj_office["ftd"])),
        ("转化率", f"{hj_office['conversion_rate']:.1%}"),
        ("首存金额", fmt_full(hj_office["ftd_amount"])),
    ]
    cfg = load_config()
    t = cfg.get("titles", {}).get("hj_staff_dev", "劫持-人员与开发")
    return _build_hj_horizontal(f"🛡️ {t} — {hj_office['platform']} | {hj_date_str}", hj_office, fields)


def build_hj_newcust_level1_image(hj_office, hj_date_str):
    """新客 & 一级数据 (cols 9-17)."""
    fields = [
        ("首存总金额", fmt_full(hj_office["total_ftd_amount"])),
        ("首提总金额", fmt_full(hj_office["total_first_wdr"])),
        ("新客总存", fmt_full(hj_office["new_cust_dps"])),
        ("新客总提", fmt_full(hj_office["new_cust_wdr"])),
        ("新客存提差", fmt_full_signed(hj_office["new_cust_diff"])),
        ("一级首存(人)", str(hj_office["level1_ftd"])),
        ("一级存款", fmt_full(hj_office["level1_dps"])),
        ("一级提款", fmt_full(hj_office["level1_wdr"])),
    ]
    cfg = load_config()
    t = cfg.get("titles", {}).get("hj_newcust", "劫持-新客与一级")
    return _build_hj_horizontal(f"🛡️ {t} — {hj_office['platform']} | {hj_date_str}", hj_office, fields)


def build_hj_ops_cost_image(hj_office, hj_date_str):
    """运营 & 成本 (cols 17-25)."""
    fields = [
        ("充值人数", str(hj_office["dps_ppl"])),
        ("复存人数", str(hj_office["retained_ppl"])),
        ("存款", fmt_full(hj_office["total_dps"])),
        ("提款", fmt_full(hj_office["total_wdr"])),
        ("存提差", fmt_full_signed(hj_office["diff"])),
        ("线上人员工资", fmt_full(hj_office["online_salary"])),
        ("平台盈亏", fmt_full_signed(hj_office["platform_pl"])),
        ("充值手续费(3.5%)", fmt_full(hj_office["topup_fee"])),
    ]
    cfg = load_config()
    t = cfg.get("titles", {}).get("hj_ops_cost", "劫持-运营与成本")
    return _build_hj_horizontal(f"🛡️ {t} — {hj_office['platform']} | {hj_date_str}", hj_office, fields)


def build_hj_profit_roi_image(hj_office, hj_date_str):
    """利润 & 效率 (cols 25-32 + derived)."""
    fields = [
        ("平台服务费(20%)", fmt_full(hj_office["platform_fee"])),
        ("现场人员工资", fmt_full(hj_office["field_salary"])),
        ("现场转线上费用", fmt_full(hj_office["field_to_online_cost"])),
        ("净利润", fmt_full_signed(hj_office["net_profit"])),
        ("新客客单价", f"{hj_office['new_cust_avg']:,.0f}" if hj_office['new_cust_avg'] else "N/A"),
        ("人均客单价", f"{hj_office['avg_cust_value']:.3f}" if hj_office['avg_cust_value'] else "N/A"),
        ("支出回报率", f"{hj_office['expense_roi']*100:.1f}%" if hj_office['expense_roi'] else "0%"),
        ("投产比(ROI)", f"{hj_office['roi']*100:.1f}%"),
    ]
    cfg = load_config()
    t = cfg.get("titles", {}).get("hj_profit_roi", "劫持-利润与效率")
    return _build_hj_horizontal(f"🛡️ {t} — {hj_office['platform']} | {hj_date_str}", hj_office, fields)


def build_hj_combined_image(hj_office, hj_date_str):
    """Combine all 4 HJ office sections into one vertically stacked image."""
    if hj_office is None:
        return None

    cfg = load_config()

    sections = [
        (cfg.get("titles", {}).get("hj_staff_dev", "劫持-人员与开发"), [
            ("现场人数", str(hj_office["field_staff"])),
            ("现场转线上", str(hj_office["field_to_online"])),
            ("线上人数", str(hj_office["online_staff"])),
            ("人均开发", f"{hj_office['avg_dev']:,.0f}" if hj_office['avg_dev'] else "N/A"),
            ("总注册", str(hj_office["registrations"])),
            ("FTD", str(hj_office["ftd"])),
            ("转化率", f"{hj_office['conversion_rate']:.1%}"),
            ("首存金额", fmt_full(hj_office["ftd_amount"])),
        ]),
        (cfg.get("titles", {}).get("hj_newcust", "劫持-新客与一级"), [
            ("首存总金额", fmt_full(hj_office["total_ftd_amount"])),
            ("首提总金额", fmt_full(hj_office["total_first_wdr"])),
            ("新客总存", fmt_full(hj_office["new_cust_dps"])),
            ("新客总提", fmt_full(hj_office["new_cust_wdr"])),
            ("新客存提差", fmt_full_signed(hj_office["new_cust_diff"])),
            ("一级首存(人)", str(hj_office["level1_ftd"])),
            ("一级存款", fmt_full(hj_office["level1_dps"])),
            ("一级提款", fmt_full(hj_office["level1_wdr"])),
        ]),
        (cfg.get("titles", {}).get("hj_ops_cost", "劫持-运营与成本"), [
            ("充值人数", str(hj_office["dps_ppl"])),
            ("复存人数", str(hj_office["retained_ppl"])),
            ("存款(累计)", fmt_full(hj_office["total_dps"])),
            ("提款(累计)", fmt_full(hj_office["total_wdr"])),
            ("存提差", fmt_full_signed(hj_office["diff"])),
            ("线上人员工资", fmt_full(hj_office["online_salary"])),
            ("平台盈亏", fmt_full_signed(hj_office["platform_pl"])),
            ("充值手续费(3.5%)", fmt_full(hj_office["topup_fee"])),
        ]),
        (cfg.get("titles", {}).get("hj_profit_roi", "劫持-利润与效率"), [
            ("平台服务费(20%)", fmt_full(hj_office["platform_fee"])),
            ("现场人员工资", fmt_full(hj_office["field_salary"])),
            ("现场转线上费用", fmt_full(hj_office["field_to_online_cost"])),
            ("净利润", fmt_full_signed(hj_office["net_profit"])),
            ("新客客单价", f"{hj_office['new_cust_avg']:,.0f}" if hj_office['new_cust_avg'] else "N/A"),
            ("人均客单价", f"{hj_office['avg_cust_value']:.3f}" if hj_office['avg_cust_value'] else "N/A"),
            ("支出回报率", f"{hj_office['expense_roi']*100:.1f}%" if hj_office['expense_roi'] else "0%"),
            ("投产比(ROI)", f"{hj_office['roi']*100:.1f}%"),
        ]),
    ]

    font_hdr = _get_font(14, bold=True)
    font_body = _get_font(13, bold=False)
    font_title = _get_font(18, bold=True)
    font_section = _get_font(14, bold=True)
    font_footer = _get_font(11, bold=False)

    BG = (255, 255, 255)
    HDR_BG = (45, 62, 80)
    HDR_FG = (255, 255, 255)
    GRID = (222, 226, 230)
    TITLE_FG = (33, 37, 41)
    FOOTER_FG = (140, 140, 140)
    SECTION_BG = (235, 240, 245)
    SECTION_FG = (60, 80, 100)
    DATA_BG = (249, 251, 253)

    row_h = 32
    section_h = 30
    title_h = 44
    footer_h = 26
    gap_h = 8

    num_cols = max(len(fields) for _, fields in sections)
    col_widths = [0] * num_cols
    for _, fields in sections:
        for ci, (h, v) in enumerate(fields):
            col_widths[ci] = max(col_widths[ci],
                                font_hdr.getbbox(h)[2] + 28,
                                font_body.getbbox(str(v))[2] + 28)

    total_w = sum(col_widths)
    total_h = title_h + 4 * (section_h + row_h + row_h) + 3 * gap_h + footer_h + 6

    img = Image.new("RGB", (total_w, total_h), BG)
    draw = ImageDraw.Draw(img)

    title = f"劫持运营 — {hj_office['platform']} | {hj_date_str}"
    tw_title = font_title.getbbox(title)[2]
    draw.text(((total_w - tw_title) // 2, 10), title, fill=TITLE_FG, font=font_title)

    y = title_h

    for sec_idx, (sec_title, fields) in enumerate(sections):
        draw.rectangle([0, y, total_w - 1, y + section_h], fill=SECTION_BG, outline=GRID)
        sec_txt = f"▸ {sec_title}"
        tw = font_section.getbbox(sec_txt)[2]
        draw.text(((total_w - tw) // 2, y + 5), sec_txt, fill=SECTION_FG, font=font_section)
        y += section_h

        x = 0
        for ci, (h, v) in enumerate(fields):
            cw = col_widths[ci]
            draw.rectangle([x, y, x + cw - 1, y + row_h], fill=HDR_BG, outline=HDR_BG)
            tw = font_hdr.getbbox(h)[2]
            draw.text((x + (cw - tw) // 2, y + 7), h, fill=HDR_FG, font=font_hdr)
            x += cw
        y += row_h

        x = 0
        for ci, (h, v) in enumerate(fields):
            cw = col_widths[ci]
            draw.rectangle([x, y, x + cw - 1, y + row_h], fill=DATA_BG, outline=GRID)
            tw = font_body.getbbox(str(v))[2]
            draw.text((x + (cw - tw) // 2, y + 7), str(v), fill=TITLE_FG, font=font_body)
            x += cw
        y += row_h

        if sec_idx < 3:
            y += gap_h

    footer_text = f"@WFHDPbot | {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    draw.text((14, y + 5), footer_text, fill=FOOTER_FG, font=font_footer)

    import io
    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=True)
    return buf.getvalue()


def build_hj_hr_image(hj_hr, hj_hr_monthly, hr_date_str):
    """Render HJ HR snapshot as horizontal table image."""
    if hj_hr is None:
        return None
    headers = ["", "简历", "面试", "通过", "失败", "培训中", "上岗", "淘汰"]
    rows = [
        ["今日"] + [str(hj_hr[k]) for k in ["resumes", "interviews", "passed", "failed", "in_training", "officially_started", "backout"]],
    ]
    if hj_hr_monthly:
        rows.append(["当月累计"] + [str(hj_hr_monthly[k]) for k in ["resumes", "interviews", "passed", "failed", "in_training", "officially_started", "backout"]])
    cfg = load_config()
    t = cfg.get("titles", {}).get("hijack_hr", "劫持人事汇总")
    return render_table_image(f"👤 {t} — {hr_date_str} | {hj_hr['hr_name']}", headers, rows)


def build_anomaly_image(anomaly_lines):
    """Render anomaly alerts as image."""
    if not anomaly_lines:
        return None
    headers = ["状态", "站点", "FTD", "DPS", "WDR", "异常原因"]
    rows = []
    for line in anomaly_lines:
        # Parse: "🔴 PH09: FTD=0 DPS=5 WDR=8 (FTD归零, 提款>充值)"
        m = re.match(r'(🔴|🟡)\s+([A-Z0-9-]+):\s*FTD=(\d+)\s+DPS=(\d+)\s+WDR=(\d+)\s*\((.+)\)', line)
        if m:
            rows.append([m.group(1), m.group(2), m.group(3), m.group(4), m.group(5), m.group(6)])
        else:
            # Fallback: try looser parsing
            icon_match = re.match(r'(🔴|🟡)', line)
            icon = icon_match.group(1) if icon_match else ""
            rest = line[len(icon_match.group(0)):] if icon_match else line
            name_m = re.match(r'\s*([A-Z0-9-]+):', rest)
            ftd_m = re.search(r'FTD=(\d+)', rest)
            dps_m = re.search(r'DPS=(\d+)', rest)
            wdr_m = re.search(r'WDR=(\d+)', rest)
            reason_m = re.search(r'\((.+)\)', rest)
            rows.append([
                icon,
                name_m.group(1) if name_m else "",
                ftd_m.group(1) if ftd_m else "",
                dps_m.group(1) if dps_m else "",
                wdr_m.group(1) if wdr_m else "",
                reason_m.group(1) if reason_m else rest.strip(),
            ])
    return render_table_image("⚠️ 异常站点", headers, rows)


def build_dod_image(today, yesterday, prev_date, latest_date):
    """Render DoD comparison table as image."""
    if not yesterday:
        return None
    headers = ["站点", "FTD昨日", "FTD今日", "FTD变化", "ROI变化", "充提差变化"]
    rows = []
    for name in ALL_PLATFORMS:
        if name not in today or name not in yesterday:
            continue
        t = today[name]
        y = yesterday[name]
        trend = "↑" if t["ftd"] > y["ftd"] else ("↓" if t["ftd"] < y["ftd"] else "→")
        rows.append([
            name,
            str(y["ftd"]),
            str(t["ftd"]),
            f"{t['ftd'] - y['ftd']:+d} {trend}",
            delta_roi(t["roi"], y["roi"]),
            delta_str(t["diff"], y["diff"]),
        ])
    return render_table_image(f"📈 环比昨日 — {prev_date} → {latest_date}", headers, rows)


# ── Delta helpers ──────────────────────────────────────────────────

def delta_str(curr, prev):
    if prev is None or prev == 0:
        if curr is None or curr == 0:
            return "N/A"
        return "NEW"
    if curr is None:
        return "N/A"
    diff = curr - prev
    if diff > 0:
        return f"+{diff:,.0f}"
    elif diff < 0:
        return f"{diff:,.0f}"
    return "0"


def delta_roi(curr, prev):
    if prev is None or prev == 0:
        if curr is None or curr == 0:
            return "N/A"
        return "NEW"
    if curr is None:
        return "N/A"
    diff = curr - prev
    if diff > 0:
        return f"+{diff:.1f}"
    elif diff < 0:
        return f"{diff:.1f}"
    return "0"


# ── Main ───────────────────────────────────────────────────────────

def generate_push(target_date=None, target_month=None, override_sections=None):
    now = datetime.now()
    timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
    tip = TIPS[now.minute % len(TIPS)]
    cfg = load_config()
    titles = cfg.get("titles", {})
    sections = cfg.get("sections", {})
    if override_sections is not None:
        for k in sections:
            sections[k] = False
        for s in override_sections.split(","):
            s = s.strip()
            if s in sections:
                sections[s] = True
    fmt = cfg.get("format", {})
    if target_month:
        print(f"[{timestamp}] Target month: {target_month}")
    elif target_date:
        print(f"[{timestamp}] Target date: {target_date}")

    # Dynamically find Excel files based on target date/month
    if target_month:
        resolve_date = datetime.strptime(target_month + "-01", "%Y-%m-%d")
        use_last_row = True  # For month mode, use the last data row
    elif target_date:
        resolve_date = datetime.strptime(target_date, "%Y-%m-%d")
        use_last_row = False
    else:
        resolve_date = datetime.now()
        use_last_row = False

    excel_file = find_monthly_file(resolve_date, ["线上办公数据汇"])
    hj_office_file = find_monthly_file(resolve_date, ["劫持", "办公数据汇总"], exclude_kw=["人事"])
    hj_hr_file = find_monthly_file(resolve_date, ["劫持", "人事数据汇总"])

    if not excel_file:
        msg = f"未找到 {resolve_date.strftime('%Y年%m月') if hasattr(resolve_date, 'strftime') else resolve_date[:7]} 的地推数据文件"
        print(f"[{timestamp}] ERROR: {msg}")
        send_telegram(f"❌ {msg}\n请确认文件已放入 新建文件夹")
        return

    print(f"[{timestamp}] Using: {os.path.basename(excel_file)}")
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    headcount = read_headcount(wb)

    today = {}
    yesterday = {}
    monthly = {}
    latest_data_date = None

    # Primary: read today's data from 当日汇总 sheet
    daily_summary = read_ground_push_daily(wb)
    # Read daily sheet rows (exact Excel replica for image)
    daily_sheet_rows, _ = read_daily_sheet_rows(wb)
    # Read monthly cost sheet rows (exact Excel replica)
    monthly_sheet_rows = read_monthly_sheet_rows(wb)
    # Read price summary
    price_summary_text = read_price_summary(wb)
    if daily_summary:
        for name, d in daily_summary.items():
            d["office"] = headcount.get(name, {}).get("office", 0)
            d["online"] = headcount.get(name, {}).get("online", 0)
            today[name] = d
        # Get dates from individual sheets (当日汇总 doesn't have dates)
        for name in ALL_PLATFORMS:
            if name in wb.sheetnames:
                ws = wb[name]
                # Find the latest date from individual sheets
                row = _find_last_data_row(ws)
                if row is not None:
                    dval = ws.cell(row=row, column=1).value
                    if dval is not None and hasattr(dval, 'strftime') and name in today:
                        today[name]["date"] = dval
                        if latest_data_date is None:
                            latest_data_date = dval

    # Read yesterday (prev day) and monthly from individual sheets
    monthly_full = {}
    for name in ALL_PLATFORMS:
        if name in wb.sheetnames:
            ws = wb[name]
            p = read_platform_prev(ws, None if use_last_row else target_date)
            if p:
                yesterday[name] = p
            m = read_platform_monthly(ws)
            if m:
                m["status"] = evaluate_status(m)
                monthly[name] = m
            mf = read_platform_monthly_full(ws)
            if mf:
                mf["office"] = headcount.get(name, {}).get("office", 0)
                mf["online"] = headcount.get(name, {}).get("online", 0)
                monthly_full[name] = mf
    wb.close()

    # ── Read HJ data ─────────────────────────────────────────────
    hj_office = None
    hj_office_prev = None
    hj_hr = None
    hj_hr_prev = None
    hj_hr_monthly = None
    hj_warnings = []

    hj_daily_summary = None
    hj_monthly_summary = None
    if hj_office_file and os.path.exists(hj_office_file):
        print(f"[{timestamp}] Using: {os.path.basename(hj_office_file)}")
        hj_wb = openpyxl.load_workbook(hj_office_file, data_only=True)
        hj_office = read_hj_office(hj_wb)
        hj_office_prev = read_hj_office_prev(hj_wb)
        hj_daily_summary = read_hj_daily_summary(hj_wb)
        hj_monthly_summary = read_hj_monthly_summary(hj_wb)
        hj_wb.close()
    elif resolve_date and hasattr(resolve_date, 'month') and resolve_date.month >= 3:
        hj_warnings.append("⚠️ 劫持运营文件未找到（3月起应有此数据）")

    if hj_hr_file and os.path.exists(hj_hr_file):
        print(f"[{timestamp}] Using: {os.path.basename(hj_hr_file)}")
        hr_wb = openpyxl.load_workbook(hj_hr_file, data_only=True)
        hj_hr = read_hj_hr(hr_wb)
        if hj_hr and hj_hr.get("date"):
            hj_hr_prev = read_hj_hr_prev(hr_wb, hj_hr["date"])
        hj_hr_monthly = read_hj_hr_monthly(hr_wb)
        hr_wb.close()
    elif resolve_date and hasattr(resolve_date, 'month') and resolve_date.month >= 3:
        hj_warnings.append("⚠️ 劫持人资文件未找到（3月起应有此数据）")

    # Latest date / date range
    dates = [d["date"] for d in today.values() if d.get("date")]
    latest_date = "未知"
    date_label = "数据日期"
    if dates:
        d0 = dates[0]
        latest_date = d0.strftime("%Y-%m-%d") if hasattr(d0, 'strftime') else str(d0)[:10]
    if target_month:
        # Show month range: e.g. "2026-04 (截至04-28)"
        date_label = "数据月份"
        if latest_date != "未知":
            latest_date = f"{target_month} (截至{latest_date[-5:]})"

    prev_date = ""
    if yesterday:
        pd_dates = [d["date"] for d in yesterday.values() if d.get("date")]
        if pd_dates:
            pd0 = pd_dates[0]
            prev_date = pd0.strftime("%Y-%m-%d") if hasattr(pd0, 'strftime') else str(pd0)[:10]

    # Region totals
    def sum_field(data, field):
        return int(sum(d[field] for d in data.values()))

    for label, plats in [("ph", PH_PLATFORMS), ("bd", BD_PLATFORMS), ("mm", MM_PLATFORMS)]:
        pass  # computed below

    ph_d = {n: today[n] for n in PH_PLATFORMS if n in today}
    bd_d = {n: today[n] for n in BD_PLATFORMS if n in today}
    mm_d = {n: today[n] for n in MM_PLATFORMS if n in today}

    ph_ftd = sum_field(ph_d, "ftd")
    ph_ftd_amt = sum(d["ftd_amount"] for d in ph_d.values())
    ph_dps_amt = sum(d["total_dps_amt"] for d in ph_d.values())
    ph_wdr_amt = sum(d["total_wdr_amt"] for d in ph_d.values())
    ph_diff = sum(d["diff"] for d in ph_d.values())

    bd_ftd = sum_field(bd_d, "ftd")
    bd_ftd_amt = sum(d["ftd_amount"] for d in bd_d.values())
    bd_dps_amt = sum(d["total_dps_amt"] for d in bd_d.values())
    bd_wdr_amt = sum(d["total_wdr_amt"] for d in bd_d.values())
    bd_diff = sum(d["diff"] for d in bd_d.values())

    mm_ftd = sum_field(mm_d, "ftd")
    mm_ftd_amt = sum(d["ftd_amount"] for d in mm_d.values())
    mm_dps_amt = sum(d["total_dps_amt"] for d in mm_d.values())
    mm_wdr_amt = sum(d["total_wdr_amt"] for d in mm_d.values())
    mm_diff = sum(d["diff"] for d in mm_d.values())

    total_ftd = ph_ftd + bd_ftd + mm_ftd
    total_ftd_amt = ph_ftd_amt + bd_ftd_amt + mm_ftd_amt
    total_dps_amt = ph_dps_amt + bd_dps_amt + mm_dps_amt
    total_wdr_amt = ph_wdr_amt + bd_wdr_amt + mm_wdr_amt
    total_diff = ph_diff + bd_diff + mm_diff

    # ── Build all-sites box table ──
    box_table = build_box_table(today)

    # ── Region summary lines ──
    summary_ph = f"菲区首存{ph_ftd}人/首存{fmt_k(ph_ftd_amt)}/充提差{fmt_k_signed(ph_diff)}"
    summary_bd = f"孟区首存{bd_ftd}人/充提差{fmt_k_signed(bd_diff)}"
    summary_mm = f"缅区首存{mm_ftd}人/充提差{fmt_k_signed(mm_diff)}"
    summary_all = f"整体充提差{fmt_k_signed(total_diff)}"

    # ── Anomaly sites (ALL 10 sites) ──
    anomaly_lines = []
    for name in ALL_PLATFORMS:
        if name not in today:
            continue
        d = today[name]
        icon = {"critical": "🔴", "warning": "🟡"}.get(d["status"], "🟢")
        reasons = []
        if d["ftd"] == 0:
            reasons.append("FTD归零")
        if d["wdr_ppl"] > d["dps_ppl"]:
            reasons.append(f"提款>充值(DIFF={d['diff']:,.0f})")
        if 0 < d["ftd"] < 10:
            reasons.append("FTD个位数")
        if d["roi"] < 0:
            reasons.append(f"ROI负({d['roi']:.1f})")
        if not reasons:
            reasons.append("正常")
        anomaly_lines.append(f"{icon} {name}: FTD={d['ftd']} DPS={d['dps_ppl']} WDR={d['wdr_ppl']} ({', '.join(reasons)})")

    # ── Fraud alerts (ALL 10 sites) ──
    fraud_lines = []
    for name in ALL_PLATFORMS:
        if name not in today:
            continue
        risks = today[name].get("fraud_risks", [])
        if risks:
            for risk in risks:
                fraud_lines.append(f"🚨 {name}: {risk}")
        else:
            fraud_lines.append(f"🟢 {name}: 正常")

    # ── DoD comparison ──
    dod_lines = []
    if yesterday:
        dod_lines.append(f"环比昨日 ({prev_date} → {latest_date}):")
        for name in ALL_PLATFORMS:
            if name in today:
                t = today[name]
                y = yesterday.get(name)
                if y:
                    f_d = delta_str(t["ftd"], y["ftd"])
                    r_d = delta_roi(t["roi"], y["roi"])
                    diff_d = delta_str(t["diff"], y["diff"])
                    trend = "↑" if (t["ftd"] - y["ftd"]) > 0 else ("↓" if (t["ftd"] - y["ftd"]) < 0 else "→")
                    dod_lines.append(f"  {name}: FTD {y['ftd']}→{t['ftd']}({trend}) ROI {r_d} 差 {diff_d}")

    # ── Headcount & Efficiency ──
    hc_lines = []
    hc_lines.append("| 站点 | 编制 | 办公 | 远程 | 人均FTD | 人均充提差 |")
    hc_lines.append("|------|------|------|------|--------|-----------|")
    for name in ALL_PLATFORMS:
        if name in today:
            d = today[name]
            office = d.get("office", 0)
            online = d.get("online", 0)
            total_hc = office + online
            if total_hc > 0:
                avg_ftd = f"{d['ftd'] / total_hc:.1f}"
                avg_diff = f"{d['diff'] / total_hc:,.0f}"
            else:
                avg_ftd = "N/A"
                avg_diff = "N/A"
            hc_lines.append(f"| {name} | {total_hc} | {office} | {online} | {avg_ftd} | {avg_diff} |")

    # ── Assemble content ──
    content = f"""# 线上办公数据汇总
**推送时间**: {timestamp}
**数据日期**: {latest_date}
**下次推送**: 明日21:07

---

```
{box_table}
```

> {summary_ph}，{summary_bd}，{summary_mm}。{summary_all}。

## 编制与人效
{chr(10).join(hc_lines)}
"""

    if dod_lines:
        content += "\n## 环比昨日\n"
        content += "\n".join(dod_lines) + "\n"

    # ── Per-site tips ──
    tip_lines = []
    for name in ALL_PLATFORMS:
        if name not in today:
            continue
        d = today[name]
        tips = []
        if d["ftd"] == 0:
            tips.append("FTD归零，立刻联系了解原因")
        if 0 < d["ftd"] < 10:
            tips.append("FTD个位数，关注渠道质量")
        if d["roi"] < 0:
            tips.append("ROI为负，控制成本或提升转化")
        if d.get("wdr_ppl", 0) > d.get("dps_ppl", 0):
            tips.append("提款>充值，检查套利风险")
        if d.get("register", 0) > 0 and d["ftd"] > 0:
            conv = d["ftd"] / max(d["register"], 1)
            if conv < 0.1:
                tips.append(f"注册转化率仅{conv:.0%}，优化注册渠道")
        if not tips:
            tips.append("数据正常，保持运营节奏")
        tip_lines.append(f"  {name}: {'; '.join(tips)}")
    content += f"\n## 提醒操作\n{chr(10).join(tip_lines)}\n"

    content += "\n## 异常站点\n"
    for line in anomaly_lines:
        content += f"- {line}\n"

    content += "\n## 防作弊风控\n"
    for line in fraud_lines:
        content += f"- {line}\n"

    # ── HJ Office section (image-based, text summary only) ──
    if hj_daily_summary:
        content += "\n## 劫持运营\n"
        content += f"**{hj_daily_summary['platform']} | {latest_date}** — 当天FTD={hj_daily_summary['ftd']} 注册={hj_daily_summary['registrations']} 存提差={fmt_k_signed(hj_daily_summary['cumulative_diff'])} 净利润={fmt_k_signed(hj_daily_summary['net_profit'])}\n"
        content += f"> 详情见劫持运营数据图\n"

    # ── HJ HR section ──
    if hj_hr:
        content += "\n## 劫持人事汇总\n"
        hr_date = hj_hr["date"]
        hr_date_str = hr_date.strftime("%Y-%m-%d") if hasattr(hr_date, 'strftime') else str(hr_date)[:10]
        content += f"**{hj_hr['hr_name']} | {hr_date_str}**\n\n"
        content += f"| 简历 | 面试 | 通过 | 失败 | 培训中 | 上岗 | 淘汰 |\n"
        content += f"|------|------|------|------|--------|------|------|\n"
        content += f"| {hj_hr['resumes']} | {hj_hr['interviews']} | {hj_hr['passed']} | {hj_hr['failed']} | {hj_hr['in_training']} | {hj_hr['officially_started']} | {hj_hr['backout']} |\n"
        if hj_hr_prev:
            content += f"\n环比昨日 简历 {hj_hr_prev['resumes']}→{hj_hr['resumes']} | 面试 {hj_hr_prev['interviews']}→{hj_hr['interviews']}\n"
        if hj_hr_monthly:
            content += f"\n当月累计 简历{hj_hr_monthly['resumes']} 面试{hj_hr_monthly['interviews']} 通过{hj_hr_monthly['passed']} 上岗{hj_hr_monthly['officially_started']}\n"

    content += f"\n---\n*自动推送 | {timestamp} | 下次推送: 明日21:07*"

    # Write push file
    push_file = os.path.join(TARGET_DIR, "LATEST_PUSH.md")
    with open(push_file, "w", encoding="utf-8") as f:
        f.write(content)

    # History
    history_file = os.path.join(TARGET_DIR, f"push_history_{now.strftime('%Y%m%d_%H%M%S')}.md")
    with open(history_file, "w", encoding="utf-8") as f:
        f.write(content)

    history_files = sorted([f for f in os.listdir(TARGET_DIR) if f.startswith("push_history_")])
    for old_file in history_files[:-50]:
        os.remove(os.path.join(TARGET_DIR, old_file))

    print(f"[{timestamp}] Push generated: {push_file}")

    # ========== TELEGRAM PUSH ==========
    print(f"[{timestamp}] Sending to Telegram...")

    tg_parts = []

    # Part 1: Box table (only if daily_table section is on)
    if sections.get("daily_table", True):
        daily_title = titles.get("daily", "线上办公数据汇总")
        tg_parts.append(f"📊 {daily_title} — {latest_date}\n\n<pre>{box_table}</pre>\n\n{summary_ph}\n{summary_bd}\n{summary_mm}\n{summary_all}")

    # Part 2: Fraud alerts only (DoD + anomalies are image-only)
    if sections.get("fraud_alerts", True) and fraud_lines:
        tg2 = "<b>🚨 风控告警</b>\n"
        for line in fraud_lines:
            tg2 += esc_html(line) + "\n"
        tg_parts.append(tg2.strip())

    # Part 3: Price summary
    if sections.get("price_summary", True) and price_summary_text:
        tg_parts.append(price_summary_text)

    # HJ data is image-only (see photos below)
    # Warnings about missing data
    if hj_warnings:
        tg_parts.append("\n".join(hj_warnings))

    sent = send_telegram_parts(tg_parts)
    print(f"[{timestamp}] Telegram text: {sent}/{len(tg_parts)} messages sent")

    # ========== TELEGRAM PHOTOS ==========
    photos_sent = 0
    photos_total = 0

    # Image 1: Daily 当日汇总 exact replica (Document, lossless)
    if sections.get("daily_table", True) and daily_sheet_rows:
        img_daily = build_daily_full_image_from_sheet(daily_sheet_rows, latest_date)
        if img_daily:
            photos_total += 1
            if send_telegram_document(img_daily, caption=f"📊 当日汇总 — {latest_date}"):
                photos_sent += 1

    # Image 2: Monthly cost sheet exact replica (33 columns, Document)
    if sections.get("monthly_table", True) and monthly_sheet_rows:
        monthly_date_label = latest_date
        if latest_date and latest_date != "未知":
            parts = latest_date.split("-")
            if len(parts) >= 3:
                monthly_date_label = f"{parts[0]}-{parts[1]}-01 至 {latest_date}"
        img_monthly_full = build_monthly_full_image_from_sheet(monthly_sheet_rows, monthly_date_label)
        if img_monthly_full:
            photos_total += 1
            if send_telegram_document(img_monthly_full, caption=f"📊 当月累计汇总 — {monthly_date_label}"):
                photos_sent += 1

    # Image 6: Headcount & efficiency
    if sections.get("headcount", True):
        img_hc = build_headcount_image(today, latest_date)
        if img_hc:
            photos_total += 1
            if send_telegram_photo(img_hc):
                photos_sent += 1

    # Image 7: DoD comparison
    if sections.get("dod_comparison", True):
        img_dod = build_dod_image(today, yesterday, prev_date, latest_date)
        if img_dod:
            photos_total += 1
            if send_telegram_photo(img_dod):
                photos_sent += 1

    # Image 8: Anomaly alerts
    if sections.get("anomaly_alerts", True):
        img_anom = build_anomaly_image(anomaly_lines)
        if img_anom:
            photos_total += 1
            if send_telegram_photo(img_anom):
                photos_sent += 1

    # Image 9: HJ Daily Full (all 32 cols)
    if sections.get("hijack_office", True) and hj_daily_summary:
        img_hj_daily = build_hj_daily_full_image(hj_daily_summary, latest_date)
        if img_hj_daily:
            photos_total += 1
            if send_telegram_document(img_hj_daily, caption=f"劫持运营 — 当天数据汇总 | {latest_date}"):
                photos_sent += 1

    # Image 10: HJ Monthly Full (all 28 cols)
    if sections.get("hijack_office", True) and hj_monthly_summary:
        img_hj_monthly = build_hj_monthly_full_image(hj_monthly_summary, latest_date)
        if img_hj_monthly:
            photos_total += 1
            if send_telegram_document(img_hj_monthly, caption=f"劫持运营 — 当月数据汇总 | 5月"):
                photos_sent += 1

    # Image 13: HJ HR
    if sections.get("hijack_hr", True):
        img3 = build_hj_hr_image(hj_hr, hj_hr_monthly, hr_date_str if hj_hr else "")
        if img3:
            photos_total += 1
            if send_telegram_photo(img3):
                photos_sent += 1

    print(f"[{timestamp}] Telegram photos: {photos_sent}/{photos_total} images sent")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", type=str, default=None, help="Target date YYYY-MM-DD")
    parser.add_argument("--month", type=str, default=None, help="Target month YYYY-MM (full month summary)")
    parser.add_argument("--sections", type=str, default=None, help="Comma-separated section names to enable (overrides config)")
    args = parser.parse_args()
    generate_push(target_date=args.date, target_month=args.month, override_sections=args.sections)
