"""Comparison Push — Day-over-Day and Month-over-Month data comparison.
/compare       → auto-detect current file, compare vs previous day and previous month
/compare_date  → specify exact date, read from archive
"""

import os, sys, json, io, hashlib, shutil
from datetime import datetime, timedelta

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

from PIL import Image, ImageDraw, ImageFont
import openpyxl

with open(os.path.join(SCRIPT_DIR, "config.json"), "r", encoding="utf-8") as f:
    CFG = json.load(f)

DATA_FOLDER = CFG.get("data_folder", SCRIPT_DIR)
ARCHIVE_DIR = os.path.join(SCRIPT_DIR, "data", "comparison_archive")
GEN_DIR = os.path.join(SCRIPT_DIR, "data", "generated")
os.makedirs(ARCHIVE_DIR, exist_ok=True)
os.makedirs(GEN_DIR, exist_ok=True)

FONT_PATH = "C:/Windows/Fonts/msyh.ttc"
FONT_BOLD_PATH = "C:/Windows/Fonts/msyhbd.ttc"

NEW_MAP = {"现场人数": 4, "转线上人数": 5, "线上人数": 6, "人均开发": 7,
           "总注册": 8, "总开发人数": 9, "首存金额": 10,
           "充值人数": 19, "复存人数": 20, "存款": 21, "提款": 22, "存提差": 23}
OLD_MAP = {"现场人数": 3, "转线上人数": None, "线上人数": 5, "人均开发": 6,
           "总注册": 7, "总开发人数": 8, "首存金额": 9,
           "充值人数": 18, "复存人数": 19, "存款": 20, "提款": 21, "存提差": 22}
LABELS = list(NEW_MAP.keys())


def _get_map(ws):
    return NEW_MAP if "Onsite HC" in str(ws.cell(row=3, column=4).value or "") else OLD_MAP


def _find_total_row(ws):
    for r in range(1, ws.max_row + 1):
        for c in (1, 2, 3):
            if str(ws.cell(row=r, column=c).value or "").strip() == "总":
                return r
    return None


def _read_internal_date(filepath):
    """Extract internal data cutoff date from Excel.
    Priority:
      1. Parse 汇总 sheet title: '菲律宾 07月01-25' → 2026-07-25
         If only month found ('菲律宾 07月01-'), use month + scan sheets for latest day
      2. Scan platform sheets for latest data row with non-zero FTD
      3. Fallback: datetime cells in rows 1-5
    Returns 'YYYY-MM-DD' or None."""
    import re
    if not os.path.isfile(filepath):
        return None
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    year = datetime.now().year

    # 1. Parse title text
    for r in range(1, 4):
        for c in range(1, 6):
            v = str(ws.cell(row=r, column=c).value or "")
            # Full pattern: '07月01-25'
            m = re.search(r'(\d{1,2})\s*月\s*\d{1,2}\s*[-–—]\s*(\d{1,2})', v)
            if m:
                wb.close()
                return f"{year}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
            # Partial pattern: '07月01-' (month only, need to find day from sheets)
            m2 = re.search(r'(\d{1,2})\s*月', v)
            if m2:
                month = int(m2.group(1))
                # Scan platform sheets for latest day in this month
                latest_day = _find_latest_day_in_month(wb, year, month)
                if latest_day:
                    wb.close()
                    return f"{year}-{month:02d}-{latest_day:02d}"

    # 2. Fallback: datetime cells
    for r in range(1, 6):
        for c in range(1, 6):
            dv = ws.cell(row=r, column=c).value
            if dv and hasattr(dv, 'strftime'):
                wb.close()
                return dv.strftime('%Y-%m-%d')

    # 3. Platform sheets
    latest_day = _find_latest_day_in_month(wb, year, None)
    wb.close()
    if latest_day:
        today = datetime.now()
        return f"{year}-{today.month:02d}-{latest_day:02d}"
    return None


def _find_latest_day_in_month(wb, year, month):
    """Scan platform sheets for the latest day with real business data.
    Checks: register(col7), FTD(col8), FTD amount(col20), depositors(col11),
            deposit(col23), withdraw(col24). At least one must be non-zero."""
    import re
    latest_day = 0
    for sn in wb.sheetnames:
        if not re.match(r'^(PH|BD|MM)\d', sn):
            continue
        try:
            pws = wb[sn]
            for row_idx in range(pws.max_row, 5, -1):
                d = pws.cell(row=row_idx, column=1).value
                if not d or not hasattr(d, 'strftime'):
                    continue
                if month is not None and d.month != month:
                    continue
                # Check multiple business fields — at least one non-zero
                fields = [pws.cell(row=row_idx, column=c).value for c in (7, 8, 20, 11, 23, 24)]
                has_data = False
                for v in fields:
                    try:
                        if v is not None and float(v) != 0:
                            has_data = True
                            break
                    except (ValueError, TypeError):
                        pass
                if has_data and d.day > latest_day:
                    latest_day = d.day
        except Exception:
            continue
    return latest_day if latest_day > 0 else None


# ── Platform sheet column mapping (same for June and July templates) ──
# Daily row fields to sum cumulatively:
_DAILY_COLS = {
    "注册": 7, "FTD": 8, "首存金额": 20,
    "充值人数": 11, "复存人数": 12,
    "存款": 23, "提款": 24,
}
# 汇总 sheet headcount columns (per-platform rows)
_HC_COLS = {"现场人数": 4, "转线上人数": 5, "线上人数": 6}


def _read_as_of_date(filepath, target_date, dev_platforms):
    """Sum platform sheet daily rows up to target_date for each dev platform.
    Returns (data_dict, metadata_dict).
    HC fields are only valid when the file is an exact-date snapshot."""
    import re
    if not os.path.isfile(filepath):
        raise FileNotFoundError(f"文件不存在: {os.path.basename(filepath)}")

    # ── Determine source type ──
    file_cutoff_date = _read_internal_date(filepath)
    is_exact_snapshot = (file_cutoff_date is not None and file_cutoff_date == target_date)
    source_type = "exact_snapshot" if is_exact_snapshot else "full_month_as_of"
    warning = None if is_exact_snapshot else (
        f"⚠️ {_date_label(target_date)}缺少独立历史快照，"
        f"人数及人均开发无法从全月文件准确回溯。")

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # ── Read headcount from 汇总 sheet ──
    ws_summary = wb[wb.sheetnames[0]]
    row4_col3 = str(ws_summary.cell(row=4, column=3).value or "").strip()
    is_june_template = (row4_col3 == "1" or row4_col3 == "" or row4_col3.isdigit())

    if is_june_template:
        PLAT_COL, ONSITE_COL, TRANSFER_COL, ONLINE_COL = 2, 3, 4, 5
    else:
        PLAT_COL, ONSITE_COL, TRANSFER_COL, ONLINE_COL = 3, 4, 5, 6

    hc = {}
    for r in range(4, ws_summary.max_row + 1):
        platform = str(ws_summary.cell(row=r, column=PLAT_COL).value or "").strip()
        if platform in dev_platforms and platform not in hc:
            onsite = ws_summary.cell(row=r, column=ONSITE_COL).value
            transfer = ws_summary.cell(row=r, column=TRANSFER_COL).value
            online = ws_summary.cell(row=r, column=ONLINE_COL).value
            try: onsite = int(float(onsite)) if onsite else 0
            except: onsite = 0
            try: transfer = int(float(transfer)) if transfer else 0
            except: transfer = 0
            try: online = int(float(online)) if online else 0
            except: online = 0
            hc[platform] = {"现场": onsite, "转线上": transfer, "线上": online}

    # ── Sum daily data from each platform sheet ──
    totals = {k: 0.0 for k in _DAILY_COLS}
    total_hc_onsite = 0
    total_hc_transfer = 0
    total_hc_online = 0

    for plat in dev_platforms:
        if plat not in wb.sheetnames:
            continue
        ws = wb[plat]
        phc = hc.get(plat, {"现场": 0, "转线上": 0, "线上": 0})
        if is_exact_snapshot:
            total_hc_onsite += phc["现场"]
            total_hc_transfer += phc["转线上"]
            total_hc_online += phc["线上"]

        for r in range(6, ws.max_row + 1):
            d = ws.cell(row=r, column=1).value
            if not d or not hasattr(d, 'strftime'):
                continue
            if d.strftime('%Y-%m-%d') > target_date:
                break
            for field, col in _DAILY_COLS.items():
                v = ws.cell(row=r, column=col).value
                try: totals[field] += float(v) if v is not None else 0.0
                except: pass

    wb.close()

    # Compute derived fields (only if HC is valid)
    total_ftd = int(totals["FTD"])
    day_of_month = int(target_date.split('-')[2])
    if is_exact_snapshot and total_hc_online > 0 and day_of_month > 0:
        avg_ftd = total_ftd / total_hc_online / day_of_month
        headcount_available = True
    else:
        avg_ftd = None
        headcount_available = False

    net_dep = totals["存款"] - totals["提款"]

    data = {
        "现场人数": total_hc_onsite if is_exact_snapshot else None,
        "转线上人数": total_hc_transfer if is_exact_snapshot else None,
        "线上人数": total_hc_online if is_exact_snapshot else None,
        "人均开发": avg_ftd,
        "总注册": int(totals["注册"]),
        "总开发人数": total_ftd,
        "首存金额": totals["首存金额"],
        "充值人数": int(totals["充值人数"]),
        "复存人数": int(totals["复存人数"]),
        "存款": totals["存款"],
        "提款": totals["提款"],
        "存提差": net_dep,
    }

    meta = {
        "source_type": source_type,
        "source_file": os.path.basename(filepath),
        "file_cutoff_date": file_cutoff_date,
        "target_date": target_date,
        "headcount_available": headcount_available,
        "warning": warning,
    }

    return data, meta


def _read_total_row(filepath):
    if not os.path.isfile(filepath):
        raise FileNotFoundError(f"文件不存在: {os.path.basename(filepath)}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    col_map = _get_map(ws)
    total_row = _find_total_row(ws)
    if total_row is None:
        wb.close()
        raise ValueError(f"未找到 '总' 行: {os.path.basename(filepath)}")
    data = {}
    for label in LABELS:
        c = col_map.get(label)
        if c is None: data[label] = None; continue
        v = ws.cell(row=total_row, column=c).value
        try: data[label] = float(v) if v is not None else 0.0
        except (ValueError, TypeError): data[label] = 0.0
    wb.close()
    return data


def _find_latest(data_folder, keyword, exclude_kw=None):
    best, best_mtime = None, 0
    for f in os.listdir(data_folder):
        if not f.endswith('.xlsx') or f.startswith('~$'): continue
        if keyword not in f: continue
        if exclude_kw and exclude_kw in f: continue
        path = os.path.join(data_folder, f)
        mtime = os.path.getmtime(path)
        if mtime > best_mtime: best_mtime = mtime; best = path
    return best


def _snapshot(src_path, archive_key):
    """Save snapshot. Returns (dest_path, warning_or_None)."""
    date_str = _read_internal_date(src_path)
    if not date_str:
        date_str = datetime.now().strftime('%Y-%m-%d')
    dest = os.path.join(ARCHIVE_DIR, f"{archive_key}_{date_str}.xlsx")
    with open(src_path, 'rb') as f:
        src_md5 = hashlib.md5(f.read()).hexdigest()
    if os.path.isfile(dest):
        with open(dest, 'rb') as f:
            if hashlib.md5(f.read()).hexdigest() == src_md5:
                return dest, None
        return dest, f"今天历史快照已存在，实时Excel已更新（MD5不同），是否覆盖？"
    shutil.copy2(src_path, dest)
    return dest, None


def _archive_path(category, date_str):
    return os.path.join(ARCHIVE_DIR, f"{category}_{date_str}.xlsx")


def _date_label(date_str):
    """'2026-07-25' → '7月25日'"""
    try:
        d = datetime.strptime(date_str, '%Y-%m-%d')
        return f"{d.month}月{d.day}日"
    except: return date_str


# ── Calculation / formatting ──

def _calc(cur, prev):
    if cur is None or prev is None: return 0, None, "—"
    diff = cur - prev
    if prev == 0: return diff, None, "—"
    pct = (diff / prev) * 100
    return diff, pct, "↑" if diff > 0 else ("↓" if diff < 0 else "—")


def _fmt_val(label, val):
    if val is None: return "—"
    if "人数" in label or "注册" in label or "开发" in label: return f"{int(round(val)):,}"
    if "人均" in label: return f"{val:,.2f}"
    return f"{val:,.2f}"


def _fmt_diff(label, diff):
    if "人数" in label or "注册" in label or "开发" in label: return f"{int(round(diff)):+,}"
    if "人均" in label: return f"{diff:+,.2f}"
    return f"{diff:+,.2f}"


def _fmt_pct(pct):
    return "—" if pct is None else f"{pct:+.1f}%"


# ── Image builder ──

def build_comparison_image(title, subtitle, cur_data, prev_data, cutoff_note=None, theme="blue"):
    font_title = ImageFont.truetype(FONT_BOLD_PATH, 18)
    font_sub = ImageFont.truetype(FONT_PATH, 14)
    font_hdr = ImageFont.truetype(FONT_BOLD_PATH, 13)
    font_body = ImageFont.truetype(FONT_PATH, 12)
    font_footer = ImageFont.truetype(FONT_PATH, 10)
    font_cutoff = ImageFont.truetype(FONT_PATH, 11)

    headers = ["指标", "当前日", "对比日", "增减值", "增减比例", "变化"]
    col_widths = [140, 120, 120, 120, 100, 60]
    rows = []
    for label in LABELS:
        diff, pct, arrow = _calc(cur_data.get(label, 0), prev_data.get(label, 0))
        rows.append((label,
                     _fmt_val(label, cur_data.get(label, 0)),
                     _fmt_val(label, prev_data.get(label, 0)),
                     _fmt_diff(label, diff), _fmt_pct(pct), arrow))

    row_h, hdr_h, title_h, sub_h = 28, 32, 42, 26
    cut_h = 28 if cutoff_note else 0
    total_w = sum(col_widths)
    total_h = title_h + sub_h + hdr_h + len(rows) * row_h + cut_h + 28

    # Theme colors
    if theme == "orange":
        DARK_BG = (60, 60, 65); LIGHT_BG = (245, 240, 230)
        TITLE_BG = (55, 55, 60); SUB_BG = (250, 240, 225)
    else:
        DARK_BG = (31, 56, 100); LIGHT_BG = (220, 235, 252)
        TITLE_BG = (31, 56, 100); SUB_BG = (220, 235, 252)

    WHITE = (255, 255, 255)
    GREEN_BG = (235, 255, 235); RED_BG = (255, 235, 235); GRAY_BG = (245, 245, 245)
    DARK = (33, 37, 41); GRAY = (140, 140, 140)
    YELLOW_BG = (255, 255, 220); ORANGE = (180, 120, 0)

    img = Image.new("RGB", (total_w, total_h), WHITE)
    draw = ImageDraw.Draw(img)
    draw.rectangle([0, 0, total_w - 1, title_h], fill=TITLE_BG)
    tw = font_title.getbbox(title)[2]
    draw.text(((total_w - tw) // 2, 10), title, fill=WHITE, font=font_title)
    y = title_h
    draw.rectangle([0, y, total_w - 1, y + sub_h], fill=SUB_BG)
    sw = font_sub.getbbox(subtitle)[2]
    draw.text(((total_w - sw) // 2, y + 4), subtitle, fill=DARK, font=font_sub)
    y += sub_h
    x = 0
    for ci, h in enumerate(headers):
        cw = col_widths[ci]
        draw.rectangle([x, y, x + cw - 1, y + hdr_h], fill=DARK_BG)
        tw = font_hdr.getbbox(h)[2]
        draw.text((x + (cw - tw) // 2, y + 7), h, fill=WHITE, font=font_hdr)
        x += cw
    y += hdr_h
    for ri, row in enumerate(rows):
        arrow = row[5]
        bg = GREEN_BG if arrow == "↑" else (RED_BG if arrow == "↓" else (GRAY_BG if ri % 2 == 0 else WHITE))
        x = 0
        for ci, val in enumerate(row):
            cw = col_widths[ci]
            draw.rectangle([x, y, x + cw - 1, y + row_h], fill=bg, outline=(230, 230, 230))
            tw = font_body.getbbox(str(val))[2]
            draw.text((x + (cw - tw) // 2, y + 5), str(val), fill=DARK, font=font_body)
            x += cw
        y += row_h
    if cutoff_note:
        draw.rectangle([0, y, total_w - 1, y + cut_h], fill=YELLOW_BG)
        dw = font_cutoff.getbbox(cutoff_note)[2]
        draw.text(((total_w - dw) // 2, y + 4), cutoff_note, fill=ORANGE, font=font_cutoff)
        y += cut_h
    ts = datetime.now().strftime('%Y-%m-%d %H:%M')
    footer = f"WFHDPbot  |  Generated: {ts}"
    tw_f = font_footer.getbbox(footer)[2]
    draw.text((total_w - tw_f - 10, total_h - 22), footer, fill=GRAY, font=font_footer)
    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=True)
    return buf.getvalue()


# ── Main generation ──

def generate_comparison(target_date_str=None):
    """Generate comparison. If target_date_str is given (YYYY-MM-DD), use archive.
    Otherwise auto-detect from current live file."""
    errors = []

    if target_date_str:
        # /compare_date mode: read from archive
        cur_date = target_date_str
        try:
            cur_dt = datetime.strptime(cur_date, '%Y-%m-%d')
            prev_day_dt = cur_dt - timedelta(days=1)
            prev_month_dt = cur_dt.replace(month=cur_dt.month - 1) if cur_dt.month > 1 else cur_dt.replace(year=cur_dt.year - 1, month=12)
        except:
            return False, f"日期格式错误: {target_date_str}", [], ""

        cur_path = _archive_path("development", cur_date)
        prev_day_path = _archive_path("development", prev_day_dt.strftime('%Y-%m-%d'))
        prev_month_path = _archive_path("development", prev_month_dt.strftime('%Y-%m-%d'))

        cur_label = _date_label(cur_date)
        prev_label = _date_label(prev_day_dt.strftime('%Y-%m-%d'))
        month_label = _date_label(prev_month_dt.strftime('%Y-%m-%d'))

    else:
        # /compare mode: resolve live file via active_month.json first
        import _platform_config as _pcfg
        _am_path, _am_name, _am_date, _am_errs = _pcfg.get_active_excel("development")
        if _am_path:
            live_path = _am_path
        else:
            live_path = _find_latest(DATA_FOLDER, "线上办公数据汇总", "劫持")
        if not live_path:
            return False, "未找到当前实时开发Excel", [], ""

        cur_date = _read_internal_date(live_path)
        if not cur_date:
            return False, "无法读取实时Excel内部日期", [], ""

        # Snapshot live file
        _snapshot(live_path, "development")
        _hij_path, _hij_name, _hij_date, _hij_errs = _pcfg.get_active_excel("hijack")
        if _hij_path:
            hij_path = _hij_path
        else:
            hij_path = _find_latest(DATA_FOLDER, "劫持（线上办公数据汇总）")
        if hij_path:
            _snapshot(hij_path, "hijack")

        cur_path = live_path
        try:
            cur_dt = datetime.strptime(cur_date, '%Y-%m-%d')
            prev_day_dt = cur_dt - timedelta(days=1)
            prev_month_dt = cur_dt.replace(month=cur_dt.month - 1) if cur_dt.month > 1 else cur_dt.replace(year=cur_dt.year - 1, month=12)
        except:
            return False, f"内部日期解析失败: {cur_date}", [], ""

        prev_day_path = _archive_path("development", prev_day_dt.strftime('%Y-%m-%d'))
        prev_month_path = _archive_path("development", prev_month_dt.strftime('%Y-%m-%d'))

        cur_label = _date_label(cur_date)
        prev_label = _date_label(prev_day_dt.strftime('%Y-%m-%d'))
        month_label = _date_label(prev_month_dt.strftime('%Y-%m-%d'))

    # Collect files — allow full-month files when exact archive snapshots missing
    def _resolve_file(archive_path, target_date_str, label):
        if os.path.isfile(archive_path):
            return archive_path
        # Fallback: find any full-month file containing target_date
        candidates = []
        for folder in [DATA_FOLDER, os.path.join(SCRIPT_DIR, 'data', 'comparison_archive')]:
            if not os.path.isdir(folder): continue
            for f in os.listdir(folder):
                if not f.endswith('.xlsx') or f.startswith('~$'): continue
                if '劫持' in f: continue
                if '线上办公' not in f: continue
                candidates.append(os.path.join(folder, f))
        for cand in candidates:
            try:
                wb = openpyxl.load_workbook(cand, data_only=True)
                has_date = False
                for sn in wb.sheetnames:
                    if not sn.startswith('PH'): continue
                    if len(sn) > 6: continue
                    ws = wb[sn]
                    for r in range(6, ws.max_row + 1):
                        d = ws.cell(row=r, column=1).value
                        if d and hasattr(d, 'strftime') and d.strftime('%Y-%m-%d') == target_date_str:
                            has_date = True; break
                    if has_date: break
                wb.close()
                if has_date:
                    return cand
            except: pass
        return None

    files = {}
    for key, archive_path, target_date_str, label in [
        ("cur", cur_path, cur_date, cur_label),
        ("prev_day", prev_day_path, prev_day_dt.strftime('%Y-%m-%d'), prev_label),
        ("prev_month", prev_month_path, prev_month_dt.strftime('%Y-%m-%d'), month_label),
    ]:
        resolved = _resolve_file(archive_path, target_date_str, label)
        if resolved:
            files[key] = resolved
        else:
            errors.append(f"缺少文件: {label} ({os.path.basename(archive_path)})")

    if errors:
        return False, "\n".join(errors), [], ""

    # Read data as-of specific dates using platform sheet summation
    import _platform_config as _pcfg
    dev_platforms = _pcfg.get_development_platforms()

    date_map = {"cur": cur_date, "prev_day": prev_day_dt.strftime('%Y-%m-%d'),
                "prev_month": prev_month_dt.strftime('%Y-%m-%d')}
    data = {}
    meta = {}
    warnings_list = []
    for key in files:
        try:
            d, m = _read_as_of_date(files[key], date_map[key], dev_platforms)
            data[key] = d
            meta[key] = m
            if m.get("warning"):
                warnings_list.append(f"{_date_label(date_map[key])}: {m['warning']}")
        except Exception as e:
            errors.append(f"读取失败 {os.path.basename(files[key])}: {e}")
    if errors:
        return False, "\n".join(errors), [], ""

    cur = data["cur"]; prev_day = data["prev_day"]; prev_month = data["prev_month"]

    # Images
    results = []
    cutoff = None
    if CFG.get("comparison", {}).get("previous_day_cutoff_note"):
        cutoff = (f"⚠️ 标注为「{prev_label}」的源文件，工作簿内部【汇总】截止日显示为22日。"
                  f"本次仍按用户指定名称列为{prev_label}。")

    # Collect all warnings for image footers
    all_warnings = list(warnings_list)
    day_warning = None
    month_warning = None
    if meta.get("prev_day", {}).get("warning"):
        day_warning = meta["prev_day"]["warning"]
    if meta.get("prev_month", {}).get("warning"):
        month_warning = meta["prev_month"]["warning"]

    img1 = build_comparison_image(
        f"线上办公数据对比 — {cur_label} vs {prev_label}",
        f"数据口径：{cur_label} 与 {prev_label} 当日汇总「总」行对比",
        cur, prev_day, cutoff_note=cutoff or day_warning)
    results.append((img1, f"📊 {cur_label} vs {prev_label}"))

    img2 = build_comparison_image(
        f"线上办公数据对比 — {cur_label} vs {month_label}",
        f"数据口径：{cur_label} 与 {month_label} 当日汇总「总」行对比",
        cur, prev_month, cutoff_note=month_warning)
    results.append((img2, f"📊 {cur_label} vs {month_label}"))

    # Conclusion — handle None HC gracefully
    def _pct(cur_d, prev_d, label):
        cv = cur_d.get(label); pv = prev_d.get(label)
        if cv is None or pv is None: return f"• {label}: 无法计算（缺少历史快照数据）"
        diff, pct, arrow = _calc(cv, pv)
        if pct is None: return f"• {label}: 无法计算（对比日数值为0）"
        direction = "增长" if pct > 0 else ("下降" if pct < 0 else "持平")
        return f"• {label}{direction} {abs(pct):.1f}%"

    def _build_day_conclusion():
        lines = []
        # Always include cumulative indicators
        lines.append(_pct(cur, prev_day, '总开发人数'))
        lines.append(_pct(cur, prev_day, '首存金额'))
        lines.append(_pct(cur, prev_day, '存提差'))
        # 人均开发 only if both sides have HC
        if _pct(cur, prev_day, '人均开发') is not None:
            lines.append(_pct(cur, prev_day, '人均开发'))
        return "\n".join(lines)

    def _build_month_conclusion():
        lines = []
        lines.append(_pct(cur, prev_month, '线上人数'))
        lines.append(_pct(cur, prev_month, '人均开发'))
        lines.append(_pct(cur, prev_month, '总开发人数'))
        # 提款 vs 存提差
        _, wdr_pct, _ = _calc(cur.get("提款", 0), prev_month.get("提款", 0))
        _, diff_pct, _ = _calc(cur.get("存提差", 0), prev_month.get("存提差", 0))
        if wdr_pct is not None and diff_pct is not None and wdr_pct > 0 and diff_pct < 0:
            lines.append(f"• 提款增长较快，导致存提差下降 {abs(diff_pct):.1f}%")
        elif wdr_pct is not None and diff_pct is not None:
            lines.append(f"• 提款变化 {wdr_pct:+.1f}%，存提差变化 {diff_pct:+.1f}%")
        else:
            lines.append("• 提款与存提差无法计算")
        return "\n".join(lines)

    conclusion = (
        f"📊 线上办公数据对比结论\n\n"
        f"【{cur_label} vs {prev_label}】\n"
        f"{_build_day_conclusion()}\n\n"
        f"【{cur_label} vs {month_label}】\n"
        f"{_build_month_conclusion()}"
    )

    return True, "", results, conclusion


def check_comparison_status():
    """Read-only status check. No PNG, no snapshot, no Excel modification.
    Returns formatted text for /compare_check."""
    lines = ["📊 数据对比状态检查", ""]

    # ── Realtime files ──
    lines.append("【实时文件】")
    dev_path = _find_latest(DATA_FOLDER, "线上办公数据汇总", "劫持")
    hij_path = _find_latest(DATA_FOLDER, "劫持（线上办公数据汇总）")

    dev_date = _read_internal_date(dev_path) if dev_path else None
    hij_date = _read_internal_date(hij_path) if hij_path else None

    if dev_path:
        lines.append(f"开发：{'✅' if dev_date else '❌ 无法识别日期'} {dev_date or ''}")
        lines.append(f"路径：{dev_path}")
    else:
        lines.append("开发：❌ 未找到实时文件")
    if hij_path:
        lines.append(f"劫持：{'✅' if hij_date else '❌ 无法识别日期'} {hij_date or ''}")
        lines.append(f"路径：{hij_path}")
    else:
        lines.append("劫持：❌ 未找到实时文件")

    if not dev_date:
        lines.append("")
        lines.append("❌ 无法识别实时Excel实际数据截止日期")
        return "\n".join(lines)

    # ── Default /compare ──
    lines.append("")
    lines.append("【默认 /compare】")
    try:
        cur_dt = datetime.strptime(dev_date, '%Y-%m-%d')
        prev_day_dt = cur_dt - timedelta(days=1)
        prev_month_dt = (cur_dt.replace(month=cur_dt.month - 1) if cur_dt.month > 1
                         else cur_dt.replace(year=cur_dt.year - 1, month=12))
    except Exception:
        lines.append("❌ 日期解析失败")
        return "\n".join(lines)

    cur_label = _date_label(dev_date)
    prev_label = _date_label(prev_day_dt.strftime('%Y-%m-%d'))
    month_label = _date_label(prev_month_dt.strftime('%Y-%m-%d'))

    lines.append(f"当前：{dev_date} ({cur_label})")
    lines.append(f"前一日：{prev_day_dt.strftime('%Y-%m-%d')} ({prev_label})")
    lines.append(f"上月同日：{prev_month_dt.strftime('%Y-%m-%d')} ({month_label})")
    lines.append("")

    default_ok = True
    missing = []
    for label, date_str in [("当前实时", dev_date),
                             ("前一日", prev_day_dt.strftime('%Y-%m-%d')),
                             ("上月同日", prev_month_dt.strftime('%Y-%m-%d'))]:
        if label == "当前实时":
            exists = dev_path is not None
        else:
            p = _archive_path("development", date_str)
            exists = os.path.isfile(p)
        icon = "✅" if exists else "❌"
        lines.append(f"  {icon} {label}")
        if not exists:
            default_ok = False
            if label == "前一日":
                missing.append(f"development_{prev_day_dt.strftime('%Y-%m-%d')}.xlsx")
            elif label == "上月同日":
                missing.append(f"development_{prev_month_dt.strftime('%Y-%m-%d')}.xlsx")

    if default_ok:
        lines.append("")
        lines.append("状态：✅ 默认 /compare 可以执行")
    else:
        lines.append("")
        lines.append("状态：❌ 默认 /compare 无法执行")
        if missing:
            lines.append(f"缺少：{', '.join(missing)}")

    # ── /compare_date 2026-07-25 ──
    lines.append("")
    lines.append("【/compare_date 2026-07-25】")
    spec_ok = True
    for label, date_str in [("当前", "2026-07-25"),
                             ("前一日", "2026-07-24"),
                             ("上月同日", "2026-06-25")]:
        p = _archive_path("development", date_str)
        exists = os.path.isfile(p)
        icon = "✅" if exists else "❌"
        lines.append(f"  {icon} {label} ({date_str})")
        if not exists:
            spec_ok = False
    lines.append("")
    lines.append(f"状态：{'✅ 可以执行' if spec_ok else '❌ 无法执行'}")

    # ── Archive summary ──
    lines.append("")
    lines.append("【Archive】")
    dev_archives = sorted([f for f in os.listdir(ARCHIVE_DIR)
                           if f.startswith("development_") and f.endswith(".xlsx")])
    hij_archives = sorted([f for f in os.listdir(ARCHIVE_DIR)
                           if f.startswith("hijack_") and f.endswith(".xlsx")])
    lines.append(f"Development：{len(dev_archives)} 份")
    lines.append(f"Hijack：{len(hij_archives)} 份")

    # Count quarantined snapshots
    quar_dir = os.path.join(ARCHIVE_DIR, "quarantine")
    quar_count = 0
    if os.path.isdir(quar_dir):
        quar_count = len([f for f in os.listdir(quar_dir) if f.endswith('.xlsx')])
    lines.append(f"隔离异常快照：{quar_count} 份")

    if dev_archives:
        first = dev_archives[0].replace("development_", "").replace(".xlsx", "")
        last = dev_archives[-1].replace("development_", "").replace(".xlsx", "")
        lines.append(f"最早：{first}")
        lines.append(f"最新：{last}")

    # Warn about snapshots newer than realtime date
    if dev_date and dev_archives:
        try:
            rt = datetime.strptime(dev_date, '%Y-%m-%d')
            newer = []
            for a in dev_archives:
                d_str = a.replace("development_", "").replace(".xlsx", "")
                try:
                    ad = datetime.strptime(d_str, '%Y-%m-%d')
                    if ad > rt: newer.append(a)
                except: pass
            if newer:
                lines.append("")
                lines.append("⚠️ Archive存在晚于实时日期的快照：")
                for n in newer: lines.append(f"  {n}")
                lines.append("请检查是否为旧日期识别逻辑产生的错误命名文件。")
        except: pass

    return "\n".join(lines)


def send_comparison(send_photo_fn, send_message_fn, target_date=None):
    ok, error, images, conclusion = generate_comparison(target_date)
    if not ok:
        send_message_fn(f"❌ 数据对比生成失败:\n{error}")
        return
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    os.makedirs(GEN_DIR, exist_ok=True)
    for i, (img_bytes, caption) in enumerate(images):
        fname = f"comparison_{ts}_{i+1}.png"
        with open(os.path.join(GEN_DIR, fname), "wb") as f:
            f.write(img_bytes)
    for img_bytes, caption in images:
        try:
            send_photo_fn(img_bytes, caption)
        except Exception as e:
            send_message_fn(f"⚠️ 图片发送失败: {e}")
    send_message_fn(conclusion)
