"""PH33 当天数据汇总图 — pure Pillow renderer (Railway/Linux).
Replaces the Windows Excel COM screenshot for the summary image.
Data is read from 当天数据汇总 sheet of the hijack Excel via openpyxl.
"""

import os
import io
import sys
from datetime import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

import openpyxl
from PIL import Image, ImageDraw, ImageFont

try:
    import _runtime
    FONT = _runtime.font_path() or "C:/Windows/Fonts/msyh.ttc"
    FONT_BOLD = _runtime.font_bold_path() or "C:/Windows/Fonts/msyhbd.ttc"
except ImportError:
    FONT = "C:/Windows/Fonts/msyh.ttc"
    FONT_BOLD = "C:/Windows/Fonts/msyhbd.ttc"


def render_hijack_summary(filepath, output_dir=None):
    """Render PH33 当天数据汇总 as a horizontal PNG via Pillow.
    Returns unified result dict, never raises on data errors.
    """
    result = {"success": False, "renderer": "pillow"}

    if not os.path.isfile(filepath):
        result["error"] = f"文件不存在: {os.path.basename(filepath)}"
        return result
    result["source_file"] = filepath

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        result["error"] = f"无法打开 Excel: {str(e)[:60]}"
        return result

    if "当天数据汇总" not in wb.sheetnames:
        wb.close()
        result["error"] = "缺少 当天数据汇总 sheet"
        return result
    ws = wb["当天数据汇总"]

    # data_date from sheet
    data_date = None
    for r in range(1, 4):
        for c in range(1, 20):
            v = ws.cell(row=r, column=c).value
            if v and hasattr(v, 'strftime'):
                data_date = v.strftime('%Y-%m-%d')
                break
        if data_date:
            break
    if not data_date:
        wb.close()
        result["error"] = "无法识别 data_date"
        return result
    result["data_date"] = data_date

    # Read headers (row 3) + data (row 4)
    headers = []
    values = []
    for c in range(1, 33):
        h = str(ws.cell(row=3, column=c).value or f"C{c}").replace("\n", " ")
        v = ws.cell(row=4, column=c).value
        # Format numeric
        try:
            fv = float(v) if v is not None else 0.0
            if abs(fv) >= 1000:
                vs = f"{fv:,.0f}"
            elif fv != int(fv):
                vs = f"{fv:,.2f}"
            else:
                vs = f"{int(fv):,}"
        except (ValueError, TypeError):
            vs = str(v) if v is not None else "0"
        headers.append(h[:12])
        values.append(vs)
    wb.close()

    # Fonts
    try:
        ft = ImageFont.truetype(FONT_BOLD, 15)
        fh = ImageFont.truetype(FONT_BOLD, 10)
        fb = ImageFont.truetype(FONT, 10)
        ff = ImageFont.truetype(FONT, 8)
    except Exception as e:
        result["error"] = f"字体无法加载: {str(e)[:60]}"
        return result

    # Column widths
    col_w = []
    for ci in range(32):
        hw = fh.getbbox(headers[ci])[2] + 12
        vw = fb.getbbox(values[ci])[2] + 12
        col_w.append(max(hw, vw, 46))

    tw = sum(col_w)
    th = 42 + 28 + 28 + 30

    DARK = (55, 55, 60)   # deep gray (orange theme family)
    W = (255, 255, 255)
    DK = (33, 37, 41)
    GY = (140, 140, 140)

    img = Image.new("RGB", (tw, th), W)
    d = ImageDraw.Draw(img)

    d.rectangle([0, 0, tw - 1, 42], fill=DARK)
    title = f"劫持运营 — 当天数据汇总 | {data_date}"
    tw_t = ft.getbbox(title)[2]
    d.text(((tw - tw_t) // 2, 10), title, fill=W, font=ft)

    y = 42
    x = 0
    for ci, h in enumerate(headers):
        wc = col_w[ci]
        d.rectangle([x, y, x + wc - 1, y + 28], fill=DARK)
        tw_h = fh.getbbox(h)[2]
        d.text((x + (wc - tw_h) // 2, y + 7), h, fill=W, font=fh)
        x += wc
    y += 28

    x = 0
    for ci, v in enumerate(values):
        wc = col_w[ci]
        d.rectangle([x, y, x + wc - 1, y + 28], fill=(249, 251, 253), outline=(222, 226, 230))
        tw_v = fb.getbbox(v)[2]
        d.text((x + (wc - tw_v) // 2, y + 6), v, fill=DK, font=fb)
        x += wc
    y += 28

    footer = f"WFHDPbot | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Pillow"
    tw_f = ff.getbbox(footer)[2]
    d.text((tw - tw_f - 10, y + 4), footer, fill=GY, font=ff)

    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=True)

    if output_dir is None:
        try:
            output_dir = _runtime.generated_dir()
        except Exception:
            output_dir = os.path.join(SCRIPT_DIR, "data", "generated")
    os.makedirs(output_dir, exist_ok=True)
    out_path = os.path.join(output_dir, f"hijack_summary_{data_date}.png")
    with open(out_path, "wb") as f:
        f.write(buf.getvalue())

    result.update({
        "success": True,
        "path": out_path,
        "width": img.width,
        "height": img.height,
        "size": len(buf.getvalue()),
    })
    return result
