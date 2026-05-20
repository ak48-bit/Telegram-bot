import openpyxl
from datetime import datetime

EXCEL_FILE = r"C:\Users\ak481\OneDrive\Desktop\新建文件夹\26年04月 线上办公数据汇.xlsx"
TARGET_DIR = r"C:\Users\ak481\OneDrive\Desktop\ak 线上办公部门skills建议和调用"

PLATFORMS = ["PH09", "PH09-2", "PH25", "PH18", "PH30", "PH05", "PH16", "BD02", "BD05"]

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

results = []
total_ftd = 0
total_ftd_amt = 0

for name in PLATFORMS:
    if name not in wb.sheetnames:
        results.append((name, None, None, None, None, "Sheet not found"))
        continue

    ws = wb[name]
    last_row = None
    for row_idx in range(ws.max_row, 5, -1):
        date_val = ws.cell(row=row_idx, column=1).value
        ftd_val = ws.cell(row=row_idx, column=8).value
        if date_val is not None and hasattr(date_val, "strftime") and ftd_val is not None:
            try:
                float(ftd_val)
                last_row = row_idx
                break
            except:
                continue

    if last_row is None:
        results.append((name, None, None, None, None, "No data"))
        continue

    def cell(c):
        v = ws.cell(row=last_row, column=c).value
        if v is None:
            return 0
        try:
            return float(v)
        except:
            return 0

    date_val = ws.cell(row=last_row, column=1).value
    date_str = date_val.strftime("%Y-%m-%d") if hasattr(date_val, "strftime") else str(date_val)[:10]
    ftd = int(cell(8))
    ftd_amt = cell(20)
    unit_price = cell(15)

    total_ftd += ftd
    total_ftd_amt += ftd_amt

    results.append((name, date_str, ftd, ftd_amt, unit_price, "OK"))

wb.close()

now = datetime.now()
timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
avg_price = total_ftd_amt / total_ftd if total_ftd > 0 else 0

lines = []
lines.append("# 各站点昨日FTD & 单价汇总")
lines.append(f"**推送时间**: {timestamp}")
lines.append("**数据日期**: 昨日（各站点最新数据日）")
lines.append("")
lines.append("---")
lines.append("")
lines.append("## 昨日总首存人数 & 单价明细")
lines.append("| 站点 | 数据日期 | 首存人数(FTD) | 首存金额 | 新客单价 |")
lines.append("|------|----------|---------------|----------|----------|")

for name, date_str, ftd, ftd_amt, unit_price, status in results:
    if status != "OK":
        lines.append(f"| {name} | - | - | - | - | ({status})")
    else:
        def fmt_k(v):
            if abs(v) >= 1000:
                return f"{v/1000:.0f}K"
            return f"{v:,.0f}"
        lines.append(f"| {name} | {date_str} | {ftd} | {fmt_k(ftd_amt)} | {unit_price:,.0f} |")

lines.append("")
lines.append("## 汇总")
lines.append(f"- **总首存人数(FTD)**: {total_ftd} 人")
lines.append(f"- **总首存金额**: {total_ftd_amt:,.0f}")
lines.append(f"- **综合新客单价**: {avg_price:,.0f} (总首存金额/总首存人数)")
lines.append("")
lines.append("---")
lines.append(f"*专项推送 | {timestamp}*")

content = "\n".join(lines)

push_file = TARGET_DIR + "\\FTD_PRICE_PUSH.md"
with open(push_file, "w", encoding="utf-8") as f:
    f.write(content)

with open(TARGET_DIR + "\\LATEST_PUSH.md", "w", encoding="utf-8") as f:
    f.write(content)

print(f"[{timestamp}] FTD/Price Push generated: {push_file}")
print(content)
