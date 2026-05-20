import os
from datetime import datetime
import openpyxl

TARGET_DIR = r"C:\Users\ak481\OneDrive\Desktop\ak 线上办公部门skills建议和调用"
EXCEL_FILE = r"C:\Users\ak481\OneDrive\Desktop\新建文件夹\26年05月 线上办公数据汇总.xlsx"

PH_PLATFORMS = ["PH09", "PH09-2", "PH25", "PH18", "PH30", "PH05", "PH16"]
BD_PLATFORMS = ["BD02", "BD05"]
MM_PLATFORMS = ["MM01"]

TIPS = [
    "Check zero-order sites today, contact immediately for reasons",
    "Review day-over-day FTD changes per site",
    "Monitor sites with withdrawal rate above 70%",
    "Check for sites with 3 consecutive days of volume drop",
    "Verify the new version data comparison table is updated",
    "Spot-check player sources for 1-2 high-FTD sites",
    "Confirm today's anomaly alerts have been resolved",
    "Remind team to use one more social platform",
    "Compare this week vs last week FTD trends per site",
    "Watch for signs of virtual numbers or same-IP activity",
    "Confirm performance evaluation criteria sent to all staff",
    "Organize today's data for tomorrow morning's report",
]


def read_platform_data(ws):
    last_row = None
    for row_idx in range(ws.max_row, 5, -1):
        date_val = ws.cell(row=row_idx, column=1).value
        ftd_val = ws.cell(row=row_idx, column=8).value
        if date_val is not None and hasattr(date_val, 'strftime') and ftd_val is not None:
            try:
                float(ftd_val)
                last_row = row_idx
                break
            except (ValueError, TypeError):
                continue

    if last_row is None:
        return None

    def cell(c):
        v = ws.cell(row=last_row, column=c).value
        if v is None:
            return 0
        try:
            return float(v)
        except (ValueError, TypeError):
            return 0

    return {
        "date": ws.cell(row=last_row, column=1).value,
        "register": int(cell(7)),
        "ftd": int(cell(8)),
        "dps_ppl": int(cell(11)),
        "ftd_amount": cell(20),
        "total_dps_amt": cell(23),
        "total_wdr_amt": cell(24),
        "diff": cell(25),
        "new_cust_avg": cell(15),
        "wdr_ppl": int(cell(12)),
        "roi": cell(34),
    }


def read_headcount(wb):
    ws = wb[wb.sheetnames[0]]
    hc = {}
    for row_idx in range(4, ws.max_row + 1):
        platform = ws.cell(row=row_idx, column=2).value
        if platform is None:
            continue
        if isinstance(platform, str) and platform.strip():
            office = ws.cell(row=row_idx, column=3).value or 0
            online = ws.cell(row=row_idx, column=5).value or 0
            try:
                office = int(float(office))
            except (ValueError, TypeError):
                office = 0
            try:
                online = int(float(online))
            except (ValueError, TypeError):
                online = 0
            hc[platform.strip()] = {"office": office, "online": online}
    return hc


def evaluate_status(d):
    if d["ftd"] == 0:
        return "critical"
    if d["wdr_ppl"] > d["dps_ppl"] and d["dps_ppl"] > 0:
        ratio = d["wdr_ppl"] / d["dps_ppl"]
        if ratio > 1.5:
            return "critical"
        return "warning"
    if d["ftd"] < 10 and d["ftd"] > 0:
        return "warning"
    if d["roi"] < 0:
        return "warning"
    return "ok"


def evaluate_fraud_risks(d):
    risks = []
    reg = d.get("register", 0)
    ftd = d.get("ftd", 0)
    dps = d.get("total_dps_amt", 0)
    wdr = d.get("total_wdr_amt", 0)
    diff = d.get("diff", 0)
    if reg > 0 and ftd > 0:
        conversion = ftd / reg
        if conversion > 0.7:
            risks.append(f"Reg-to-FTD rate={conversion:.0%}(>70% redline)")
    if dps > 0 and ftd > 0:
        wdr_rate = wdr / dps
        if wdr_rate > 0.9:
            risks.append(f"Withdrawal rate={wdr_rate:.0%}(>90% redline)")
        diff_ratio = diff / dps
        if diff_ratio < 0.1:
            risks.append(f"Net deposit ratio={diff_ratio:.0%}(<10% redline)")
    return risks


def fmt_k(v):
    if abs(v) >= 1_000_000:
        return f"{v/1_000_000:.1f}M"
    if abs(v) >= 1_000:
        return f"{v/1_000:.0f}K"
    return f"{v:,.0f}"


def generate_push():
    now = datetime.now()
    timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
    tip = TIPS[now.minute % len(TIPS)]

    if not os.path.exists(EXCEL_FILE):
        content = f"""# Online Field Team - Real-time Push
**Push Time**: {timestamp}
**Next Push**: 1 hour

---

## Error
> Excel data source not found: {EXCEL_FILE}

---
*Auto push | {timestamp} | Next push: 1 hour*"""
        push_file = os.path.join(TARGET_DIR, "LATEST_PUSH_EN.md")
        with open(push_file, "w", encoding="utf-8") as f:
            f.write(content)
        print(f"[{timestamp}] EN Push generated (ERROR): {push_file}")
        return

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    headcount = read_headcount(wb)

    # --- PH platforms ---
    ph_data = {}
    for name in PH_PLATFORMS:
        if name in wb.sheetnames:
            d = read_platform_data(wb[name])
            if d:
                d["status"] = evaluate_status(d)
                d["fraud_risks"] = evaluate_fraud_risks(d)
                d["office"] = headcount.get(name, {}).get("office", 0)
                d["online"] = headcount.get(name, {}).get("online", 0)
                ph_data[name] = d

    # --- BD platforms ---
    bd_data = {}
    for name in BD_PLATFORMS:
        if name in wb.sheetnames:
            d = read_platform_data(wb[name])
            if d:
                d["status"] = evaluate_status(d)
                d["fraud_risks"] = evaluate_fraud_risks(d)
                d["office"] = headcount.get(name, {}).get("office", 0)
                d["online"] = headcount.get(name, {}).get("online", 0)
                bd_data[name] = d

    # --- MM platforms ---
    mm_data = {}
    for name in MM_PLATFORMS:
        if name in wb.sheetnames:
            d = read_platform_data(wb[name])
            if d:
                d["status"] = evaluate_status(d)
                d["fraud_risks"] = evaluate_fraud_risks(d)
                d["office"] = headcount.get(name, {}).get("office", 0)
                d["online"] = headcount.get(name, {}).get("online", 0)
                mm_data[name] = d

    wb.close()

    def sum_field(data, field):
        return int(sum(d[field] for d in data.values()))

    ph_reg = sum_field(ph_data, "register")
    ph_ftd = sum_field(ph_data, "ftd")
    ph_dps_ppl = sum_field(ph_data, "dps_ppl")
    ph_ftd_amt = sum(d["ftd_amount"] for d in ph_data.values())
    ph_dps_amt = sum(d["total_dps_amt"] for d in ph_data.values())
    ph_wdr_amt = sum(d["total_wdr_amt"] for d in ph_data.values())
    ph_diff = sum(d["diff"] for d in ph_data.values())
    ph_office = sum(d["office"] for d in ph_data.values())
    ph_online = sum(d["online"] for d in ph_data.values())

    bd_reg = sum_field(bd_data, "register")
    bd_ftd = sum_field(bd_data, "ftd")
    bd_dps_ppl = sum_field(bd_data, "dps_ppl")
    bd_ftd_amt = sum(d["ftd_amount"] for d in bd_data.values())
    bd_dps_amt = sum(d["total_dps_amt"] for d in bd_data.values())
    bd_wdr_amt = sum(d["total_wdr_amt"] for d in bd_data.values())
    bd_diff = sum(d["diff"] for d in bd_data.values())
    bd_office = sum(d["office"] for d in bd_data.values())
    bd_online = sum(d["online"] for d in bd_data.values())

    mm_reg = sum_field(mm_data, "register")
    mm_ftd = sum_field(mm_data, "ftd")
    mm_dps_ppl = sum_field(mm_data, "dps_ppl")
    mm_ftd_amt = sum(d["ftd_amount"] for d in mm_data.values())
    mm_dps_amt = sum(d["total_dps_amt"] for d in mm_data.values())
    mm_wdr_amt = sum(d["total_wdr_amt"] for d in mm_data.values())
    mm_diff = sum(d["diff"] for d in mm_data.values())
    mm_office = sum(d["office"] for d in mm_data.values())
    mm_online = sum(d["online"] for d in mm_data.values())

    total_reg = ph_reg + bd_reg + mm_reg
    total_ftd = ph_ftd + bd_ftd + mm_ftd
    total_dps_ppl = ph_dps_ppl + bd_dps_ppl + mm_dps_ppl
    total_ftd_amt = ph_ftd_amt + bd_ftd_amt + mm_ftd_amt
    total_dps_amt = ph_dps_amt + bd_dps_amt + mm_dps_amt
    total_wdr_amt = ph_wdr_amt + bd_wdr_amt + mm_wdr_amt
    total_diff = ph_diff + bd_diff + mm_diff
    total_office = ph_office + bd_office + mm_office
    total_online = ph_online + bd_online + mm_online

    dates = [d["date"] for d in list(ph_data.values()) + list(bd_data.values()) + list(mm_data.values()) if d["date"]]
    latest_date = "Unknown"
    if dates:
        d0 = dates[0]
        if hasattr(d0, 'strftime'):
            latest_date = d0.strftime("%Y-%m-%d")
        else:
            latest_date = str(d0)[:10]

    # --- Build PH table ---
    ph_rows = ""
    for name in PH_PLATFORMS:
        if name in ph_data:
            d = ph_data[name]
            icon = {"critical": "RED", "warning": "YEL"}.get(d["status"], "GRN")
            roi_str = f"{d['roi']:.1f}" if d['roi'] != 0 else "N/A"
            ph_rows += (
                f"| {name} | {d['office']}+{d['online']}={d['office']+d['online']} | "
                f"{d['register']} | {d['ftd']} | {d['dps_ppl']} | "
                f"{fmt_k(d['ftd_amount'])} | {fmt_k(d['total_dps_amt'])} | "
                f"{fmt_k(d['total_wdr_amt'])} | {fmt_k(d['diff'])} | "
                f"{roi_str} | {icon} |\n"
            )
        else:
            ph_rows += f"| {name} | - | - | - | - | - | - | - | - | - | N/A |\n"

    # --- Build BD table ---
    bd_rows = ""
    for name in BD_PLATFORMS:
        if name in bd_data:
            d = bd_data[name]
            icon = {"critical": "RED", "warning": "YEL"}.get(d["status"], "GRN")
            roi_str = f"{d['roi']:.1f}" if d['roi'] != 0 else "N/A"
            bd_rows += (
                f"| {name} | {d['office']}+{d['online']}={d['office']+d['online']} | "
                f"{d['register']} | {d['ftd']} | {d['dps_ppl']} | "
                f"{fmt_k(d['ftd_amount'])} | {fmt_k(d['total_dps_amt'])} | "
                f"{fmt_k(d['total_wdr_amt'])} | {fmt_k(d['diff'])} | "
                f"{roi_str} | {icon} |\n"
            )
        else:
            bd_rows += f"| {name} | - | - | - | - | - | - | - | - | - | N/A |\n"

    # --- Build MM table ---
    mm_rows = ""
    for name in MM_PLATFORMS:
        if name in mm_data:
            d = mm_data[name]
            icon = {"critical": "RED", "warning": "YEL"}.get(d["status"], "GRN")
            roi_str = f"{d['roi']:.1f}" if d['roi'] != 0 else "N/A"
            mm_rows += (
                f"| {name} | {d['office']}+{d['online']}={d['office']+d['online']} | "
                f"{d['register']} | {d['ftd']} | {d['dps_ppl']} | "
                f"{fmt_k(d['ftd_amount'])} | {fmt_k(d['total_dps_amt'])} | "
                f"{fmt_k(d['total_wdr_amt'])} | {fmt_k(d['diff'])} | "
                f"{roi_str} | {icon} |\n"
            )
        else:
            mm_rows += f"| {name} | - | - | - | - | - | - | - | - | - | N/A |\n"

    # --- Assemble content ---
    content = f"""# Online Field Team - Real-time Push
**Push Time**: {timestamp}
**Data Date**: {latest_date}
**Next Push**: 1 hour

---

## Philippines Sites
| Site | Staff(Onsite+Online) | Reg | FTD | Deposits | FTD Amt | Total Deposits | Total W/D | Net Deposit | ROI | Status |
|------|----------------------|-----|-----|----------|---------|----------------|-----------|-------------|-----|--------|
{ph_rows}
> **PH Total** | Onsite {ph_office}+Online {ph_online}={ph_office+ph_online} ppl | Reg {ph_reg} | FTD {ph_ftd} | Deposits {ph_dps_ppl} | FTD {fmt_k(ph_ftd_amt)} | Deposits {fmt_k(ph_dps_amt)} | W/D {fmt_k(ph_wdr_amt)} | Net {fmt_k(ph_diff)}

## Bangladesh Sites
| Site | Staff(Onsite+Online) | Reg | FTD | Deposits | FTD Amt | Total Deposits | Total W/D | Net Deposit | ROI | Status |
|------|----------------------|-----|-----|----------|---------|----------------|-----------|-------------|-----|--------|
{bd_rows}
> **BD Total** | Onsite {bd_office}+Online {bd_online}={bd_office+bd_online} ppl | Reg {bd_reg} | FTD {bd_ftd} | Deposits {bd_dps_ppl} | FTD {fmt_k(bd_ftd_amt)} | Deposits {fmt_k(bd_dps_amt)} | W/D {fmt_k(bd_wdr_amt)} | Net {fmt_k(bd_diff)}

## Myanmar Sites
| Site | Staff(Onsite+Online) | Reg | FTD | Deposits | FTD Amt | Total Deposits | Total W/D | Net Deposit | ROI | Status |
|------|----------------------|-----|-----|----------|---------|----------------|-----------|-------------|-----|--------|
{mm_rows}
> **MM Total** | Onsite {mm_office}+Online {mm_online}={mm_office+mm_online} ppl | Reg {mm_reg} | FTD {mm_ftd} | Deposits {mm_dps_ppl} | FTD {fmt_k(mm_ftd_amt)} | Deposits {fmt_k(mm_dps_amt)} | W/D {fmt_k(mm_wdr_amt)} | Net {fmt_k(mm_diff)}

## Overall Summary
| Metric | Philippines | Bangladesh | Myanmar | Total |
|--------|-------------|------------|---------|-------|
| Onsite Staff | {ph_office} | {bd_office} | {mm_office} | {total_office} |
| Online Staff | {ph_online} | {bd_online} | {mm_online} | {total_online} |
| Total Reg | {ph_reg} | {bd_reg} | {mm_reg} | {total_reg} |
| FTD Count | {ph_ftd} | {bd_ftd} | {mm_ftd} | {total_ftd} |
| Deposit Count | {ph_dps_ppl} | {bd_dps_ppl} | {mm_dps_ppl} | {total_dps_ppl} |
| FTD Amount | {fmt_k(ph_ftd_amt)} | {fmt_k(bd_ftd_amt)} | {fmt_k(mm_ftd_amt)} | {fmt_k(total_ftd_amt)} |
| Total Deposit Amt | {fmt_k(ph_dps_amt)} | {fmt_k(bd_dps_amt)} | {fmt_k(mm_dps_amt)} | {fmt_k(total_dps_amt)} |
| Total W/D Amt | {fmt_k(ph_wdr_amt)} | {fmt_k(bd_wdr_amt)} | {fmt_k(mm_wdr_amt)} | {fmt_k(total_wdr_amt)} |
| Net Deposit | {fmt_k(ph_diff)} | {fmt_k(bd_diff)} | {fmt_k(mm_diff)} | {fmt_k(total_diff)} |

## Reminder
> {tip}
"""

    # --- Anomalies ---
    anomalies = []
    for name in PH_PLATFORMS:
        if name in ph_data and ph_data[name]["status"] in ("critical", "warning"):
            anomalies.append((name, ph_data[name]))
    for name in BD_PLATFORMS:
        if name in bd_data and bd_data[name]["status"] in ("critical", "warning"):
            anomalies.append((name, bd_data[name]))
    for name in MM_PLATFORMS:
        if name in mm_data and mm_data[name]["status"] in ("critical", "warning"):
            anomalies.append((name, mm_data[name]))

    if anomalies:
        content += "\n## Anomaly Sites\n"
        for name, d in anomalies:
            icon = "RED" if d["status"] == "critical" else "YEL"
            reasons = []
            if d["ftd"] == 0:
                reasons.append("FTD=0")
            if d["wdr_ppl"] > d["dps_ppl"]:
                reasons.append(f"W/D>Deposits(DIFF={d['diff']:,.0f})")
            if 0 < d["ftd"] < 10:
                reasons.append("FTD single-digit")
            if d["roi"] < 0:
                reasons.append(f"ROI negative({d['roi']:.1f})")
            content += f"- {icon} **{name}**: FTD={d['ftd']}, Deposits={d['dps_ppl']:,}, W/D={d['wdr_ppl']:,} ({', '.join(reasons)})\n"
    else:
        content += "\n## Anomaly Sites\n- No anomalies\n"

    # --- Fraud risks ---
    fraud_items = []
    for name in PH_PLATFORMS:
        if name in ph_data and ph_data[name].get("fraud_risks"):
            for risk in ph_data[name]["fraud_risks"]:
                fraud_items.append((name, risk))
    for name in BD_PLATFORMS:
        if name in bd_data and bd_data[name].get("fraud_risks"):
            for risk in bd_data[name]["fraud_risks"]:
                fraud_items.append((name, risk))
    for name in MM_PLATFORMS:
        if name in mm_data and mm_data[name].get("fraud_risks"):
            for risk in mm_data[name]["fraud_risks"]:
                fraud_items.append((name, risk))

    if fraud_items:
        content += "\n## Fraud Risk Alerts\n"
        for name, risk in fraud_items:
            content += f"- ALERT **{name}**: {risk}\n"
    else:
        content += "\n## Fraud Risk Alerts\n- All clear\n"

    content += f"\n---\n*Auto push | {timestamp} | Next push: 1 hour*"

    # Write push file
    push_file = os.path.join(TARGET_DIR, "LATEST_PUSH_EN.md")
    with open(push_file, "w", encoding="utf-8") as f:
        f.write(content)

    print(f"[{timestamp}] EN Push generated: {push_file}")


if __name__ == "__main__":
    generate_push()
